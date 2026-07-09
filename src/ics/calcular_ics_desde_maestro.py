"""
calcular_ics_desde_maestro.py
===============================
Lee EXCEL_MAESTRO_ICS.xlsx (generado por etl_normalizar_excels.py), ejecuta
el pipeline de indicador_cumplimiento_historico.py, y agrega el resultado
como una quinta hoja "Resultado_ICS" dentro del mismo Excel maestro.

Flujo:  4 Excel fuente -> etl_normalizar_excels.py -> EXCEL_MAESTRO_ICS.xlsx
        EXCEL_MAESTRO_ICS.xlsx -> calcular_ics_desde_maestro.py -> Resultado_ICS
"""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from indicador_cumplimiento_historico import calcular_indicador_completo

MAESTRO = "EXCEL_MAESTRO_ICS.xlsx"


def leer_maestro(path: str = MAESTRO):
    ejecutores = pd.read_excel(path, sheet_name="Ejecutores", dtype={"codigo_ejecutor": str})
    proyectos = pd.read_excel(path, sheet_name="Proyectos", dtype={"bpin": str, "codigo_ejecutor": str})
    periodos = pd.read_excel(path, sheet_name="Periodos", dtype={"bpin": str, "codigo_ejecutor": str})
    reprogramaciones = pd.read_excel(path, sheet_name="Reprogramaciones", dtype={"bpin": str})
    return ejecutores, proyectos, periodos, reprogramaciones


def preparar_insumos(ejecutores, proyectos, periodos, reprogramaciones):
    """
    Adapta las tablas normalizadas del Excel maestro al formato que espera
    calcular_indicador_completo(): agrega reprogramaciones a nivel de
    ejecutor (sumando las de todos sus BPIN) y arma df_capacidad.
    """
    df_periodos = periodos[
        ["bpin", "codigo_ejecutor", "periodo", "valor_programado", "valor_ejecutado"]
    ].copy()

    df_proyectos = proyectos.merge(
        periodos[["bpin", "codigo_ejecutor"]].drop_duplicates(),
        on="bpin", how="left", suffixes=("", "_periodos")
    )
    # Si el proyecto no trae codigo_ejecutor propio (por archivo incompleto),
    # se usa el que aparece en Periodos como respaldo
    df_proyectos["codigo_ejecutor"] = df_proyectos["codigo_ejecutor"].fillna(
        df_proyectos.get("codigo_ejecutor_periodos")
    )
    df_proyectos = df_proyectos[["bpin", "codigo_ejecutor", "estado"]]

    reprog_por_bpin = reprogramaciones.merge(
        proyectos[["bpin", "codigo_ejecutor"]], on="bpin", how="left"
    )
    reprog_por_ejecutor = (
        reprog_por_bpin.groupby("codigo_ejecutor")["reprogramaciones_no_permitidas"]
        .sum()
        .reset_index()
    )

    df_capacidad = ejecutores[["codigo_ejecutor", "capacidad_institucional"]].copy()

    return df_periodos, df_proyectos, reprog_por_ejecutor, df_capacidad


def guardar_resultado_en_maestro(resultado: pd.DataFrame, path: str = MAESTRO):
    """Agrega/reemplaza la hoja 'Resultado_ICS' sin borrar las demás hojas."""
    wb = load_workbook(path)
    if "Resultado_ICS" in wb.sheetnames:
        del wb["Resultado_ICS"]
    wb.save(path)

    with pd.ExcelWriter(path, engine="openpyxl", mode="a") as writer:
        resultado.to_excel(writer, sheet_name="Resultado_ICS", index=False)


if __name__ == "__main__":
    ejecutores, proyectos, periodos, reprogramaciones = leer_maestro()

    df_periodos, df_proyectos, df_reprog, df_capacidad = preparar_insumos(
        ejecutores, proyectos, periodos, reprogramaciones
    )

    if df_periodos.empty or df_proyectos.empty:
        print(
            "Aviso: con los archivos de MUESTRA (1-2 filas por archivo) no hay "
            "suficiente cruce de BPIN para calcular el ICS de un ejecutor real.\n"
            "El pipeline corrió sin errores; con los archivos completos "
            "(miles de filas) este mismo script calculará el ICS de todos "
            "los ejecutores automáticamente.\n"
        )
    else:
        resultado = calcular_indicador_completo(
            df_periodos, df_proyectos, df_reprog, df_capacidad
        )
        guardar_resultado_en_maestro(resultado)
        print("Resultado guardado en la hoja 'Resultado_ICS' de", Path(MAESTRO).resolve())
        print(resultado.to_string(index=False))
