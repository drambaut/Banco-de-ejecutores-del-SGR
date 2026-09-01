"""
construir_maestro.py
=====================
Construye el Excel Maestro (EXCEL_MAESTRO.xlsx) a partir de 6 Excel
fuente:

    1) Balance_seguimiento_SGR.xlsx        -> Ejecutores + Proyectos
    2) Curva_Sl_*.xlsx                     -> Periodos (avance mes a mes)
    3) Reprogramaciones_no_permitidas.xlsx -> Reprogramaciones
    4) 5. Ajustes <mes> <año>.xlsx         -> Ajustes (detalle de ajustes
                                               al valor de los proyectos,
                                               hoja "BASE CON AJUSTES")
    5) *Universo*continuidad*.xlsx         -> Continuidad (períodos
                                               obligados vs. reportados
                                               por BPIN por corte, para
                                               el ICCI)
    6) *Consolidado*habilitaci*.xlsx       -> Habilitaciones (solicitudes
                                               de reapertura de períodos
                                               por BPIN, para el ICCI)

Hojas de salida en el maestro:

    Ejecutores        -> 1 fila por codigo_ejecutor. Incluye
                         fecha_inicio_ejecutor: la fecha MÍNIMA de
                         'INICIO DE OBLIGACIÓN COMO EJECUTOR' entre TODOS
                         los proyectos de ese ejecutor (sin filtrar por
                         estado — ver nota abajo).
                         OJO: esta fecha es la obligación del ejecutor
                         sobre el proyecto, NO la fecha de programación
                         del proyecto (fecha_inicial_programacion), que
                         es otra columna distinta del mismo Balance y
                         puede no coincidir si hubo cambio de ejecutor.
    Proyectos         -> 1 fila por BPIN, con TODOS los estados que trae
                         el Balance (Terminado, En Ejecución, Sin
                         Contratar, Sin Migrar, y cualquier otro que
                         aparezca), SIN FILTRAR. Por decisión explícita
                         del usuario: este maestro ya no decide qué
                         estados son "válidos" — esa decisión la toma
                         cada script consumidor (ver ESTADOS_VALIDOS_ICH
                         en calcular_ICH.py/metodologia_ICH.py y
                         ESTADOS_VALIDOS_IE en calcular_IE.py/
                         metodologia_IE.py), como hiperparámetro
                         editable en cada uno.
                         Incluye también valor_total_inicial y
                         fecha_aprobacion (usadas por calcular_ima.py).
    Ajustes           -> 1 fila por BPIN, viene de "5. Ajustes <mes>
                         <año>.xlsx" (hoja "BASE CON AJUSTES"), NO de
                         Balance. Trae ¿EL PROYECTO PRESENTA AJUSTES?,
                         conteos y valores de ajuste por alza/disminución
                         al SGR, ajustes a otras fuentes, % de ajuste, y
                         hasta 4 ajustes individuales con fecha y valor
                         propios (1er a 4to ajuste SGR) — esto sí trae la
                         fecha del PRIMER ajuste, a diferencia de Balance.
                         Se valida por BPIN contra Proyectos (alerta de
                         cobertura, sin filtrar nada).
    Periodos          -> 1 fila por BPIN + periodo (Curva Sl)
    Reprogramaciones  -> 1 fila por BPIN, con la resta
                         (realizadas - permitidas) calculada por este
                         script, comparada contra la columna que ya trae
                         el archivo fuente
    Continuidad       -> 1 fila por BPIN + fecha_corte. Viene del excel
                         de Universo reportes continuidad (hoja
                         UNIVERSO_REPORTADO_NATALIA). Trae
                         nro_periodos_a_reportar, nro_periodos_reportados,
                         fecha_migracion_proyecto, estado_detalle y
                         fecha_corte. Insumo del Score Continuidad del
                         ICCI. Al acumular varios cortes mensuales, cada
                         BPIN tendrá una fila por corte.
    Habilitaciones    -> 1 fila por solicitud de habilitación de período.
                         Viene del excel Consolidado habilitación proyectos
                         (hoja LISTA HABILITACIÓN). Trae bpin,
                         fecha_radicado, periodos_a_habilitar y
                         fecha_habilitacion. Insumo del Score Calidad de
                         Información del ICCI.

USO
----
    python construir_maestro.py --path_files ruta/carpeta
    python construir_maestro.py --path_files ruta/carpeta --verbose
    python construir_maestro.py --path_files ruta/carpeta --verbose --log
    python construir_maestro.py --path_files ruta/carpeta --output otro_nombre.xlsx

Los 4 excels fuente deben estar dentro de --path_files. El script los
identifica por patrón de nombre (ver PATRON_ARCHIVO_* abajo), así que no
importa si el nombre exacto trae fechas o corte distinto (p.ej.
"Curva Sl 23062026.xlsx" o "Curva Sl 30072026.xlsx" ambos calzan).

CONFIGURACIÓN GLOBAL
----------------------
Todo lo que puede variar entre entregas (nombre de archivo por patrón,
nombre de hoja, fila de encabezado) está centralizado en las constantes
de la sección siguiente. Si el equipo de datos cambia el nombre de una
hoja, se ajusta AQUÍ, no en la lógica de abajo.
"""

from __future__ import annotations

import argparse
import logging
import re
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# =============================================================================
# CONFIGURACIÓN GLOBAL — AJUSTAR AQUÍ SI CAMBIAN NOMBRES DE ARCHIVO / HOJA
# =============================================================================

# Patrones (glob simplificado, case-insensitive) para ubicar cada excel
# fuente dentro de --path_files.
PATRON_ARCHIVO_BALANCE = "*Balance*seguimiento*SGR*.xlsx"
PATRON_ARCHIVO_CURVA = "*Curva*S*.xlsx"
PATRON_ARCHIVO_REPROGRAMACIONES = "*Reprogramacion*"
PATRON_ARCHIVO_AJUSTES = "*Ajustes*.xlsx"
PATRON_ARCHIVO_CONTINUIDAD = "*Universo*continuidad*.xlsx"
PATRON_ARCHIVO_HABILITACIONES = "*habilitaci*.xlsx"

# Nombres de hoja esperados. No hace falta que calcen carácter por carácter
# (espacios extra, mayúsculas/minúsculas): se buscan por coincidencia
# flexible contra las hojas reales del archivo (ver _buscar_hoja).
HOJA_BALANCE = "PROYECTOS APROBADOS"
HOJA_CURVA = "Curva S"
HOJA_REPROGRAMACIONES = "BASE FINAL"
HOJA_AJUSTES = "BASE CON AJUSTES"
HOJA_CONTINUIDAD = "UNIVERSO_REPORTADO_NATALIA"
HOJA_HABILITACIONES = "LISTA HABILITACIÓN"

# Fila (0-indexed, como espera pandas header=) donde está el encabezado
# real en la hoja de Balance (las primeras filas son título/notas).
FILA_ENCABEZADO_BALANCE = 7

# Columna del Balance con la fecha de inicio de obligación del ejecutor
# sobre CADA proyecto. fecha_inicio_ejecutor (hoja Ejecutores) es el
# mínimo de esta columna entre los proyectos filtrados de cada ejecutor.
COLUMNA_INICIO_OBLIGACION_EJECUTOR = "INICIO DE OBLIGACIÓN COMO EJECUTOR"

NOMBRE_SALIDA_DEFAULT = "EXCEL_MAESTRO.xlsx"

# Ruta de salida por defecto: siempre dentro de outputs/ relativo al script,
# independientemente de dónde esté --path_files.
# Se puede sobreescribir con --output en la CLI.
RUTA_SALIDA_DEFAULT = Path(__file__).parent / "exceles" / "outputs" / NOMBRE_SALIDA_DEFAULT


logger = logging.getLogger("construir_maestro")


# =============================================================================
# UTILIDADES DE BÚSQUEDA FLEXIBLE (archivos, hojas, columnas)
# =============================================================================

def _normalizar_texto(s: str) -> str:
    s = str(s).lower().strip()
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    s = s.replace('"', "").replace("'", "")
    return " ".join(s.split())


def _detectar_fila_encabezado(path: Path, hoja: str, columna_ancla: str = "BPIN", max_filas: int = 15) -> int:
    """Escanea las primeras `max_filas` filas de la hoja buscando aquella
    que contiene una celda igual a `columna_ancla` (tolerando espacios,
    tildes y mayúsculas). Devuelve el índice 0-based listo para usarse como
    header= de pandas.read_excel. Útil para archivos con encabezados de
    varios niveles (ej. Ajustes), donde la fila exacta puede variar entre
    cortes mensuales."""
    vista = pd.read_excel(path, sheet_name=hoja, header=None, nrows=max_filas)
    ancla_norm = _normalizar_texto(columna_ancla)
    for i in range(len(vista)):
        fila_norm = vista.iloc[i].map(_normalizar_texto)
        if (fila_norm == ancla_norm).any():
            return i
    raise ValueError(
        f"No se encontró una fila de encabezado con '{columna_ancla}' en las "
        f"primeras {max_filas} filas de la hoja '{hoja}' de {path.name}. "
        f"Ajusta max_filas o columna_ancla si el archivo cambió de formato."
    )


def _resolver_archivo(path_files: Path, patron: str, descripcion: str) -> Path:
    """Busca un único archivo dentro de path_files que calce con el patrón
    (case-insensitive, tolera acentos). Falla con mensaje claro si no
    encuentra exactamente uno."""
    candidatos = [
        p for p in path_files.iterdir()
        if p.is_file()
        and p.suffix.lower() in (".xlsx", ".xls")
        and not p.name.startswith("~$")  # excluir archivos temporales de Excel
    ]
    fragmentos = [_normalizar_texto(f) for f in patron.split("*") if f.strip()]

    def calza(nombre: str) -> bool:
        nombre_norm = _normalizar_texto(nombre)
        return all(frag in nombre_norm for frag in fragmentos)

    encontrados = [p for p in candidatos if calza(p.name)]

    if not encontrados:
        disponibles = ", ".join(p.name for p in candidatos) or "(ninguno)"
        raise FileNotFoundError(
            f"No se encontró el archivo de '{descripcion}' en {path_files} "
            f"(se buscó algo que contenga: {fragmentos}). "
            f"Archivos disponibles en la carpeta: {disponibles}"
        )
    if len(encontrados) > 1:
        nombres = ", ".join(p.name for p in encontrados)
        raise FileNotFoundError(
            f"Se encontró más de un archivo que calza con '{descripcion}': {nombres}. "
            f"Deja solo uno en la carpeta o ajusta el patrón PATRON_ARCHIVO_* en el script."
        )
    return encontrados[0]


def _buscar_hoja(path: Path, hoja_esperada: str) -> str:
    """Devuelve el nombre REAL de la hoja en el excel que mejor calza con
    hoja_esperada (tolera espacios extra, mayúsculas/minúsculas, tildes)."""
    hojas_reales = pd.ExcelFile(path).sheet_names
    esperado_norm = _normalizar_texto(hoja_esperada)

    for hoja in hojas_reales:
        if _normalizar_texto(hoja) == esperado_norm:
            return hoja
    for hoja in hojas_reales:
        if esperado_norm in _normalizar_texto(hoja) or _normalizar_texto(hoja) in esperado_norm:
            return hoja

    raise ValueError(
        f"No se encontró una hoja parecida a '{hoja_esperada}' en {path.name}. "
        f"Hojas disponibles: {hojas_reales}. Ajusta HOJA_* en el script si el "
        f"nombre real cambió."
    )


def _resolver_columnas(df: pd.DataFrame, columnas_esperadas: list[str], archivo: str) -> dict[str, str]:
    """Empareja cada nombre de columna esperado contra las columnas reales del
    DataFrame (tolerando espacios extra/mayúsculas). Devuelve
    {nombre_esperado: nombre_real}. Falla si falta alguna."""
    mapa_real = {_normalizar_texto(c): c for c in df.columns}
    resultado = {}
    faltantes = []
    for esperada in columnas_esperadas:
        clave = _normalizar_texto(esperada)
        if clave in mapa_real:
            resultado[esperada] = mapa_real[clave]
        else:
            faltantes.append(esperada)
    if faltantes:
        raise ValueError(
            f"Columnas esperadas no encontradas en {archivo}: {faltantes}. "
            f"Columnas reales disponibles: {list(df.columns)}"
        )
    return resultado


def _normalizar_estado(valor) -> str:
    if pd.isna(valor):
        return "Desconocido"
    v = str(valor).strip().upper()
    if "TERMIN" in v:
        return "Terminado"
    if "EJECUC" in v:
        return "En Ejecución"
    return str(valor).strip().title()


# =============================================================================
# 1. EJECUTORES + PROYECTOS  <-  Balance_seguimiento_SGR.xlsx
# =============================================================================

def cargar_proyectos_y_ejecutores(path_balance: Path):
    hoja = _buscar_hoja(path_balance, HOJA_BALANCE)
    df = pd.read_excel(path_balance, sheet_name=hoja, header=FILA_ENCABEZADO_BALANCE)

    columnas_proyecto = {
        "BPIN": "bpin",
        "NOMBRE DEL PROYECTO": "nombre_proyecto",
        "SECTOR": "sector",
        "ESTADO GENERAL": "estado",
        "TOTAL PROYECTO": "valor_total_proyecto",
        "CÓDIGO EJECUTOR": "codigo_ejecutor",
        "FECHA INICIAL DE LA PROGRAMACIÓN": "fecha_inicial_programacion",
        "FECHA FINAL DE LA PROGRAMACIÓN": "fecha_final_programacion",
        "VALOR TOTAL INICIAL": "valor_total_inicial",
        "FECHA APROBACIÓN": "fecha_aprobacion",
    }
    columnas_ejecutor = {
        "CÓDIGO EJECUTOR": "codigo_ejecutor",
        "ENTIDAD EJECUTORA": "nombre_ejecutor",
        "NIT ENTIDAD EJECUTORA": "nit",
        "DEPARTAMENTO LOCALIZACIÓN DEL EJECUTOR": "departamento",
        "REGIÓN LOCALIZACIÓN DEL EJECUTOR": "region",
        "TIPO EJECUTOR": "tipo_ejecutor",
        "CAPACIDAD INSTITUCIONAL": "capacidad_institucional",
    }

    todas_esperadas = sorted(
        set(columnas_proyecto) | set(columnas_ejecutor) | {COLUMNA_INICIO_OBLIGACION_EJECUTOR}
    )
    mapa_cols = _resolver_columnas(df, todas_esperadas, path_balance.name)
    df = df.rename(columns={v: k for k, v in mapa_cols.items()})

    df["ESTADO GENERAL"] = df["ESTADO GENERAL"].apply(_normalizar_estado)

    n_total = df["BPIN"].nunique()
    conteo_estados = df.drop_duplicates(subset="BPIN")["ESTADO GENERAL"].value_counts().to_dict()
    logger.info(
        "Balance: %s BPIN únicos en el archivo fuente, SIN FILTRAR por estado "
        "(decisión explícita del usuario — cada script consumidor decide qué "
        "estados usar). Desglose por estado: %s",
        n_total, conteo_estados,
    )

    df_proyectos = df[list(columnas_proyecto.keys())].rename(columns=columnas_proyecto)
    df_proyectos["bpin"] = df_proyectos["bpin"].astype(str).str.strip()
    df_proyectos["codigo_ejecutor"] = df_proyectos["codigo_ejecutor"].astype(str).str.strip()
    df_proyectos = df_proyectos.drop_duplicates(subset="bpin")

    df_ejecutores = df[list(columnas_ejecutor.keys())].rename(columns=columnas_ejecutor)
    df_ejecutores["codigo_ejecutor"] = df_ejecutores["codigo_ejecutor"].astype(str).str.strip()
    df_ejecutores = df_ejecutores.drop_duplicates(subset="codigo_ejecutor")

    # fecha_inicio_ejecutor = mínima fecha de "INICIO DE OBLIGACIÓN COMO
    # EJECUTOR" entre TODOS los proyectos de cada ejecutor (sin filtrar por
    # estado, igual que el resto de esta hoja).
    df["_codigo_ejecutor_str"] = df["CÓDIGO EJECUTOR"].astype(str).str.strip()
    df["_fecha_inicio_obligacion"] = pd.to_datetime(
        df[COLUMNA_INICIO_OBLIGACION_EJECUTOR], errors="coerce"
    )
    n_fechas_invalidas = df["_fecha_inicio_obligacion"].isna().sum()
    if n_fechas_invalidas:
        logger.warning(
            "ALERTA — %s de %s filas del Balance no tienen una fecha válida en "
            "'%s' (vacía o con formato no reconocido); no aportan al mínimo de "
            "fecha_inicio_ejecutor.",
            n_fechas_invalidas, len(df), COLUMNA_INICIO_OBLIGACION_EJECUTOR,
        )

    fecha_inicio_por_ejecutor = (
        df.groupby("_codigo_ejecutor_str")["_fecha_inicio_obligacion"]
        .min()
        .rename("fecha_inicio_ejecutor")
    )
    df_ejecutores = df_ejecutores.merge(
        fecha_inicio_por_ejecutor, left_on="codigo_ejecutor", right_index=True, how="left"
    )

    n_ejecutores_sin_fecha = df_ejecutores["fecha_inicio_ejecutor"].isna().sum()
    if n_ejecutores_sin_fecha:
        logger.warning(
            "ALERTA — %s de %s ejecutores quedaron sin fecha_inicio_ejecutor "
            "(ninguno de sus proyectos tenía '%s' válida).",
            n_ejecutores_sin_fecha, len(df_ejecutores), COLUMNA_INICIO_OBLIGACION_EJECUTOR,
        )

    return df_proyectos, df_ejecutores


# =============================================================================
# 2. PERIODOS  <-  Curva_Sl_*.xlsx
# =============================================================================

def cargar_periodos(path_curva: Path) -> pd.DataFrame:
    hoja = _buscar_hoja(path_curva, HOJA_CURVA)
    df = pd.read_excel(path_curva, sheet_name=hoja)

    columnas = {
        "CODIGO_EJECUTOR": "codigo_ejecutor",
        "BPIN": "bpin",
        "PERIODO": "periodo",
        "PERIODO_FECHA": "periodo_fecha",
        "PV_VALOR_MES": "valor_programado",
        "EV_VALOR_MES": "valor_ejecutado",
        "FECHA_CORTE": "fecha_corte",
    }
    mapa_cols = _resolver_columnas(df, list(columnas.keys()), path_curva.name)
    df = df[[mapa_cols[c] for c in columnas]].rename(
        columns={mapa_cols[c]: nuevo for c, nuevo in columnas.items()}
    )
    df["bpin"] = df["bpin"].astype(str).str.strip()
    df["codigo_ejecutor"] = df["codigo_ejecutor"].astype(str).str.strip()
    df = df.sort_values(["codigo_ejecutor", "bpin", "periodo"]).reset_index(drop=True)

    logger.info(
        "Curva S (%s): %s filas de periodo, %s BPIN únicos con avance",
        path_curva.name, len(df), df["bpin"].nunique(),
    )
    return df


# =============================================================================
# 3. REPROGRAMACIONES  <-  Reprogramaciones_no_permitidas.xlsx
# =============================================================================

def cargar_reprogramaciones(path: Path) -> pd.DataFrame:
    hoja = _buscar_hoja(path, HOJA_REPROGRAMACIONES)
    df = pd.read_excel(path, sheet_name=hoja)

    columnas = {
        "BPIN": "bpin",
        "TOTAL REPROGRAMACIONES PERMITIDAS AJUSTADA": "reprogramaciones_permitidas",
        "TOTAL REPROGRAMACIONES REALIZADAS": "reprogramaciones_realizadas",
        "TOTAL REPROGRAMACIONES NO PERMITIDAS": "reprogramaciones_no_permitidas_fuente",
        "TECHO DE MEDICIÓN": "techo_medicion",
    }
    mapa_cols = _resolver_columnas(df, list(columnas.keys()), path.name)
    df = df[[mapa_cols[c] for c in columnas]].rename(
        columns={mapa_cols[c]: nuevo for c, nuevo in columnas.items()}
    )
    df["bpin"] = df["bpin"].astype(str).str.strip()

    # La "no permitidas" se calcula aquí a partir de permitidas y realizadas
    # (no se confía ciegamente en la columna que ya trae el archivo fuente).
    # no_permitidas = realizadas - permitidas_ajustada, sin bajar de 0.
    diferencia = df["reprogramaciones_realizadas"] - df["reprogramaciones_permitidas"]
    df["reprogramaciones_no_permitidas"] = diferencia.clip(lower=0)

    # Alerta de calidad de dato: si el cálculo no coincide con lo que ya
    # traía el archivo fuente, se reporta (siempre, no solo en --verbose).
    discrepancias = df[
        df["reprogramaciones_no_permitidas"].fillna(0) != df["reprogramaciones_no_permitidas_fuente"].fillna(0)
    ]
    if not discrepancias.empty:
        ejemplos = discrepancias["bpin"].head(5).tolist()
        logger.warning(
            "ALERTA — %s de %s BPIN tienen (realizadas - permitidas) distinto de la "
            "columna 'TOTAL REPROGRAMACIONES NO PERMITIDAS' del archivo fuente. "
            "Ejemplos de BPIN con discrepancia: %s",
            len(discrepancias), len(df), ejemplos,
        )
    else:
        logger.info(
            "Reprogramaciones: el cálculo (realizadas - permitidas) coincide con la "
            "columna del archivo fuente en los %s BPIN.", len(df),
        )

    columnas_finales = [
        "bpin",
        "reprogramaciones_permitidas",
        "reprogramaciones_realizadas",
        "reprogramaciones_no_permitidas",
        "reprogramaciones_no_permitidas_fuente",
        "techo_medicion",
    ]
    return df[columnas_finales].drop_duplicates(subset="bpin")


# =============================================================================
# 4. AJUSTES  <-  "5. Ajustes <mes> <año>.xlsx" (hoja "BASE CON AJUSTES")
# =============================================================================

# Rango de años plausible para una fecha de ajuste real (el SGR arrancó en
# 2012). Última validación de sanidad, no el mecanismo principal de parseo.
AÑO_MIN_PLAUSIBLE = 2010
AÑO_MAX_PLAUSIBLE = 2035

MESES_ES = {
    "ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
    "jul": 7, "ago": 8, "sep": 9, "set": 9, "oct": 10, "nov": 11, "dic": 12,
}

# Formato real confirmado por el usuario para estas 4 columnas: la mayoría
# de las celdas son texto "<mes-abrev-español><año-2-dígitos>", ej. "feb18"
# = febrero de 2018 (SIN día exacto — el dato fuente no lo trae). Mayúsculas
# y minúsculas NO son significativas: "feb18" = "Feb18" = "FEB18" = "FEB18".
# Algunas celdas, en cambio, sí son fechas nativas de Excel completas (con
# día) — esas se usan tal cual, no se reinterpretan como texto.
# OJO: pd.to_datetime() sobre esta columna NO sirve — abreviaturas como
# "feb"/"mar"/"may"/"jun"/"jul"/"ago"/"sep"/"oct"/"nov" coinciden con
# abreviaturas en INGLÉS y pandas las parsea silenciosamente mal (año 1,
# sin ninguna alerta); "abr"/"ene"/"dic" no coinciden con nada en inglés y
# dan NaT directo. Por eso se parsea el patrón explícito ANTES de intentar
# cualquier función de fecha genérica.
_PATRON_MES_AÑO = re.compile(r"^\s*([A-Za-z]{3})[\s\-\./]*?(\d{1,4})\s*$")


def _parsear_texto_mes_año(texto: str):
    """Interpreta `texto` como '<mes-abrev><año>' (ej. 'feb18', 'Abr25',
    'FEB18', también tolera separadores como 'abr-2015'). Devuelve un
    Timestamp en el día 1 de ese mes/año (no hay día exacto en este
    formato), o None si el texto no calza con el patrón."""
    m = _PATRON_MES_AÑO.match(texto)
    if not m:
        return None
    mes_str, año_str = m.group(1).lower(), m.group(2)
    mes = MESES_ES.get(mes_str)
    if mes is None:
        return None
    if len(año_str) == 2:
        año = 2000 + int(año_str)
    elif len(año_str) == 4:
        año = int(año_str)
    else:
        return None
    if not (AÑO_MIN_PLAUSIBLE <= año <= AÑO_MAX_PLAUSIBLE):
        return None
    return pd.Timestamp(year=año, month=mes, day=1)


def _parsear_fecha_ajuste_serie(serie: pd.Series, nombre_col: str, bpins: pd.Series) -> pd.Series:
    """Convierte una serie de fechas de ajuste a datetime, manejando los 2
    formatos reales que trae esta columna (confirmado por el usuario
    revisando el archivo original en Excel):
      1) Fecha nativa de Excel (datetime/date de verdad, con día exacto) ->
         se usa tal cual, validando que el año sea plausible.
      2) Texto "<mes-abrev-español><año-2-dígitos>" (ej. 'feb18', sin
         distinguir mayúsculas/minúsculas) -> se interpreta como el día 1
         de ese mes/año. Se pierde el día exacto porque el dato fuente no
         lo trae; se reporta cuántas fechas quedaron en este caso.
    Cualquier valor que no calce con ninguno de los 2 formatos se descarta
    como NaT y se reporta con el valor original, en vez de adivinar (una
    corrección automática mal hecha sería peor que no tener el dato)."""
    resultado = []
    n_nativa = 0
    n_texto = 0
    descartadas = []

    for pos, valor in enumerate(serie):
        bpin = bpins.iloc[pos] if pos < len(bpins) else "?"

        if pd.isna(valor) or valor == "":
            resultado.append(pd.NaT)
            continue

        if isinstance(valor, (pd.Timestamp, datetime, date)):
            ts = pd.Timestamp(valor)
            if AÑO_MIN_PLAUSIBLE <= ts.year <= AÑO_MAX_PLAUSIBLE:
                resultado.append(ts)
            else:
                descartadas.append((bpin, valor, "fecha nativa de Excel con año implausible"))
                resultado.append(pd.NaT)
            n_nativa += 1
            continue

        texto = str(valor).strip()
        parseada = _parsear_texto_mes_año(texto)
        if parseada is not None:
            resultado.append(parseada)
            n_texto += 1
            continue

        # Último intento genérico, por si aparece alguna fecha completa
        # bien formada como texto (ej. '2023-08-01'); se descarta si el
        # año no da plausible en vez de aceptar un parseo ambiguo.
        intento = pd.to_datetime(texto, errors="coerce", dayfirst=True)
        if pd.notna(intento) and AÑO_MIN_PLAUSIBLE <= intento.year <= AÑO_MAX_PLAUSIBLE:
            resultado.append(intento)
        else:
            descartadas.append((bpin, valor, "formato no reconocido"))
            resultado.append(pd.NaT)

    if n_texto:
        logger.info(
            "%s: %s fechas venían en formato texto 'mesaño' (ej. 'feb18', sin "
            "distinguir mayúsculas/minúsculas) y se interpretaron como el día 1 "
            "del mes indicado (el dato fuente no trae día exacto). %s fechas "
            "venían como fecha nativa de Excel (con día exacto) y se usaron tal cual.",
            nombre_col, n_texto, n_nativa,
        )
    if descartadas:
        logger.warning(
            "ALERTA — %s: %s valores no se pudieron convertir a una fecha "
            "confiable y se descartan (NaT). Ejemplos (bpin, valor crudo, motivo): %s",
            nombre_col, len(descartadas), descartadas[:5],
        )

    return pd.Series(resultado, index=serie.index)


def cargar_ajustes(path: Path) -> pd.DataFrame:
    """Carga el detalle de ajustes al valor de los proyectos (columnas W a
    AK de la hoja 'BASE CON AJUSTES'): si presenta ajustes, conteos y
    valores por alza/disminución al SGR, ajustes a otras fuentes, % de
    ajuste, y hasta 4 ajustes individuales con su propia fecha y valor.
    A diferencia de Balance, este archivo SÍ trae la fecha del PRIMER
    ajuste (columna '1ER FECHA AJUSTE SGR') — en un formato de texto mixto
    que hay que resolver (ver _parsear_fecha_ajuste_serie)."""
    hoja = _buscar_hoja(path, HOJA_AJUSTES)
    fila_encabezado = _detectar_fila_encabezado(path, hoja, columna_ancla="BPIN")
    df = pd.read_excel(path, sheet_name=hoja, header=fila_encabezado)

    columnas = {
        "BPIN": "bpin",
        "¿EL PROYECTO PRESENTA AJUSTES?": "tiene_ajuste",
        "AJUSTES POR ALZA AL SGR": "n_ajustes_alza_sgr",
        "AJUSTES DISMINUCIÓN AL SGR": "n_ajustes_disminucion_sgr",
        "VALOR TOTAL AJUSTES SGR": "valor_total_ajustes_sgr",
        "AJUSTES OTRAS FUENTES": "ajustes_otras_fuentes",
        "TOTAL AJUSTE AL PROYECTO": "total_ajuste_proyecto",
        "PORCENTAJE DE AJUSTE": "porcentaje_ajuste",
        "1ER FECHA AJUSTE SGR": "fecha_ajuste_1",
        "1ER VALOR AJUSTE SGR": "valor_ajuste_1",
        "2DA FECHA AJUSTE SGR": "fecha_ajuste_2",
        "2DO VALOR AJUSTE SGR": "valor_ajuste_2",
        "3ER FECHA AJUSTE SGR": "fecha_ajuste_3",
        "3ER VALOR AJUSTE SGR": "valor_ajuste_3",
        "4TA FECHA AJUSTE SGR": "fecha_ajuste_4",
        "4TO VALOR AJUSTE SGR": "valor_ajuste_4",
    }
    mapa_cols = _resolver_columnas(df, list(columnas.keys()), path.name)
    df = df[[mapa_cols[c] for c in columnas]].rename(
        columns={mapa_cols[c]: nuevo for c, nuevo in columnas.items()}
    )
    df["bpin"] = df["bpin"].astype(str).str.strip()

    for col in ("fecha_ajuste_1", "fecha_ajuste_2", "fecha_ajuste_3", "fecha_ajuste_4"):
        df[col] = _parsear_fecha_ajuste_serie(df[col], col, df["bpin"])

    n_total = df["bpin"].nunique()
    conteo_tiene_ajuste = df.drop_duplicates(subset="bpin")["tiene_ajuste"].value_counts().to_dict()
    logger.info(
        "Ajustes (%s, hoja '%s', encabezado detectado en fila %s): %s BPIN únicos. "
        "Desglose de '¿EL PROYECTO PRESENTA AJUSTES?': %s",
        path.name, hoja, fila_encabezado + 1, n_total, conteo_tiene_ajuste,
    )

    return df.drop_duplicates(subset="bpin")


# =============================================================================
# 5. CONTINUIDAD  <-  *Universo*continuidad*.xlsx
# =============================================================================

def cargar_continuidad(path: Path) -> pd.DataFrame:
    """Carga el universo de períodos obligados vs. reportados por BPIN y
    corte de GESPROY. Es el insumo del Score Continuidad del ICCI.

    Al acumular varios cortes mensuales (un archivo por corte), la hoja
    Continuidad del Maestro tendrá una fila por BPIN × fecha_corte.
    Los scripts calculadores del ICCI deben filtrar por fecha_corte para
    trabajar con el corte deseado."""
    hoja = _buscar_hoja(path, HOJA_CONTINUIDAD)
    df = pd.read_excel(path, sheet_name=hoja)

    columnas = {
        "BPIN": "bpin",
        "ID_PROYECTO": "id_proyecto",
        "FECHA_MIGRACION_PROYECTO": "fecha_migracion_proyecto",
        "ESTADO_DETALLE": "estado_detalle",
        "FECHA_CORTE_GESPROY_PARA_ESTADO_Y_FECHAS": "fecha_corte",
        "NRO_PERIODOS_A_REPORTAR": "nro_periodos_a_reportar",
        "NRO_PERIODOS_REPORTADOS": "nro_periodos_reportados",
        "MARCA_REPORTO_MAS": "marca_reporto_mas",
        "DIFERENCIA_A_REPOTAR_VS_REPORTADO": "diferencia_a_reportar_vs_reportado",
    }
    mapa_cols = _resolver_columnas(df, list(columnas.keys()), path.name)
    df = df[[mapa_cols[c] for c in columnas]].rename(
        columns={mapa_cols[c]: nuevo for c, nuevo in columnas.items()}
    )
    df["bpin"] = df["bpin"].astype(str).str.strip()
    df["fecha_migracion_proyecto"] = pd.to_datetime(df["fecha_migracion_proyecto"], errors="coerce")
    df["fecha_corte"] = pd.to_datetime(df["fecha_corte"], errors="coerce")
    df = df.sort_values(["bpin", "fecha_corte"]).reset_index(drop=True)

    n_bpin = df["bpin"].nunique()
    cortes = df["fecha_corte"].dropna().unique()
    logger.info(
        "Continuidad (%s, hoja '%s'): %s filas, %s BPIN únicos, corte(s): %s",
        path.name, hoja, len(df), n_bpin,
        sorted(str(c)[:10] for c in cortes),
    )
    return df


# =============================================================================
# PARSEO DE PERÍODOS A HABILITAR
# =============================================================================

# Nombres de mes en español → número de mes
_MESES_ES_HAB = {
    "ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
    "jul": 7, "ago": 8, "sep": 9, "set": 9, "oct": 10, "nov": 11, "dic": 12,
}

def _parsear_n_periodos_a_habilitar(texto) -> int | None:
    """
    Convierte el texto libre de 'PERIODOS A HABILITAR' en un número entero
    de meses. Se estandariza aquí para que la hoja Habilitaciones del Maestro
    siempre tenga un valor numérico comparable entre filas.

    Casos que maneja:
      - Número puro ("6")                         → 6
      - Un mes ("ENERO 2026", "diciembre de 2025") → 1
      - Varios meses con Y/coma ("DIC 2025 Y ENE 2026")
                                                  → cuenta menciones de mes
      - Rango guionado ("ENERO - ABRIL 2026")     → 4 (inclusive)
      - "TODOS LOS PERIODOS"                      → None (indefinido)
    """
    if pd.isna(texto):
        return 0
    t = str(texto).upper().strip()

    # Caso especial: texto indefinido
    if "TODOS" in t and "PERIODO" in t:
        return None

    # Número puro
    if re.fullmatch(r"\d+", t):
        return int(t)

    # Rango explícito: "MES1 - MES2" o "MES1 – MES2"
    rango = re.search(r"([A-ZÁÉÍÓÚ]{3,})\s*[-–]\s*([A-ZÁÉÍÓÚ]{3,})", t)
    if rango:
        m1 = rango.group(1)[:3].lower()
        m2 = rango.group(2)[:3].lower()
        if m1 in _MESES_ES_HAB and m2 in _MESES_ES_HAB:
            i1, i2 = _MESES_ES_HAB[m1], _MESES_ES_HAB[m2]
            # Rango puede cruzar año (ej. DIC→FEB = 3 meses)
            return (i2 - i1 + 1) if i2 >= i1 else (12 - i1 + 1 + i2)

    # Contar menciones de nombre de mes (mínimo 3 letras)
    # Cada posición de inicio = un mes distinto (evita doble conteo)
    posiciones = {
        m.start()
        for m in re.finditer(r"\b([A-ZÁÉÍÓÚ]{3,})", t)
        if m.group(1)[:3].lower() in _MESES_ES_HAB
    }
    return max(len(posiciones), 1)


# =============================================================================
# FORMATO DE TABLAS EXCEL (openpyxl)
# =============================================================================

def _aplicar_formato_tabla(writer: "pd.ExcelWriter",
                           nombre_hoja: str,
                           nombre_tabla: str,
                           estilo: str = "TableStyleMedium2",
                           anchos_col: list[int] | None = None) -> None:
    """
    Convierte el rango de datos de una hoja en una tabla Excel con filtros,
    encabezado navy DNP y filas alternas.

    Para hojas con muchas filas (ej. Periodos con 500k+) el formateo
    celda-a-celda es inviable. En esos casos solo se aplica:
      - La definición de Tabla (filtros + estilo de franja) — solo metadata XML
      - El estilo del encabezado (fila 1 únicamente)
      - Anchos de columna derivados del texto del encabezado

    El estilo visual de las celdas de datos (fuente, borde, alineación) lo
    hereda Excel del TableStyle, sin tocar cada celda desde Python.
    """
    ws = writer.sheets[nombre_hoja]
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < 2 or max_col < 1:
        return

    col_fin = get_column_letter(max_col)
    ref = f"A1:{col_fin}{max_row}"

    # ── Tabla Excel (solo metadata, no itera celdas) ─────────────────────────
    tab = Table(displayName=nombre_tabla, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name=estilo,
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tab)

    # ── Encabezado: fila 1 únicamente ────────────────────────────────────────
    NAVY = "002060"
    borde = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF", size=10, name="Arial")
        cell.fill      = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = borde
    ws.row_dimensions[1].height = 28

    # ── Anchos de columna ────────────────────────────────────────────────────
    if anchos_col:
        # Anchos explícitos (rápido, sin leer datos)
        for i, w in enumerate(anchos_col, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    else:
        # Derivar del texto del encabezado + muestra de 200 filas de datos
        MUESTRA = min(200, max_row)
        for col in ws.iter_cols(min_row=1, max_row=MUESTRA, max_col=max_col):
            max_len = max(
                (len(str(c.value)) if c.value is not None else 0) for c in col
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 52)

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


# =============================================================================
# 6. HABILITACIONES  <-  *Consolidado*habilitaci*.xlsx
# =============================================================================

def cargar_habilitaciones(path: Path) -> pd.DataFrame:
    """Carga el consolidado de solicitudes de habilitación (reapertura de
    períodos ya cerrados en GESPROY). Es el insumo del Score Calidad de
    Información del ICCI.

    Cada fila es una solicitud: un BPIN puede tener varias (una por
    radicado). La columna periodos_a_habilitar viene como texto del archivo
    fuente (ej. 'DICIEMBRE 2025 Y ENERO 2026', '6', 'ENERO 2026'); el
    conteo de períodos lo debe hacer el script calculador del ICCI."""
    hoja = _buscar_hoja(path, HOJA_HABILITACIONES)
    df = pd.read_excel(path, sheet_name=hoja)

    columnas = {
        "BPIN": "bpin",
        "ESTADO DEL PROYECTO": "estado_proyecto",
        "FECHA DEL RADICADO": "fecha_radicado",
        "PERIODOS A HABILITAR": "periodos_a_habilitar",
        "FECHA DE HABILITACIÓN": "fecha_habilitacion",
    }
    mapa_cols = _resolver_columnas(df, list(columnas.keys()), path.name)
    df = df[[mapa_cols[c] for c in columnas]].rename(
        columns={mapa_cols[c]: nuevo for c, nuevo in columnas.items()}
    )
    df["bpin"] = df["bpin"].astype(str).str.strip()
    df["fecha_radicado"] = pd.to_datetime(df["fecha_radicado"], errors="coerce")
    df["fecha_habilitacion"] = pd.to_datetime(df["fecha_habilitacion"], errors="coerce")
    df["periodos_a_habilitar"] = df["periodos_a_habilitar"].astype(str).str.strip()

    # n_periodos_a_habilitar: conteo estandarizado de meses por solicitud.
    # El texto original viene en formatos heterogéneos ("DICIEMBRE 2025 Y
    # ENERO 2026", "6", "ENERO - ABRIL 2026", etc.).  Esta columna normaliza
    # a un entero para que los scripts consumidores puedan sumar directamente
    # sin necesidad de parsear texto.  Las filas con "TODOS LOS PERIODOS"
    # quedan como NaN (se reportan abajo).
    df["n_periodos_a_habilitar"] = df["periodos_a_habilitar"].apply(
        _parsear_n_periodos_a_habilitar
    )
    n_indefinidos = df["n_periodos_a_habilitar"].isna().sum()
    if n_indefinidos:
        ejemplos = df.loc[df["n_periodos_a_habilitar"].isna(), "periodos_a_habilitar"].head(3).tolist()
        logger.warning(
            "ALERTA — %s solicitudes de habilitación tienen texto indefinido "
            "('TODOS LOS PERIODOS' u otro no parseado) y quedan con "
            "n_periodos_a_habilitar = NaN. Ejemplos: %s",
            n_indefinidos, ejemplos,
        )

    df = df.sort_values(["bpin", "fecha_radicado"]).reset_index(drop=True)

    n_bpin = df["bpin"].nunique()
    logger.info(
        "Habilitaciones (%s, hoja '%s'): %s solicitudes, %s BPIN únicos",
        path.name, hoja, len(df), n_bpin,
    )
    # Reordenar columnas: texto original primero, conteo parseado al lado
    cols_finales = [
        "bpin", "estado_proyecto", "fecha_radicado",
        "periodos_a_habilitar", "n_periodos_a_habilitar", "fecha_habilitacion",
    ]
    return df[cols_finales]


def validar_cobertura_ajustes(df_proyectos: pd.DataFrame, df_ajustes: pd.DataFrame) -> None:
    """'Join' de validación por BPIN entre Proyectos (Balance) y Ajustes: no
    filtra nada en el maestro, solo reporta discrepancias de cobertura,
    igual que ya se hace con Reprogramaciones."""
    bpins_proyectos = set(df_proyectos["bpin"])
    bpins_ajustes = set(df_ajustes["bpin"])

    solo_en_ajustes = bpins_ajustes - bpins_proyectos
    solo_en_proyectos = bpins_proyectos - bpins_ajustes

    if solo_en_ajustes:
        logger.warning(
            "ALERTA — %s BPIN están en Ajustes pero NO en Proyectos (Balance). "
            "Ejemplos: %s",
            len(solo_en_ajustes), list(solo_en_ajustes)[:5],
        )
    if solo_en_proyectos:
        logger.warning(
            "ALERTA — %s BPIN están en Proyectos (Balance) pero NO en Ajustes. "
            "Se asume que no tienen ajuste (los scripts consumidores deben "
            "tratar su ausencia en 'Ajustes' como 'sin ajuste', no como dato "
            "faltante). Ejemplos: %s",
            len(solo_en_proyectos), list(solo_en_proyectos)[:5],
        )
    if not solo_en_ajustes and not solo_en_proyectos:
        logger.info("Ajustes: cobertura de BPIN coincide exactamente con Proyectos (Balance).")


# =============================================================================
# PIPELINE COMPLETO -> EXCEL MAESTRO
# =============================================================================

def construir_excel_maestro(path_files: Path, salida: Path) -> dict[str, pd.DataFrame]:
    path_balance = _resolver_archivo(path_files, PATRON_ARCHIVO_BALANCE, "Balance seguimiento SGR")
    path_curva = _resolver_archivo(path_files, PATRON_ARCHIVO_CURVA, "Curva S / Curva Sl")
    path_reprog = _resolver_archivo(path_files, PATRON_ARCHIVO_REPROGRAMACIONES, "Reprogramaciones no permitidas")
    path_ajustes = _resolver_archivo(path_files, PATRON_ARCHIVO_AJUSTES, "Ajustes")
    path_continuidad = _resolver_archivo(path_files, PATRON_ARCHIVO_CONTINUIDAD, "Universo reportes continuidad (ICCI)")
    path_habilitaciones = _resolver_archivo(path_files, PATRON_ARCHIVO_HABILITACIONES, "Consolidado habilitación proyectos (ICCI)")

    logger.info("Balance seguimiento SGR: %s", path_balance.name)
    logger.info("Curva S: %s", path_curva.name)
    logger.info("Reprogramaciones no permitidas: %s", path_reprog.name)
    logger.info("Ajustes: %s", path_ajustes.name)
    logger.info("Continuidad (ICCI): %s", path_continuidad.name)
    logger.info("Habilitaciones (ICCI): %s", path_habilitaciones.name)

    df_proyectos, df_ejecutores = cargar_proyectos_y_ejecutores(path_balance)
    df_periodos = cargar_periodos(path_curva)
    df_reprogramaciones = cargar_reprogramaciones(path_reprog)
    df_ajustes = cargar_ajustes(path_ajustes)
    df_continuidad = cargar_continuidad(path_continuidad)
    df_habilitaciones = cargar_habilitaciones(path_habilitaciones)

    validar_cobertura_ajustes(df_proyectos, df_ajustes)

    tablas = {
        "Ejecutores": df_ejecutores,
        "Proyectos": df_proyectos,
        "Periodos": df_periodos,
        "Reprogramaciones": df_reprogramaciones,
        "Ajustes": df_ajustes,
        "Continuidad": df_continuidad,
        "Habilitaciones": df_habilitaciones,
    }

    # Crear el directorio de salida si no existe
    salida.parent.mkdir(parents=True, exist_ok=True)

    # Nombres de tabla Excel (sin espacios, únicos por libro)
    # anchos_col: lista de anchos en caracteres por columna (None = auto desde muestra)
    config_hojas = {
        "Ejecutores":      ("TabMaestroEjecutores",  [20, 52, 20, 30, 20, 25, 25, 18]),
        "Proyectos":       ("TabMaestroProyectos",   [18, 52, 20, 18, 20, 18, 18, 18, 18, 18]),
        "Periodos":        ("TabMaestroPeriodos",    [18, 18, 10, 14, 18, 18, 14]),
        "Reprogramaciones":("TabMaestroReprog",      [18, 18, 18, 18, 18, 14]),
        "Ajustes":         ("TabMaestroAjustes",     None),
        "Continuidad":     ("TabMaestroContinuidad", [18, 18, 14, 14, 16, 14, 14, 18, 18, 18, 18]),
        "Habilitaciones":  ("TabMaestroHabilitaciones", [18, 20, 18, 35, 22, 18]),
    }

    with pd.ExcelWriter(salida, engine="openpyxl") as writer:
        for nombre_hoja, tabla in tablas.items():
            tabla.to_excel(writer, sheet_name=nombre_hoja, index=False)
            nombre_tab, anchos = config_hojas[nombre_hoja]
            _aplicar_formato_tabla(writer, nombre_hoja, nombre_tab,
                                   anchos_col=anchos)

    return tablas


def imprimir_resumen(tablas: dict[str, pd.DataFrame]) -> None:
    proyectos = tablas["Proyectos"]
    ejecutores = tablas["Ejecutores"]
    periodos = tablas["Periodos"]
    ajustes = tablas["Ajustes"]
    continuidad = tablas["Continuidad"]
    habilitaciones = tablas["Habilitaciones"]

    n_proyectos = proyectos["bpin"].nunique()
    n_ejecutores = ejecutores["codigo_ejecutor"].nunique()

    bpin_con_curva = set(periodos["bpin"].unique())
    con_curva = proyectos[proyectos["bpin"].isin(bpin_con_curva)]
    sin_curva = proyectos[~proyectos["bpin"].isin(bpin_con_curva)]

    logger.info("=" * 60)
    logger.info("RESUMEN EXCEL MAESTRO")
    logger.info("=" * 60)
    logger.info("Proyectos únicos (BPIN, TODOS los estados, sin filtrar): %s", n_proyectos)
    conteo_estados = proyectos["estado"].value_counts().to_dict()
    logger.info("Desglose por estado: %s", conteo_estados)
    logger.info("Ejecutores únicos: %s", n_ejecutores)
    n_con_fecha_inicio = ejecutores["fecha_inicio_ejecutor"].notna().sum()
    logger.info(
        "Ejecutores con fecha_inicio_ejecutor válida: %s de %s",
        n_con_fecha_inicio, n_ejecutores,
    )
    logger.info(
        "Proyectos con datos en Curva S: %s de %s (%.1f%%)",
        len(con_curva), n_proyectos, 100 * len(con_curva) / n_proyectos if n_proyectos else 0,
    )
    logger.info("Proyectos SIN datos en Curva S: %s de %s", len(sin_curva), n_proyectos)
    if not sin_curva.empty:
        ejemplos = sin_curva["bpin"].head(5).tolist()
        logger.info("Ejemplos de BPIN sin Curva S (máx. 5): %s", ejemplos)
    if not con_curva.empty:
        ejemplos_con = con_curva["bpin"].head(5).tolist()
        logger.info("Ejemplos de BPIN con Curva S (máx. 5): %s", ejemplos_con)
    n_con_ajuste = ajustes[ajustes["tiene_ajuste"].astype(str).str.strip().str.upper().isin(["SI", "SÍ"])]["bpin"].nunique()
    logger.info("BPIN con '¿EL PROYECTO PRESENTA AJUSTES?' = SI: %s de %s en Ajustes", n_con_ajuste, ajustes["bpin"].nunique())
    # ICCI — Continuidad
    n_cont_bpin = continuidad["bpin"].nunique()
    cortes_cont = sorted(continuidad["fecha_corte"].dropna().astype(str).str[:10].unique().tolist())
    logger.info(
        "Continuidad (ICCI): %s filas, %s BPIN únicos, corte(s) de GESPROY: %s",
        len(continuidad), n_cont_bpin, cortes_cont,
    )
    # ICCI — Habilitaciones
    n_hab_bpin = habilitaciones["bpin"].nunique()
    logger.info(
        "Habilitaciones (ICCI): %s solicitudes en %s BPIN únicos",
        len(habilitaciones), n_hab_bpin,
    )
    logger.info("=" * 60)


def configurar_logging(verbose: bool, guardar_log: bool, carpeta_log: Path) -> None:
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()

    formato = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", "%H:%M:%S")

    consola = logging.StreamHandler(sys.stdout)
    # Las alertas (WARNING) siempre se muestran; el detalle informativo
    # (resumen, avance de carga) solo se muestra con --verbose.
    consola.setLevel(logging.INFO if verbose else logging.WARNING)
    consola.setFormatter(formato)
    logger.addHandler(consola)

    if guardar_log:
        carpeta_log.mkdir(parents=True, exist_ok=True)
        nombre_log = f"construir_maestro_{datetime.now():%Y%m%d_%H%M%S}.log"
        archivo = logging.FileHandler(carpeta_log / nombre_log, encoding="utf-8")
        archivo.setLevel(logging.DEBUG)
        archivo.setFormatter(formato)
        logger.addHandler(archivo)
        logger.debug("Log guardado en: %s", carpeta_log / nombre_log)


# =============================================================================
# CLI
# =============================================================================

def parsear_argumentos(argv=None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Construye el Excel Maestro a partir de 6 excels fuente: "
                    "Balance seguimiento SGR, Curva S, Reprogramaciones no permitidas, "
                    "Ajustes, Universo reportes continuidad (ICCI) y "
                    "Consolidado habilitación proyectos (ICCI)."
    )
    parser.add_argument(
        "--path_files", required=True, type=Path,
        help="Carpeta donde están los 3 excels fuente.",
    )
    parser.add_argument(
        "--output", "-o", type=Path, default=None,
        help=f"Ruta/nombre del excel maestro de salida "
             f"(default: {NOMBRE_SALIDA_DEFAULT} dentro de --path_files).",
    )
    parser.add_argument(
        "--verbose", "-v", action="store_true",
        help="Imprime en consola el resumen: proyectos únicos, ejecutores "
             "únicos y cobertura de Curva S.",
    )
    parser.add_argument(
        "--log", action="store_true",
        help="Guarda un archivo .log con el detalle de la ejecución dentro de --path_files.",
    )
    return parser.parse_args(argv)


def main(argv=None) -> None:
    args = parsear_argumentos(argv)

    path_files = args.path_files.expanduser().resolve()
    if not path_files.is_dir():
        raise NotADirectoryError(f"--path_files no es una carpeta válida: {path_files}")

    salida = args.output.expanduser().resolve() if args.output else RUTA_SALIDA_DEFAULT

    configurar_logging(verbose=args.verbose, guardar_log=args.log, carpeta_log=path_files)

    tablas = construir_excel_maestro(path_files, salida)

    if args.verbose:
        imprimir_resumen(tablas)

    print(f"Excel maestro generado en: {salida}")


if __name__ == "__main__":
    main()