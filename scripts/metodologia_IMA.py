"""
metodologia_IMA.py
================================
Calcula el Indicador de Magnitud de Ajustes (IMA) a partir del Excel
Maestro (hojas 'Proyectos' y 'Ajustes') y genera un Excel de metodología
paso a paso, replicando el formato de "IMA - Propuesta.xlsx" (3 hojas +
una hoja base con el cruce a nivel de proyecto), con TABLAS de Excel
(TableStyleMedium9) y FÓRMULAS REALES (SUMIFS/COUNTIFS/SUMPRODUCT), no
valores pegados — así el archivo se recalcula solo si alguien edita un
dato en '0_Base'.

METODOLOGÍA (verificada celda por celda contra "IMA - Propuesta.xlsx")
-------------------------------------------------------------------------
Universo: proyectos del Excel Maestro con estado en ESTADOS_VALIDOS_IMA
(Sin Contratar, En Ejecución, Terminado — excluye Sin Migrar; Desaprobado
y Liberación de Recursos no existen en la columna `estado` del maestro,
mismo criterio que ya usa el Índice de Experiencia).

"JOIN" Proyectos <-> Ajustes: LEFT JOIN desde Proyectos (filtrado a
estados válidos) hacia Ajustes, por `bpin`. Los BPIN que están en Ajustes
pero NO en Proyectos se descartan (no forman parte del universo de
trabajo). Los que están en Proyectos pero NO en Ajustes se tratan como
"sin ajuste" (tiene_ajuste=NO), no como dato faltante.

Var 1 · Proporción de proyectos con ajuste
    = N° proyectos con ajuste efectivo / N° proyectos válidos del ejecutor

Var 1 · Proporción de proyectos con ajuste — nota sobre negativos:
    un proyecto marcado tiene_ajuste="SI" cuenta para Var1 SIEMPRE, incluso
    si su ajuste neto es negativo (reintegro/disminución). Un reintegro es
    un evento relevante (implicó trámite, un recurso inmovilizado un
    tiempo, incide en el valor total aprobado) y no debe tratarse como
    "sin ajuste" solo porque no aporta a la magnitud (Var2). Confirmado por
    Alix Obando (correo 27/08/2026, punto 2): "un reintegro sí es un evento
    relevante"; el ejecutor no queda idéntico a uno que nunca ajustó.

Var 2 · Proporción en el valor del ajuste
    = Σ ajuste_valor_usado / Σ valor_total_inicial
    "ajuste_valor_usado" depende de MODO_AJUSTE_SGR_ONLY (ver abajo).
    Correcciones importantes sobre esta suma (agregadas tras revisar datos
    reales y confirmadas por Alix Obando, correo 27/08/2026):
      (a) "Filas basura": hay proyectos donde la bandera de ajuste dice NO
          pero igual quedó cargado un monto de ajuste (dato sucio del
          archivo fuente). ajuste_valor_usado es 0 en TODA fila donde
          tiene_ajuste_efectivo no sea "SI", sin importar lo que traiga la
          columna de valor cruda — así esas filas nunca suman.
      (b) Los valores negativos NO restan — se acotan a 0 A NIVEL DE
          PROYECTO (no a nivel de ejecutor). Antes se sumaba el neto por
          ejecutor y se aplicaba MAX(0, suma), lo que permitía que un
          reintegro grande en un proyecto "neteara" y ocultara el
          sobrecosto de OTROS proyectos del mismo ejecutor — ejemplo real
          señalado por Alix: Norte de Santander tenía 59 proyectos con
          ajuste al alza (+$129 mil millones) y 5 con reintegros (-$105 mil
          millones); con el neteo por ejecutor la magnitud se desplomaba a
          $24,6 mil millones, escondiendo el riesgo real de los 59
          proyectos que sí sobrecostaron. Ahora cada proyecto aporta
          MAX(0, su propio ajuste) de forma individual ANTES de sumar por
          ejecutor, así que un reintegro en el proyecto A nunca puede
          "deshacer" el sobrecosto del proyecto B (son eventos fiscales
          distintos, marcados por BPIN). Con esto la suma por ejecutor ya
          no puede dar negativa por construcción, así que ya no hace falta
          un MAX(0,...) adicional a ese nivel.

Var 3 · Rapidez del ajuste
    = promedio en meses de (fecha_ajuste_1 − fecha_aprobacion), sobre
      proyectos con ajuste efectivo Y fecha_ajuste_1 válida (>= fecha
      aprobación). A diferencia de la implementación anterior
      (calcular_ima.py, ya obsoleta), esta SÍ es la fecha del PRIMER
      ajuste real, no un proxy con la última fecha.

Cuartiles de escala: se arman con COLUMNA_ESCALA_CUARTIL (ver
HIPERPARÁMETRO abajo — hoy: valor_total_inicial, por decisión explícita
del usuario, distinto de lo que hace "IMA - Propuesta.xlsx" que usa el
valor ACTUAL/con-ajustes; queda documentado como desviación consciente).

Percentil dentro de cada cuartil — fórmula verificada contra el excel de
referencia (NO es la fórmula ad-hoc del Índice de Experiencia):
    posición  = N° de entidades del MISMO CUARTIL cuyo valor es
                ESTRICTAMENTE MENOR (ties NO se dividen a la mitad)
    percentil = posición / N° total del cuartil × 100
Consecuencia (verificada, no es un caso especial en el código): como 0 es
el valor mínimo posible de Var1/Var2, cualquier ejecutor sin ajuste cae
en percentil 0 automáticamente. Igual para Var3, tratando "sin fecha de
ajuste válida" como riesgo mínimo (peor score de meses posible).

IMA = 30%·Pct.V1 + 32%·Pct.V2 + 38%·Pct.V3 (pesos editables en la propia
hoja 3, celdas $C$3/$F$3/$I$3 — igual que el excel de referencia).
Puntaje de riesgo = IMA (a mayor IMA, mayor riesgo). Nivel: <33 Bajo,
33-67 Medio, >=67 Alto (cortes editables en $B$6/$D$6 de la hoja 3).

HIPERPARÁMETROS (arriba de este archivo — cambiar aquí y re-ejecutar)
-------------------------------------------------------------------------
MODO_AJUSTE_SGR_ONLY (default False):
    False (default) -> "ajuste" = columna '¿EL PROYECTO PRESENTA AJUSTES?'
        tal cual viene del excel de Ajustes (SI incluye ajustes a
        cualquier fuente); Σ TOTAL AJUSTE usa 'TOTAL AJUSTE AL PROYECTO'
        (SGR + otras fuentes).
    True -> redefine el universo de "ajuste" a solo proyectos con
        valor_total_ajustes_sgr != 0. Un proyecto marcado SI en
        '¿EL PROYECTO PRESENTA AJUSTES?' pero cuyo ajuste fue SOLO de
        otras fuentes (no SGR) se trata como SIN ajuste en este modo.
        Σ TOTAL AJUSTE usa 'VALOR TOTAL AJUSTES SGR'.
    Confirmado explícitamente por el usuario: el default es el universo
    TOTAL (todas las fuentes); este flag es para poder recalcular todo
    con un cambio de una línea cuando se necesite ver solo SGR.

COLUMNA_ESCALA_CUARTIL (default "valor_total_inicial"):
    Columna de la hoja 'Proyectos' del maestro usada para sumar por
    ejecutor y formar los 4 cuartiles de escala. Alternativa documentada:
    "valor_total_proyecto" (el valor ACTUAL/con ajustes — es lo que usa
    "IMA - Propuesta.xlsx", pero el usuario pidió explícitamente usar el
    valor INICIAL). Cambiar esta constante y re-ejecutar recalcula todo
    el árbol de cuartiles/percentiles/índice consistentemente.

USO
----
    python metodologia_IMA.py --maestro EXCEL_MAESTRO_ICS.xlsx --output METODOLOGIA_IMA.xlsx
"""

from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# =============================================================================
# HIPERPARÁMETROS METODOLÓGICOS — cambiar aquí y re-ejecutar
# =============================================================================

MODO_AJUSTE_SGR_ONLY = False  # HIPERPARÁMETRO — ver docstring arriba
COLUMNA_ESCALA_CUARTIL = "valor_total_inicial"  # HIPERPARÁMETRO — alternativa: "valor_total_proyecto"

ESTADOS_VALIDOS_IMA = ("sin contratar", "en ejecución", "terminado")  # HIPERPARÁMETRO

PESO_VAR1 = 0.30  # HIPERPARÁMETRO — proporción de proyectos con ajuste
PESO_VAR2 = 0.32  # HIPERPARÁMETRO — proporción en el valor del ajuste
PESO_VAR3 = 0.38  # HIPERPARÁMETRO — rapidez del ajuste tras aprobación

UMBRAL_BAJO = 33
UMBRAL_MEDIO = 67

# =============================================================================
# ESTILO — paleta DNP
# =============================================================================

FONT_NAME = "Verdana"
PINK, TEAL, YELLOW, NAVY = "FE187B", "00C4C3", "FFCB1E", "002060"
FILL_TITLE = PatternFill("solid", fgColor=NAVY)
FILL_HEADER = PatternFill("solid", fgColor=NAVY)
FILL_NOTE = PatternFill("solid", fgColor="FFF2CC")
FILL_PARAM = PatternFill("solid", fgColor="E7F7F7")
FILL_ALT = PatternFill("solid", fgColor="F2F2F2")
FILL_BAJO = PatternFill("solid", fgColor="C6EFCE")
FILL_MEDIO = PatternFill("solid", fgColor="FFEB9C")
FILL_ALTO = PatternFill("solid", fgColor="FFC7CE")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

logger = logging.getLogger("metodologia_IMA")


def _titulo(ws, fila, texto, n_cols):
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=n_cols)
    c = ws.cell(row=fila, column=1, value=texto)
    c.font = Font(name=FONT_NAME, size=13, bold=True, color="FFFFFF")
    c.fill = FILL_TITLE
    c.alignment = LEFT
    ws.row_dimensions[fila].height = 22


def _nota(ws, fila, texto, n_cols, altura=30):
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=n_cols)
    c = ws.cell(row=fila, column=1, value=texto)
    c.font = Font(name=FONT_NAME, size=9, italic=True, color="666666")
    c.fill = FILL_NOTE
    c.alignment = LEFT
    ws.row_dimensions[fila].height = altura


def _encabezados(ws, fila, encabezados):
    for j, h in enumerate(encabezados, start=1):
        c = ws.cell(row=fila, column=j, value=h)
        c.font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
        c.fill = FILL_HEADER
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[fila].height = 30


def _tabla(ws, nombre, fila_encabezado, fila_fin, n_cols):
    if fila_fin < fila_encabezado + 1:
        return  # sin filas de datos, no se puede crear la tabla
    ref = f"A{fila_encabezado}:{get_column_letter(n_cols)}{fila_fin}"
    tabla = Table(displayName=nombre, ref=ref)
    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tabla)


# =============================================================================
# PASO 1: CARGA Y "JOIN" DESDE EL MAESTRO
# =============================================================================

def cargar_base(path_maestro: str) -> pd.DataFrame:
    proyectos = pd.read_excel(path_maestro, sheet_name="Proyectos", dtype={"codigo_ejecutor": str, "bpin": str})
    ajustes = pd.read_excel(path_maestro, sheet_name="Ajustes", dtype={"bpin": str})
    ejecutores = pd.read_excel(path_maestro, sheet_name="Ejecutores", dtype={"codigo_ejecutor": str})

    proyectos["estado_norm"] = proyectos["estado"].astype(str).str.strip().str.lower()
    validos = proyectos[proyectos["estado_norm"].isin(ESTADOS_VALIDOS_IMA)].copy()
    logger.info(
        "Universo: %s de %s proyectos del maestro tienen estado válido (%s).",
        len(validos), len(proyectos), ", ".join(ESTADOS_VALIDOS_IMA),
    )

    # "JOIN": se parte de Proyectos (universo de trabajo) y se hace LEFT JOIN
    # hacia Ajustes por bpin. Los BPIN de Ajustes que no están en Proyectos
    # quedan fuera (no entran al cálculo).
    bpins_ajustes = set(ajustes["bpin"])
    bpins_proyectos = set(validos["bpin"])
    solo_en_ajustes = bpins_ajustes - bpins_proyectos
    if solo_en_ajustes:
        logger.warning(
            "ALERTA — %s BPIN de la hoja 'Ajustes' NO están en el universo de "
            "'Proyectos' válidos y se descartan del cálculo del IMA (no forman "
            "parte de lo que ya se venía trabajando). Ejemplos: %s",
            len(solo_en_ajustes), list(solo_en_ajustes)[:5],
        )

    base = validos.merge(ajustes, on="bpin", how="left", suffixes=("", "_ajuste"))

    base["tiene_ajuste"] = base["tiene_ajuste"].fillna("NO").astype(str).str.strip().str.upper()
    base["tiene_ajuste"] = base["tiene_ajuste"].replace({"SÍ": "SI"})
    for col in ("valor_total_ajustes_sgr", "total_ajuste_proyecto"):
        base[col] = pd.to_numeric(base[col], errors="coerce").fillna(0.0)
    base["valor_total_inicial"] = pd.to_numeric(base["valor_total_inicial"], errors="coerce")
    base["fecha_aprobacion"] = pd.to_datetime(base["fecha_aprobacion"], errors="coerce")
    base["fecha_ajuste_1"] = pd.to_datetime(base["fecha_ajuste_1"], errors="coerce")

    # Etiquetas del ejecutor (nombre/tipo), solo para mostrar — se traen del
    # catálogo Ejecutores, no son parte del cálculo.
    base = base.merge(
        ejecutores[["codigo_ejecutor", "nombre_ejecutor", "tipo_ejecutor"]],
        on="codigo_ejecutor", how="left",
    )

    columnas_finales = [
        "bpin", "codigo_ejecutor", "nombre_ejecutor", "tipo_ejecutor", "estado",
        "valor_total_inicial", "valor_total_proyecto",
        "tiene_ajuste", "valor_total_ajustes_sgr", "total_ajuste_proyecto",
        "fecha_aprobacion", "fecha_ajuste_1",
    ]
    return base[columnas_finales].reset_index(drop=True)


# =============================================================================
# PASO 2: HOJA '0_Base' — cruce a nivel de proyecto, con fórmulas
# =============================================================================

def escribir_hoja_0_base(wb: Workbook, base: pd.DataFrame) -> int:
    ws = wb.create_sheet("0_Base")

    encabezados = [
        "bpin", "codigo_ejecutor", "nombre_ejecutor", "tipo_ejecutor", "estado",
        "valor_total_inicial", "valor_total_proyecto",
        "tiene_ajuste", "valor_total_ajustes_sgr", "total_ajuste_proyecto",
        "fecha_aprobacion", "fecha_ajuste_1",
        "tiene_ajuste_efectivo", "ajuste_valor_usado", "meses_a_1er_ajuste",
    ]
    fila_encabezado = 1
    _encabezados(ws, fila_encabezado, encabezados)

    n = len(base)
    fila_inicio = fila_encabezado + 1
    fila_fin = fila_encabezado + n

    # columnas letra por nombre, para armar las fórmulas de forma legible
    col = {nombre: get_column_letter(i + 1) for i, nombre in enumerate(encabezados)}

    for i, fila_datos in enumerate(base.itertuples(index=False), start=fila_inicio):
        ws.cell(row=i, column=1, value=fila_datos.bpin)
        ws.cell(row=i, column=2, value=fila_datos.codigo_ejecutor)
        ws.cell(row=i, column=3, value=fila_datos.nombre_ejecutor)
        ws.cell(row=i, column=4, value=fila_datos.tipo_ejecutor)
        ws.cell(row=i, column=5, value=fila_datos.estado)
        ws.cell(row=i, column=6, value=None if pd.isna(fila_datos.valor_total_inicial) else float(fila_datos.valor_total_inicial))
        ws.cell(row=i, column=7, value=None if pd.isna(fila_datos.valor_total_proyecto) else float(fila_datos.valor_total_proyecto))
        ws.cell(row=i, column=8, value=fila_datos.tiene_ajuste)
        ws.cell(row=i, column=9, value=float(fila_datos.valor_total_ajustes_sgr))
        ws.cell(row=i, column=10, value=float(fila_datos.total_ajuste_proyecto))
        ws.cell(row=i, column=11, value=None if pd.isna(fila_datos.fecha_aprobacion) else fila_datos.fecha_aprobacion.to_pydatetime())
        ws.cell(row=i, column=12, value=None if pd.isna(fila_datos.fecha_ajuste_1) else fila_datos.fecha_ajuste_1.to_pydatetime())

        # --- columnas calculadas, con FÓRMULA (no valor pegado) ---
        # tiene_ajuste_efectivo primero; ajuste_valor_usado depende de ella
        # (si dice NO, el valor usado es 0 SIN IMPORTAR lo que traiga la
        # columna de valor cruda — hay filas "basura" donde la bandera dice
        # NO pero igual quedó un monto de ajuste cargado; no deben sumar).
        if MODO_AJUSTE_SGR_ONLY:
            f_tiene_efectivo = f'=IF({col["valor_total_ajustes_sgr"]}{i}<>0,"SI","NO")'
            col_valor_crudo = col["valor_total_ajustes_sgr"]
        else:
            f_tiene_efectivo = f'={col["tiene_ajuste"]}{i}'
            col_valor_crudo = col["total_ajuste_proyecto"]
        ws.cell(row=i, column=13, value=f_tiene_efectivo)

        col_tiene_efectivo = col["tiene_ajuste_efectivo"]
        # ajuste_valor_usado a nivel de PROYECTO: si el proyecto tiene ajuste
        # efectivo, se usa el valor crudo, pero NUNCA negativo — un ajuste
        # negativo (reintegro/disminución) no debe restar valor a la Var2 del
        # ejecutor. Se acota a 0 aquí, a nivel de proyecto, para que un
        # reintegro en el proyecto A no pueda compensar/ocultar el sobrecosto
        # de otro proyecto B del mismo ejecutor al sumarlos (regla del correo
        # de Alix Obando, 27/08/2026: "sumar únicamente las partes positivas;
        # los valores negativos no restan, se quedan en cero").
        f_valor_usado = f'=IF({col_tiene_efectivo}{i}="SI",MAX(0,{col_valor_crudo}{i}),0)'
        ws.cell(row=i, column=14, value=f_valor_usado)

        f_meses = (
            f'=IF(AND({col["tiene_ajuste_efectivo"]}{i}="SI",{col["fecha_ajuste_1"]}{i}<>"",'
            f'({col["fecha_ajuste_1"]}{i}-{col["fecha_aprobacion"]}{i})>=0),'
            f'({col["fecha_ajuste_1"]}{i}-{col["fecha_aprobacion"]}{i})/30.44,"")'
        )
        ws.cell(row=i, column=15, value=f_meses)

        # formatos
        ws.cell(row=i, column=6).number_format = "#,##0"
        ws.cell(row=i, column=7).number_format = "#,##0"
        ws.cell(row=i, column=9).number_format = "#,##0"
        ws.cell(row=i, column=10).number_format = "#,##0"
        ws.cell(row=i, column=11).number_format = "yyyy-mm-dd"
        ws.cell(row=i, column=12).number_format = "yyyy-mm-dd"
        ws.cell(row=i, column=14).number_format = "#,##0"
        ws.cell(row=i, column=15).number_format = "0.0"

    anchos = [16, 14, 42, 18, 14, 16, 16, 10, 16, 16, 13, 13, 10, 16, 12]
    for j, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    _tabla(ws, "TablaBase", fila_encabezado, fila_fin, len(encabezados))
    ws.freeze_panes = f"A{fila_inicio}"

    # Reporte de calidad de dato: fechas de ajuste inconsistentes (anteriores
    # a la aprobación) — se reporta por consola, la fórmula ya las excluye
    # del promedio (quedan como "" en meses_a_1er_ajuste).
    dias = (base["fecha_ajuste_1"] - base["fecha_aprobacion"]).dt.days
    inconsistentes = base[base["tiene_ajuste"].eq("SI") & dias.notna() & (dias < 0)]
    if not inconsistentes.empty:
        logger.warning(
            "ALERTA — %s proyectos con ajuste tienen fecha_ajuste_1 ANTERIOR a "
            "fecha_aprobacion (dato inconsistente). La fórmula de "
            "meses_a_1er_ajuste los excluye automáticamente. Ejemplos de BPIN: %s",
            len(inconsistentes), inconsistentes["bpin"].head(5).tolist(),
        )

    return fila_fin


# =============================================================================
# PASO 3: HOJA '1_DatosFuente' — agregación por ejecutor, con fórmulas
# =============================================================================

def escribir_hoja_1(wb: Workbook, base: pd.DataFrame, fila_fin_base: int) -> tuple[int, int]:
    ws = wb.create_sheet("1_DatosFuente")

    # Universo de ejecutores: TODO ejecutor con al menos 1 proyecto válido
    # (aunque sea sin ningún ajuste, o con valor_total_inicial faltante para
    # todos sus proyectos) debe aparecer aquí — con sus variables en 0 en
    # ese caso, NO excluido. Antes se excluía a quien tuviera suma_escala
    # <= 0 o NaN, lo cual dejaba fuera ejecutores sin ajuste (su índice de
    # riesgo real es 0, no "sin dato"), y además distorsionaba el
    # percentil de TODOS los demás porque el N del cuartil quedaba más
    # chico de lo que debía ser. Corregido por pedido explícito del
    # usuario: solo se excluye a quien NO tenga NINGÚN proyecto válido
    # (ese si es "sin historial en el SGR", y ni siquiera llega a este
    # groupby porque `base` ya viene filtrado a proyectos válidos).
    resumen = base.groupby("codigo_ejecutor").agg(
        nombre_ejecutor=("nombre_ejecutor", "first"),
        tipo_ejecutor=("tipo_ejecutor", "first"),
        suma_escala=(COLUMNA_ESCALA_CUARTIL, "sum"),
    ).reset_index()
    n_sin_escala = int(resumen["suma_escala"].isna().sum() + (resumen["suma_escala"] <= 0).sum())
    resumen["suma_escala"] = resumen["suma_escala"].fillna(0.0).clip(lower=0.0)
    if n_sin_escala:
        logger.info(
            "Aviso: %s ejecutores no tienen '%s' > 0 en ninguno de sus proyectos "
            "válidos (falta el dato, o no tienen ajustes). Se incluyen igual, con "
            "escala 0 — quedan en Q1 (menor escala) y su índice de riesgo da 0 si "
            "efectivamente no tienen ajustes.", n_sin_escala, COLUMNA_ESCALA_CUARTIL,
        )
    resumen = resumen.sort_values("codigo_ejecutor").reset_index(drop=True)

    nota = (
        f"Universo = proyectos {', '.join(e.upper() for e in ESTADOS_VALIDOS_IMA)}. "
        f"'JOIN' con la hoja Ajustes por BPIN (solo BPIN que ya estaban en Proyectos). "
        f"Ajuste efectivo = columna 'tiene_ajuste_efectivo' de 0_Base "
        f"(modo {'SOLO SGR' if MODO_AJUSTE_SGR_ONLY else 'TODAS LAS FUENTES'} — "
        f"cambiar MODO_AJUSTE_SGR_ONLY en metodologia_IMA.py para recalcular con el otro modo). "
        f"Los cuartiles se forman con Σ '{COLUMNA_ESCALA_CUARTIL}' por ejecutor "
        f"(cambiar COLUMNA_ESCALA_CUARTIL en metodologia_IMA.py para usar otra columna, "
        f"ej. 'valor_total_proyecto' para el valor actual con ajustes)."
    )
    encabezados = [
        "Cód.\nEjecutor", "Entidad Ejecutora", "Tipo\nEjecutor",
        "N° Proy.\nválidos", "N° Proy.\ncon ajuste",
        "Σ TOTAL AJUSTE\n($)", "Σ VALOR INICIAL\n($)",
        "Var 1 · Prop.\nproyectos", "Var 2 · Prop.\nvalor", "Var 3 · Meses prom.\n1er ajuste",
        f"Σ {COLUMNA_ESCALA_CUARTIL}\n($) — escala cuartil", "CUARTIL", "N°\ngrupo",
    ]

    _nota(ws, 1, nota, len(encabezados), altura=45)
    fila_encabezado = 2
    _encabezados(ws, fila_encabezado, encabezados)

    fila_inicio = fila_encabezado + 1
    fila_fin = fila_encabezado + len(resumen)

    hoja0 = "'0_Base'"
    rango_cod_base = f"{hoja0}!$B$2:$B${fila_fin_base}"
    rango_ajuste_base = f"{hoja0}!$M$2:$M${fila_fin_base}"
    rango_ajusteval_base = f"{hoja0}!$N$2:$N${fila_fin_base}"
    rango_inicial_base = f"{hoja0}!$F$2:$F${fila_fin_base}"
    _col_escala = "F" if COLUMNA_ESCALA_CUARTIL == "valor_total_inicial" else "G"
    rango_escala_base = f"{hoja0}!${_col_escala}$2:${_col_escala}${fila_fin_base}"
    rango_meses_base = f"{hoja0}!$O$2:$O${fila_fin_base}"

    for i, ejecutor in enumerate(resumen.itertuples(index=False), start=fila_inicio):
        cod = ejecutor.codigo_ejecutor
        ws.cell(row=i, column=1, value=cod)
        ws.cell(row=i, column=2, value=ejecutor.nombre_ejecutor)
        ws.cell(row=i, column=3, value=ejecutor.tipo_ejecutor)

        ws.cell(row=i, column=4, value=f'=COUNTIF({rango_cod_base},$A{i})')
        ws.cell(row=i, column=5, value=f'=COUNTIFS({rango_cod_base},$A{i},{rango_ajuste_base},"SI")')
        # Ya no hace falta MAX(0,...) aquí: ajuste_valor_usado (col N de
        # 0_Base) nunca es negativo a nivel de proyecto (ver 0_Base), así
        # que la suma por ejecutor tampoco puede serlo. Sumar directo evita
        # el "neteo oculto" que se daba antes cuando se aplicaba MAX(0,...)
        # sobre la suma del ejecutor: un reintegro grande en un proyecto ya
        # no puede compensar/ocultar el sobrecosto de otros proyectos del
        # mismo ejecutor, porque cada proyecto ya aporta 0 (nunca negativo)
        # de forma individual.
        ws.cell(row=i, column=6, value=(
            f'=SUMIFS({rango_ajusteval_base},{rango_cod_base},$A{i},'
            f'{rango_ajuste_base},"SI")'
        ))
        ws.cell(row=i, column=7, value=f'=SUMIFS({rango_inicial_base},{rango_cod_base},$A{i})')
        ws.cell(row=i, column=8, value=f'=IFERROR($E{i}/$D{i},0)')
        ws.cell(row=i, column=9, value=f'=IFERROR($F{i}/$G{i},0)')
        ws.cell(row=i, column=10, value=(
            f'=IFERROR(AVERAGEIFS({rango_meses_base},{rango_cod_base},$A{i},{rango_ajuste_base},"SI"),"")'
        ))
        ws.cell(row=i, column=11, value=f'=SUMIFS({rango_escala_base},{rango_cod_base},$A{i})')

        # RANK.EQ es posterior a Excel 2007: openpyxl necesita el prefijo
        # "_xlfn." para que Excel/LibreOffice lo reconozcan (si no, da #NAME?).
        f_rank = f'_xlfn.RANK.EQ($K{i},$K${fila_inicio}:$K${fila_fin},1)'
        f_n = f'COUNT($K${fila_inicio}:$K${fila_fin})'
        f_num_cuartil = f'MIN(4,MAX(1,ROUNDUP({f_rank}/({f_n}/4),0)))'
        ws.cell(row=i, column=12, value=(
            f'=CHOOSE({f_num_cuartil},'
            f'"Q1 — Menor escala","Q2 — Escala media-baja",'
            f'"Q3 — Escala media-alta","Q4 — Mayor escala")'
        ))
        ws.cell(row=i, column=13, value=f'=COUNTIF($L${fila_inicio}:$L${fila_fin},$L{i})')

        for j in (6, 7, 11):
            ws.cell(row=i, column=j).number_format = "#,##0"
        ws.cell(row=i, column=8).number_format = "0.0%"
        ws.cell(row=i, column=9).number_format = "0.00%"
        ws.cell(row=i, column=10).number_format = "0.0"

        if (i - fila_inicio) % 2 == 1:
            for j in range(1, len(encabezados) + 1):
                ws.cell(row=i, column=j).fill = FILL_ALT

    anchos = [10, 42, 16, 10, 10, 16, 16, 12, 12, 14, 18, 20, 8]
    for j, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    _tabla(ws, "TablaDatosFuente", fila_encabezado, fila_fin, len(encabezados))
    ws.freeze_panes = f"A{fila_inicio}"

    return fila_inicio, fila_fin


# =============================================================================
# PASO 4: HOJA '2_Rankings_Percentiles'
# =============================================================================

def escribir_hoja_2(wb: Workbook, fila_inicio_1: int, fila_fin_1: int) -> tuple[int, int]:
    ws = wb.create_sheet("2_Rankings_Percentiles")
    n_cols = 13

    _nota(ws, 1, (
        "Cada entidad se compara SOLO con las de su cuartil de escala. "
        "Percentil = N° de entidades del MISMO CUARTIL con valor ESTRICTAMENTE MENOR / N° del grupo × 100."
    ), n_cols, 26)
    _nota(ws, 2, (
        "Orientación a riesgo: más proporción de proyectos = más riesgo · mayor proporción del valor = más riesgo · "
        "MENOS meses hasta el 1er ajuste = más riesgo (por eso el percentil de Var3 se calcula sobre la rapidez, no sobre los meses)."
    ), n_cols, 26)
    _nota(ws, 3, (
        "Entidades sin fecha de ajuste válida (sin ajuste, o con ajuste sin fecha SGR) quedan en percentil 0 de Var 3 "
        "de forma automática (se tratan como el riesgo mínimo posible en esa variable, no como caso especial)."
    ), n_cols, 26)

    encabezados = [
        "Cód.", "Entidad Ejecutora", "Cuartil",
        "Var1\nProp.\nproyectos", "Pos.\nV1", "PERC.\nV1",
        "Var2\nProp.\nvalor", "Pos.\nV2", "PERC.\nV2",
        "Var3\nMeses 1er\najuste", "Pos.\nV3", "PERC.\nV3 (rapidez)",
        "Lectura",
    ]
    fila_encabezado = 5
    _encabezados(ws, fila_encabezado, encabezados)

    fila_inicio = fila_encabezado + 1
    n = fila_fin_1 - fila_inicio_1 + 1
    fila_fin = fila_encabezado + n

    hoja1 = "'1_DatosFuente'"
    rango_cuartil_1 = f"{hoja1}!$L${fila_inicio_1}:$L${fila_fin_1}"
    rango_v1_1 = f"{hoja1}!$H${fila_inicio_1}:$H${fila_fin_1}"
    rango_v2_1 = f"{hoja1}!$I${fila_inicio_1}:$I${fila_fin_1}"
    rango_v3_1 = f"{hoja1}!$J${fila_inicio_1}:$J${fila_fin_1}"

    for k, i1 in enumerate(range(fila_inicio_1, fila_fin_1 + 1)):
        i2 = fila_inicio + k
        ws.cell(row=i2, column=1, value=f"={hoja1}!$A${i1}")
        ws.cell(row=i2, column=2, value=f"={hoja1}!$B${i1}")
        ws.cell(row=i2, column=3, value=f"={hoja1}!$L${i1}")
        ws.cell(row=i2, column=4, value=f"={hoja1}!$H${i1}")
        ws.cell(row=i2, column=5, value=f'=COUNTIFS({rango_cuartil_1},$C{i2},{rango_v1_1},"<"&$D{i2})')
        ws.cell(row=i2, column=6, value=f'=IFERROR($E{i2}/COUNTIF({rango_cuartil_1},$C{i2})*100,0)')
        ws.cell(row=i2, column=7, value=f"={hoja1}!$I${i1}")
        ws.cell(row=i2, column=8, value=f'=COUNTIFS({rango_cuartil_1},$C{i2},{rango_v2_1},"<"&$G{i2})')
        ws.cell(row=i2, column=9, value=f'=IFERROR($H{i2}/COUNTIF({rango_cuartil_1},$C{i2})*100,0)')
        ws.cell(row=i2, column=10, value=f"={hoja1}!$J${i1}")
        ws.cell(row=i2, column=11, value=(
            f'=IF($J{i2}="",0,SUMPRODUCT(({rango_cuartil_1}=$C{i2})*({rango_v3_1}>$J{i2})))'
        ))
        ws.cell(row=i2, column=12, value=f'=IFERROR($K{i2}/COUNTIF({rango_cuartil_1},$C{i2})*100,0)')
        ws.cell(row=i2, column=13, value=(
            f'=IF($D{i2}=0,"Sin proyectos con ajuste — riesgo mínimo.",'
            f'"En "&$C{i2}&": más proyectos con ajuste que el "&TEXT($F{i2},"0")&"%, '
            f'mayor valor que el "&TEXT($I{i2},"0")&"%; "&'
            f'IF($J{i2}="","ajuste sin fecha SGR.",'
            f'"ajustó ~"&TEXT($J{i2},"0")&" meses tras aprobar (más rápido que el "&TEXT($L{i2},"0")&"%).") )'
        ))

        ws.cell(row=i2, column=4).number_format = "0.0%"
        ws.cell(row=i2, column=7).number_format = "0.00%"
        ws.cell(row=i2, column=10).number_format = "0.0"
        for j in (6, 9, 12):
            ws.cell(row=i2, column=j).number_format = "0.0"

        if k % 2 == 1:
            for j in range(1, n_cols + 1):
                ws.cell(row=i2, column=j).fill = FILL_ALT

    anchos = [10, 42, 20, 10, 8, 10, 10, 8, 10, 10, 8, 10, 55]
    for j, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    _tabla(ws, "TablaRankings", fila_encabezado, fila_fin, n_cols)
    ws.freeze_panes = f"A{fila_inicio}"

    return fila_inicio, fila_fin


# =============================================================================
# PASO 5: HOJA '3_Indice_Riesgo'
# =============================================================================

def escribir_hoja_3(wb: Workbook, fila_inicio_2: int, fila_fin_2: int) -> None:
    ws = wb.create_sheet("3_Indice_Riesgo")
    n_cols = 12

    _titulo(ws, 1, "PASO 2 — VARIABLE DE AJUSTES (ÍNDICE DE RIESGO)", n_cols)

    ws.cell(row=2, column=1, value="PARÁMETROS — cambiar aquí para recalcular todo")
    ws.cell(row=2, column=1).font = Font(name=FONT_NAME, size=10, bold=True)

    etiquetas_pesos = [
        (2, "Peso Var1 · Proporción de proyectos", PESO_VAR1),
        (5, "Peso Var2 · Proporción del valor", PESO_VAR2),
        (8, "Peso Var3 · Rapidez del ajuste", PESO_VAR3),
    ]
    for col_lbl, texto, valor in etiquetas_pesos:
        ws.cell(row=3, column=col_lbl, value=texto).font = Font(name=FONT_NAME, size=9)
        c = ws.cell(row=3, column=col_lbl + 1, value=valor)
        c.font = Font(name=FONT_NAME, size=11, bold=True)
        c.fill = FILL_PARAM
        c.number_format = "0%"
        c.border = BORDER
    ws.cell(row=3, column=11, value="Suma").font = Font(name=FONT_NAME, size=9)
    c = ws.cell(row=3, column=12, value="=C3+F3+I3")
    c.font = Font(name=FONT_NAME, size=11, bold=True)
    c.fill = FILL_PARAM
    c.number_format = "0%"
    c.border = BORDER

    _nota(ws, 4, (
        "Pesos derivados del método de ENTROPÍA DE SHANNON sobre los tres percentiles (ver propuesta IMA). "
        "Editables: cambiar C3/F3/I3 recalcula Aportes e Índice de toda la tabla."
    ), n_cols, 26)

    ws.cell(row=5, column=1, value="CORTES DE NIVEL DE RIESGO").font = Font(name=FONT_NAME, size=10, bold=True)
    ws.cell(row=6, column=1, value="Corte Bajo/Medio →").font = Font(name=FONT_NAME, size=9)
    c = ws.cell(row=6, column=2, value=UMBRAL_BAJO)
    c.font = Font(name=FONT_NAME, size=11, bold=True)
    c.fill = FILL_PARAM
    c.border = BORDER
    ws.cell(row=6, column=3, value="Corte Medio/Alto →").font = Font(name=FONT_NAME, size=9)
    c = ws.cell(row=6, column=4, value=UMBRAL_MEDIO)
    c.font = Font(name=FONT_NAME, size=11, bold=True)
    c.fill = FILL_PARAM
    c.border = BORDER
    ws.cell(row=6, column=6, value=f"Bajo: 0–{UMBRAL_BAJO}  |  Medio: {UMBRAL_BAJO}–{UMBRAL_MEDIO}  |  Alto: {UMBRAL_MEDIO}–100").font = Font(name=FONT_NAME, size=9, italic=True)

    _nota(ws, 7, (
        'ÍNDICE = (Pct_V1×$C$3)+(Pct_V2×$F$3)+(Pct_V3×$I$3). Los "Aporte" muestran cuánto suma cada variable (suman el índice).'
    ), n_cols, 20)

    encabezados = [
        "Cód.", "Entidad Ejecutora", "Cuartil",
        "Pct.V1\nProp.", "Pct.V2\nValor", "Pct.V3\nRapidez",
        "Aporte\nV1", "Aporte\nV2", "Aporte\nV3",
        "ÍNDICE\nDE RIESGO", "NIVEL\nDE RIESGO", "Posición\nen grupo",
    ]
    fila_encabezado = 9
    _encabezados(ws, fila_encabezado, encabezados)

    fila_inicio = fila_encabezado + 1
    n = fila_fin_2 - fila_inicio_2 + 1
    fila_fin = fila_encabezado + n

    hoja2 = "'2_Rankings_Percentiles'"

    for k, i2 in enumerate(range(fila_inicio_2, fila_fin_2 + 1)):
        i3 = fila_inicio + k
        ws.cell(row=i3, column=1, value=f"={hoja2}!$A${i2}")
        ws.cell(row=i3, column=2, value=f"={hoja2}!$B${i2}")
        ws.cell(row=i3, column=3, value=f"={hoja2}!$C${i2}")
        ws.cell(row=i3, column=4, value=f"={hoja2}!$F${i2}")
        ws.cell(row=i3, column=5, value=f"={hoja2}!$I${i2}")
        ws.cell(row=i3, column=6, value=f"={hoja2}!$L${i2}")
        ws.cell(row=i3, column=7, value=f"=$D{i3}*$C$3")
        ws.cell(row=i3, column=8, value=f"=$E{i3}*$F$3")
        ws.cell(row=i3, column=9, value=f"=$F{i3}*$I$3")
        ws.cell(row=i3, column=10, value=f"=$G{i3}+$H{i3}+$I{i3}")
        ws.cell(row=i3, column=11, value=(
            f'=IF($J{i3}<$B$6,"Riesgo Bajo",IF($J{i3}<$D$6,"Riesgo Medio","Riesgo Alto"))'
        ))
        ws.cell(row=i3, column=12, value=(
            f'=COUNTIFS($C${fila_inicio}:$C${fila_fin},$C{i3},$J${fila_inicio}:$J${fila_fin},"<"&$J{i3})'
            f'&" de "&COUNTIF($C${fila_inicio}:$C${fila_fin},$C{i3})'
        ))

        for j in (4, 5, 6, 7, 8, 9, 10):
            ws.cell(row=i3, column=j).number_format = "0.0"

        if k % 2 == 1:
            for j in range(1, n_cols + 1):
                ws.cell(row=i3, column=j).fill = FILL_ALT

    anchos = [10, 42, 20, 10, 10, 10, 10, 10, 10, 12, 14, 14]
    for j, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    _tabla(ws, "TablaIndiceRiesgo", fila_encabezado, fila_fin, n_cols)
    ws.freeze_panes = f"A{fila_inicio}"

    # Formato condicional sobre NIVEL DE RIESGO (texto, columna K) y también
    # sobre el ÍNDICE DE RIESGO (numérico, columna J) — mismos 3 colores,
    # mismos cortes (UMBRAL_BAJO/UMBRAL_MEDIO), para que el número y la
    # etiqueta de texto siempre se vean consistentes.
    from openpyxl.formatting.rule import CellIsRule

    rango_indice = f"$J${fila_inicio}:$J${fila_fin}"
    ws.conditional_formatting.add(
        rango_indice,
        CellIsRule(operator="between", formula=["0", str(UMBRAL_BAJO)], fill=FILL_BAJO),
    )
    ws.conditional_formatting.add(
        rango_indice,
        CellIsRule(operator="between", formula=[str(UMBRAL_BAJO), str(UMBRAL_MEDIO)], fill=FILL_MEDIO),
    )
    ws.conditional_formatting.add(
        rango_indice,
        CellIsRule(operator="between", formula=[str(UMBRAL_MEDIO), "100"], fill=FILL_ALTO),
    )

    rango_nivel = f"$K${fila_inicio}:$K${fila_fin}"
    ws.conditional_formatting.add(
        rango_nivel, CellIsRule(operator="equal", formula=['"Riesgo Bajo"'], fill=FILL_BAJO)
    )
    ws.conditional_formatting.add(
        rango_nivel, CellIsRule(operator="equal", formula=['"Riesgo Medio"'], fill=FILL_MEDIO)
    )
    ws.conditional_formatting.add(
        rango_nivel, CellIsRule(operator="equal", formula=['"Riesgo Alto"'], fill=FILL_ALTO)
    )


# =============================================================================
# PIPELINE COMPLETO
# =============================================================================

def generar_metodologia_ima(path_maestro: str, path_salida: str) -> None:
    logger.info("Modo de ajuste: %s", "SOLO SGR" if MODO_AJUSTE_SGR_ONLY else "TODAS LAS FUENTES (default)")
    logger.info("Columna de escala para cuartiles: %s", COLUMNA_ESCALA_CUARTIL)

    base = cargar_base(path_maestro)

    wb = Workbook()
    del wb["Sheet"]

    fila_fin_base = escribir_hoja_0_base(wb, base)
    fila_inicio_1, fila_fin_1 = escribir_hoja_1(wb, base, fila_fin_base)
    fila_inicio_2, fila_fin_2 = escribir_hoja_2(wb, fila_inicio_1, fila_fin_1)
    escribir_hoja_3(wb, fila_inicio_2, fila_fin_2)

    wb.save(path_salida)
    logger.info("Excel de metodología IMA generado en: %s", path_salida)


def configurar_logging(verbose: bool) -> None:
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()
    formato = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", "%H:%M:%S")
    consola = logging.StreamHandler(sys.stdout)
    consola.setLevel(logging.INFO if verbose else logging.WARNING)
    consola.setFormatter(formato)
    logger.addHandler(consola)


def main():
    parser = argparse.ArgumentParser(description="Genera el Excel de metodología del IMA (Indicador de Magnitud de Ajustes) con fórmulas reales.")
    parser.add_argument("--maestro", default="EXCEL_MAESTRO_ICS.xlsx")
    parser.add_argument("--output", "-o", default="METODOLOGIA_IMA.xlsx")
    parser.add_argument("--verbose", "-v", action="store_true")
    parser.add_argument(
        "--sgr-only", action="store_true",
        help="Calcula el universo de ajustes usando SOLO montos SGR (equivalente a "
             "poner MODO_AJUSTE_SGR_ONLY=True en el archivo). Por defecto usa todas las fuentes.",
    )
    args = parser.parse_args()

    configurar_logging(args.verbose)

    if args.sgr_only:
        global MODO_AJUSTE_SGR_ONLY
        MODO_AJUSTE_SGR_ONLY = True

    if not Path(args.maestro).exists():
        print(f"ERROR: no encontré el archivo '{args.maestro}'", file=sys.stderr)
        sys.exit(1)

    generar_metodologia_ima(args.maestro, args.output)


if __name__ == "__main__":
    main()