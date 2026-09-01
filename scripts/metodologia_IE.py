"""
metodologia_IE.py
====================
Versión "auditable" del cálculo del Índice de Experiencia (IE — ficha
técnica "Perfil de riesgo de ejecutores del SGR / Componente
institucional, Variable 1", ver IE - Propuesta.pdf): en vez de una sola
hoja de resultado, genera un Excel con 3 hojas replicando la estructura
del archivo de referencia entregado por el usuario (PE Experiencia SGR
Cálculo.xlsx), para poder auditar o defender el número de cualquier
ejecutor sin cálculos ocultos. Usa la misma lógica y los mismos
parámetros que calcular_IE.py — este script no cambia la metodología,
solo la desglosa paso a paso.

FÓRMULA — ver docstring de calcular_IE.py para el detalle completo.
    Índice_Experiencia_e = (Percentil_Antigüedad_e x 30%)
                          + (Percentil_N_Proyectos_e x 40%)
                          + (Percentil_Valor_e x 30%)
    Puntaje_Riesgo_e = 100 - Índice_Experiencia_e

EMPATES EN EL PERCENTIL — replica EXACTAMENTE la fórmula ad-hoc del Excel
de referencia (no el rank promedio estándar usado en el ICH), por
decisión explícita del usuario. Ver calcular_posicion_percentil().

UNIVERSO DE PROYECTOS Y EJECUTORES SIN HISTORIAL — ver docstring de
calcular_IE.py (mismo hiperparámetro ESTADOS_VALIDOS_IE — hoy Sin
Contratar/En Ejecución/Terminado, excluyendo Sin Migrar — y mismo criterio
de "no aplica" para ejecutores sin historial, ambos por decisión
explícita del usuario).

SALIDA — 3 HOJAS (misma estructura que PE Experiencia SGR Cálculo.xlsx)
-------------------------------------------------------------------------
    1_DatosFuente          -> por ejecutor: fecha primer proyecto, N°
                               proyectos, valor total, antigüedad (años),
                               valor (millones), cuartil asignado, N°
                               entidades en su grupo.
    2_Rankings_Percentiles -> posición y percentil de CADA ejecutor DENTRO
                               de su cuartil, para las 3 variables
                               (antigüedad, N° proyectos, valor), con una
                               columna de lectura en texto plano.
    3_Indice_Riesgo        -> índice de experiencia final, puntaje de
                               riesgo, nivel de riesgo (coloreado: verde
                               claro = Bajo, amarillo claro = Medio,
                               rosado = Alto) y la posición del ejecutor
                               dentro de su cuartil según el índice final.

Las 3 hojas son Tablas de Excel (con filtros y bandas de color), igual
que metodologia_ICH.py.

USO
----
    python metodologia_IE.py --path_maestro ruta/EXCEL_MAESTRO_ICS.xlsx
    python metodologia_IE.py --path_maestro ruta/carpeta --verbose
    python metodologia_IE.py --path_maestro ruta/carpeta --verbose --log
    python metodologia_IE.py --path_maestro ruta/carpeta --output Metodologia_IE.xlsx
"""

from __future__ import annotations

import argparse
import logging
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# =============================================================================
# PARÁMETROS METODOLÓGICOS (ver docstring / IE - Propuesta.pdf — no tocar
# sin acordarlo antes)
# =============================================================================

PESO_ANTIGUEDAD = 0.30
PESO_N_PROYECTOS = 0.40
PESO_VALOR = 0.30

UMBRAL_RIESGO_BAJO = 33
UMBRAL_RIESGO_ALTO = 67

N_CUARTILES = 4
NOMBRES_CUARTIL = [
    "Q1 — Menor escala", "Q2 — Escala media-baja",
    "Q3 — Escala media-alta", "Q4 — Mayor escala",
]

# HIPERPARÁMETRO — estados de proyecto que se INCLUYEN en el IE. El Excel
# Maestro ya no filtra por estado, así que el IE cuenta SOLO los proyectos
# cuyo estado esté en esta lista (el resto se descarta). Por decisión
# explícita del usuario: Sin Contratar, En Ejecución y Terminado cuentan;
# Sin Migrar se excluye (un proyecto sin migrar al sistema no refleja
# gestión real todavía). Edita esta lista si el criterio cambia — no hace
# falta tocar nada más del pipeline.
ESTADOS_VALIDOS_IE = ["Sin Contratar", "En Ejecución", "Terminado"]

NOMBRE_MAESTRO_DEFAULT = "EXCEL_MAESTRO.xlsx"
NOMBRE_SALIDA_DEFAULT = "Metodologia_IE.xlsx"
ESTILO_TABLA = "TableStyleMedium9"

# HIPERPARÁMETRO — IE SECTORIZADO. Un ejecutor puede tener proyectos en
# varios sectores; para elegir el ejecutor idóneo de un proyecto NUEVO en
# el sector S, el IE se recalcula usando SOLO los proyectos de ese
# ejecutor que son del sector S (Opción "filtro puro", decisión explícita
# del usuario tras discutir alternativas — ver metodologia_IE.py docstring
# y conversación de diseño). Si el ejecutor no tiene proyectos en S, su
# IE_sector es 0 (Riesgo Alto) — no es un dato faltante, es información
# real: cero trayectoria demostrable en ese sector.
CUARTIL_UNICO_SECTOR = "Grupo único (sector con pocos ejecutores)"
SIN_EXPERIENCIA_SECTOR = "Sin proyectos en el sector"

FILL_RIESGO_BAJO = PatternFill("solid", fgColor="C6EFCE")
FILL_RIESGO_MEDIO = PatternFill("solid", fgColor="FFEB9C")
FILL_RIESGO_ALTO = PatternFill("solid", fgColor="FFC7CE")
FONT_RIESGO_BAJO = Font(color="006100")
FONT_RIESGO_MEDIO = Font(color="9C6500")
FONT_RIESGO_ALTO = Font(color="9C0006")

logger = logging.getLogger("metodologia_IE")


# =============================================================================
# LECTURA DEL EXCEL MAESTRO
# =============================================================================

def leer_maestro(path_maestro: Path):
    ejecutores = pd.read_excel(
        path_maestro, sheet_name="Ejecutores",
        dtype={"codigo_ejecutor": str}, parse_dates=["fecha_inicio_ejecutor"],
    )
    proyectos = pd.read_excel(
        path_maestro, sheet_name="Proyectos",
        dtype={"bpin": str, "codigo_ejecutor": str},
    )

    ejecutores["codigo_ejecutor"] = ejecutores["codigo_ejecutor"].astype(str).str.strip()
    proyectos["codigo_ejecutor"] = proyectos["codigo_ejecutor"].astype(str).str.strip()
    proyectos["bpin"] = proyectos["bpin"].astype(str).str.strip()

    return ejecutores, proyectos


def filtrar_proyectos_validos_ie(proyectos: pd.DataFrame) -> pd.DataFrame:
    """
    El Excel Maestro ya NO filtra por estado. El IE cuenta SOLO los
    proyectos cuyo estado esté en ESTADOS_VALIDOS_IE (hiperparámetro
    editable arriba).
    """
    n_total = proyectos["bpin"].nunique()
    proyectos_validos = proyectos[proyectos["estado"].isin(ESTADOS_VALIDOS_IE)].copy()
    n_validos = proyectos_validos["bpin"].nunique()
    logger.info(
        "Filtro de estado para el IE: %s BPIN en el maestro -> %s conservados "
        "(estado en %s). Descartados: %s",
        n_total, n_validos, ESTADOS_VALIDOS_IE, n_total - n_validos,
    )
    return proyectos_validos


# =============================================================================
# HOJA 1: DATOS FUENTE POR EJECUTOR
# =============================================================================

def calcular_datos_fuente(ejecutores: pd.DataFrame, proyectos: pd.DataFrame, fecha_corte) -> pd.DataFrame:
    n_proy = proyectos.groupby("codigo_ejecutor")["bpin"].nunique().rename("n_proyectos")

    valores = proyectos.dropna(subset=["valor_total_proyecto"])
    valor_total = valores.groupby("codigo_ejecutor")["valor_total_proyecto"].sum().rename("valor_total_ejecutado")

    columnas_base = ["codigo_ejecutor", "capacidad_institucional", "fecha_inicio_ejecutor"]
    if "nombre_ejecutor" in ejecutores.columns:
        columnas_base.insert(1, "nombre_ejecutor")
    df = ejecutores[columnas_base].drop_duplicates("codigo_ejecutor").copy()

    df = df.merge(n_proy, on="codigo_ejecutor", how="left")
    df = df.merge(valor_total, on="codigo_ejecutor", how="left")
    df["n_proyectos"] = df["n_proyectos"].fillna(0).astype(int)
    df["valor_total_ejecutado"] = df["valor_total_ejecutado"].fillna(0.0)
    df["valor_millones"] = df["valor_total_ejecutado"] / 1e6

    df["antiguedad_anios"] = (
        pd.Timestamp(fecha_corte) - df["fecha_inicio_ejecutor"]
    ).dt.days / 365.25

    df["sin_historial"] = df["fecha_inicio_ejecutor"].isna() | (df["n_proyectos"] == 0)
    return df


def asignar_cuartil(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["cuartil"] = pd.NA
    evaluables = df.loc[~df["sin_historial"]]
    if len(evaluables) >= N_CUARTILES:
        rangos = evaluables["valor_total_ejecutado"].rank(method="first")
        cuartiles = pd.qcut(rangos, N_CUARTILES, labels=NOMBRES_CUARTIL)
        df.loc[evaluables.index, "cuartil"] = cuartiles.astype(str)

    tam_grupo = df.groupby("cuartil", dropna=True)["codigo_ejecutor"].transform("count")
    df["n_entidades_grupo"] = tam_grupo
    return df


def armar_hoja1_datos_fuente(df: pd.DataFrame) -> pd.DataFrame:
    columnas = [
        "codigo_ejecutor",
        "nombre_ejecutor" if "nombre_ejecutor" in df.columns else None,
        "fecha_inicio_ejecutor", "n_proyectos", "valor_total_ejecutado",
        "antiguedad_anios", "valor_millones", "cuartil", "n_entidades_grupo",
    ]
    columnas = [c for c in columnas if c is not None]
    return df[columnas].sort_values("codigo_ejecutor").reset_index(drop=True)


# =============================================================================
# HOJA 2: RANKING Y PERCENTIL DENTRO DEL CUARTIL
# =============================================================================

def calcular_posicion_percentil(serie: pd.Series) -> tuple[pd.Series, pd.Series]:
    """
    Replica EXACTAMENTE la fórmula de empates del Excel de referencia (PE
    Experiencia SGR Cálculo.xlsx), distinta del rank(pct=True,
    method="average") usado en el resto del proyecto (decisión explícita
    del usuario para que calce con ese archivo):

        posición  = (n° de valores estrictamente MENORES en el grupo)
                  + ceil(n° de valores EMPATADOS con este / 2)
        percentil = posición / n° de entidades del grupo x 100

    A mayor valor, mayor percentil (0=peor del grupo, 100=mejor).
    """
    n = len(serie)
    conteo_por_valor = serie.value_counts()
    count_equal = serie.map(conteo_por_valor)
    count_less = serie.rank(method="min") - 1
    posicion = count_less + np.ceil(count_equal / 2)
    percentil = posicion / n * 100
    return posicion, percentil


def calcular_rankings_en_df(df: pd.DataFrame) -> pd.DataFrame:
    """Agrega las columnas de posición/percentil al df completo (se
    reutiliza tanto para armar la Hoja 2 como para calcular el índice)."""
    df = df.copy()
    for col in ["posicion_antiguedad", "percentil_antiguedad",
                "posicion_n_proyectos", "percentil_n_proyectos",
                "posicion_valor", "percentil_valor"]:
        df[col] = np.nan

    for cuartil, grupo in df.loc[~df["sin_historial"]].groupby("cuartil"):
        pos_a, pct_a = calcular_posicion_percentil(grupo["antiguedad_anios"])
        pos_n, pct_n = calcular_posicion_percentil(grupo["n_proyectos"])
        pos_v, pct_v = calcular_posicion_percentil(grupo["valor_total_ejecutado"])
        df.loc[grupo.index, "posicion_antiguedad"] = pos_a
        df.loc[grupo.index, "percentil_antiguedad"] = pct_a
        df.loc[grupo.index, "posicion_n_proyectos"] = pos_n
        df.loc[grupo.index, "percentil_n_proyectos"] = pct_n
        df.loc[grupo.index, "posicion_valor"] = pos_v
        df.loc[grupo.index, "percentil_valor"] = pct_v
    return df


def armar_hoja2_rankings(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    def lectura(fila):
        if fila["sin_historial"]:
            return "Sin historial en el SGR — la variable no aplica."
        return (
            f"Dentro de {fila['cuartil']}: supera al {fila['percentil_antiguedad']:.0f}% "
            f"en antigüedad, al {fila['percentil_n_proyectos']:.0f}% en proyectos, "
            f"al {fila['percentil_valor']:.0f}% en valor."
        )
    df["lectura"] = df.apply(lectura, axis=1)

    columnas = [
        "codigo_ejecutor",
        "nombre_ejecutor" if "nombre_ejecutor" in df.columns else None,
        "cuartil", "antiguedad_anios", "posicion_antiguedad", "percentil_antiguedad",
        "n_proyectos", "posicion_n_proyectos", "percentil_n_proyectos",
        "valor_total_ejecutado", "posicion_valor", "percentil_valor", "lectura",
    ]
    columnas = [c for c in columnas if c is not None]
    return df[columnas].sort_values("codigo_ejecutor").reset_index(drop=True)


# =============================================================================
# HOJA 3: ÍNDICE DE EXPERIENCIA Y PUNTAJE DE RIESGO
# =============================================================================

def calcular_indice_y_riesgo(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["indice_experiencia"] = np.where(
        df["sin_historial"],
        np.nan,
        df["percentil_antiguedad"] * PESO_ANTIGUEDAD
        + df["percentil_n_proyectos"] * PESO_N_PROYECTOS
        + df["percentil_valor"] * PESO_VALOR,
    )
    df["puntaje_riesgo"] = np.where(df["sin_historial"], np.nan, 100 - df["indice_experiencia"])

    condiciones = [
        df["sin_historial"],
        df["puntaje_riesgo"] < UMBRAL_RIESGO_BAJO,
        df["puntaje_riesgo"] < UMBRAL_RIESGO_ALTO,
    ]
    niveles = ["No aplica (sin historial)", "Riesgo Bajo", "Riesgo Medio"]
    df["nivel_riesgo"] = np.select(condiciones, niveles, default="Riesgo Alto")
    return df


def armar_hoja3_indice_riesgo(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["posicion_en_grupo"] = ""
    for cuartil, grupo in df.loc[~df["sin_historial"]].groupby("cuartil"):
        # Posición DESCENDENTE por índice (1 = mejor/mayor índice del cuartil),
        # igual convención informativa del Excel de referencia. Misma fórmula
        # de empates, pero contando MAYORES en vez de MENORES.
        n = len(grupo)
        conteo_por_valor = grupo["indice_experiencia"].value_counts()
        count_equal = grupo["indice_experiencia"].map(conteo_por_valor)
        count_greater = grupo["indice_experiencia"].rank(method="min", ascending=False) - 1
        posicion = (count_greater + np.ceil(count_equal / 2)).astype(int)
        df.loc[grupo.index, "posicion_en_grupo"] = posicion.astype(str) + f" de {n}"

    columnas = [
        "codigo_ejecutor",
        "nombre_ejecutor" if "nombre_ejecutor" in df.columns else None,
        "cuartil", "percentil_antiguedad", "percentil_n_proyectos", "percentil_valor",
        "indice_experiencia", "puntaje_riesgo", "nivel_riesgo", "posicion_en_grupo",
    ]
    columnas = [c for c in columnas if c is not None]
    resultado = df[columnas].rename(columns={
        "percentil_antiguedad": "pct_antiguedad",
        "percentil_n_proyectos": "pct_n_proyectos",
        "percentil_valor": "pct_valor",
    })
    return resultado.sort_values("puntaje_riesgo", ascending=True, na_position="last").reset_index(drop=True)


# =============================================================================
# PIPELINE COMPLETO -> 3 HOJAS
# =============================================================================

def calcular_ie_por_pasos(ejecutores, proyectos, fecha_corte):
    proyectos = filtrar_proyectos_validos_ie(proyectos)
    df = calcular_datos_fuente(ejecutores, proyectos, fecha_corte)
    n_sin_historial = int(df["sin_historial"].sum())

    df = asignar_cuartil(df)
    hoja1 = armar_hoja1_datos_fuente(df)

    df_rank = calcular_rankings_en_df(df)
    hoja2 = armar_hoja2_rankings(df_rank)

    df_final = calcular_indice_y_riesgo(df_rank)
    hoja3 = armar_hoja3_indice_riesgo(df_final)

    hojas = {
        "1_DatosFuente": hoja1,
        "2_Rankings_Percentiles": hoja2,
        "3_Indice_Riesgo": hoja3,
    }
    return hojas, n_sin_historial, df_final


# =============================================================================
# IE SECTORIZADO — mismo método, aplicado por sector (filtro puro)
# =============================================================================

def calcular_confianza_muestral(n_proyectos: pd.Series) -> pd.Series:
    """Etiqueta informativa sobre qué tan sólida es la base del cálculo
    (NO cambia el índice, solo advierte al lector). 1-2 proyectos = Baja,
    3-4 = Media, 5+ = Alta. Hiperparámetro editable si el criterio cambia."""
    condiciones = [n_proyectos == 0, n_proyectos <= 2, n_proyectos <= 4]
    etiquetas = ["No aplica (sin proyectos)", "Baja", "Media"]
    return pd.Series(np.select(condiciones, etiquetas, default="Alta"), index=n_proyectos.index)


def calcular_ie_un_sector(ejecutores: pd.DataFrame, proyectos_validos: pd.DataFrame,
                           sector: str, fecha_corte) -> pd.DataFrame:
    """Replica calcular_datos_fuente + asignar_cuartil + rankings + índice,
    pero el universo de proyectos son SOLO los del sector dado, y la
    antigüedad es la antigüedad del ejecutor EN ESE SECTOR (fecha de su
    primer proyecto del sector), no su antigüedad institucional general."""
    proy_s = proyectos_validos[proyectos_validos["sector"] == sector]

    n_proy = proy_s.groupby("codigo_ejecutor")["bpin"].nunique().rename("n_proyectos_sector")
    valores = proy_s.dropna(subset=["valor_total_proyecto"])
    valor_total = valores.groupby("codigo_ejecutor")["valor_total_proyecto"].sum().rename("valor_total_sector")
    fecha_primer = proy_s.groupby("codigo_ejecutor")["fecha_aprobacion"].min().rename("fecha_primer_proyecto_sector")

    columnas_base = ["codigo_ejecutor"]
    if "nombre_ejecutor" in ejecutores.columns:
        columnas_base.append("nombre_ejecutor")
    df = ejecutores[columnas_base].drop_duplicates("codigo_ejecutor").copy()
    df["sector"] = sector

    df = df.merge(n_proy, on="codigo_ejecutor", how="left")
    df = df.merge(valor_total, on="codigo_ejecutor", how="left")
    df = df.merge(fecha_primer, on="codigo_ejecutor", how="left")
    df["n_proyectos_sector"] = df["n_proyectos_sector"].fillna(0).astype(int)
    df["valor_total_sector"] = df["valor_total_sector"].fillna(0.0)
    df["valor_millones_sector"] = df["valor_total_sector"] / 1e6

    df["sin_experiencia_sector"] = df["n_proyectos_sector"] == 0
    df["antiguedad_sector_anios"] = np.where(
        df["sin_experiencia_sector"], np.nan,
        (pd.Timestamp(fecha_corte) - df["fecha_primer_proyecto_sector"]).dt.days / 365.25,
    )
    df["confianza_muestral"] = calcular_confianza_muestral(df["n_proyectos_sector"])

    # --- cuartil por escala (valor_total_sector), con fallback a grupo único ---
    df["cuartil_sector"] = pd.NA
    evaluables = df.loc[~df["sin_experiencia_sector"]]
    if len(evaluables) >= N_CUARTILES:
        rangos = evaluables["valor_total_sector"].rank(method="first")
        cuartiles = pd.qcut(rangos, N_CUARTILES, labels=NOMBRES_CUARTIL)
        df.loc[evaluables.index, "cuartil_sector"] = cuartiles.astype(str)
    elif len(evaluables) > 0:
        # No hay suficientes ejecutores con historial en este sector para
        # formar 4 cuartiles (mínimo N_CUARTILES) -> se comparan todos entre
        # sí en un único grupo, en vez de quedar sin percentil calculado.
        df.loc[evaluables.index, "cuartil_sector"] = CUARTIL_UNICO_SECTOR

    # --- percentiles dentro del cuartil_sector (misma fórmula de empates) ---
    for col in ["pct_antiguedad_sector", "pct_n_proyectos_sector", "pct_valor_sector"]:
        df[col] = np.nan
    for _, grupo in df.loc[~df["sin_experiencia_sector"]].groupby("cuartil_sector"):
        _, pct_a = calcular_posicion_percentil(grupo["antiguedad_sector_anios"])
        _, pct_n = calcular_posicion_percentil(grupo["n_proyectos_sector"])
        _, pct_v = calcular_posicion_percentil(grupo["valor_total_sector"])
        df.loc[grupo.index, "pct_antiguedad_sector"] = pct_a
        df.loc[grupo.index, "pct_n_proyectos_sector"] = pct_n
        df.loc[grupo.index, "pct_valor_sector"] = pct_v

    # --- índice e indicador de riesgo ---
    df["indice_experiencia_sector"] = np.where(
        df["sin_experiencia_sector"], 0.0,
        df["pct_antiguedad_sector"] * PESO_ANTIGUEDAD
        + df["pct_n_proyectos_sector"] * PESO_N_PROYECTOS
        + df["pct_valor_sector"] * PESO_VALOR,
    )
    df["puntaje_riesgo_sector"] = 100 - df["indice_experiencia_sector"]

    condiciones = [
        df["sin_experiencia_sector"],
        df["puntaje_riesgo_sector"] < UMBRAL_RIESGO_BAJO,
        df["puntaje_riesgo_sector"] < UMBRAL_RIESGO_ALTO,
    ]
    niveles = [f"Riesgo Alto ({SIN_EXPERIENCIA_SECTOR})", "Riesgo Bajo", "Riesgo Medio"]
    df["nivel_riesgo_sector"] = np.select(condiciones, niveles, default="Riesgo Alto")
    df.loc[df["sin_experiencia_sector"], "cuartil_sector"] = SIN_EXPERIENCIA_SECTOR

    return df


def calcular_ie_sectorizado(ejecutores: pd.DataFrame, proyectos: pd.DataFrame, fecha_corte) -> pd.DataFrame:
    """Corre calcular_ie_un_sector() para cada sector presente en el
    maestro y devuelve una tabla larga: una fila por (ejecutor, sector)."""
    proyectos_validos = filtrar_proyectos_validos_ie(proyectos)
    sectores = sorted(proyectos_validos["sector"].dropna().unique())
    logger.info("IE sectorizado: %s sectores detectados en el maestro.", len(sectores))

    partes = [calcular_ie_un_sector(ejecutores, proyectos_validos, s, fecha_corte) for s in sectores]
    resultado = pd.concat(partes, ignore_index=True)

    columnas = [
        "codigo_ejecutor",
        "nombre_ejecutor" if "nombre_ejecutor" in resultado.columns else None,
        "sector", "n_proyectos_sector", "valor_total_sector", "valor_millones_sector",
        "antiguedad_sector_anios", "cuartil_sector", "confianza_muestral",
        "pct_antiguedad_sector", "pct_n_proyectos_sector", "pct_valor_sector",
        "indice_experiencia_sector", "puntaje_riesgo_sector", "nivel_riesgo_sector",
    ]
    columnas = [c for c in columnas if c is not None]
    return resultado[columnas].sort_values(
        ["sector", "puntaje_riesgo_sector"], na_position="last"
    ).reset_index(drop=True)


def armar_matriz_ie_sector(hoja_sectorial: pd.DataFrame) -> pd.DataFrame:
    """Vista ancha para filtrar rápido: una fila por ejecutor, una columna
    por sector, valor = indice_experiencia_sector. Para uso operativo
    ('¿quién es idóneo para un proyecto de Agro?' -> ordenar esa columna)."""
    id_cols = ["codigo_ejecutor"] + (["nombre_ejecutor"] if "nombre_ejecutor" in hoja_sectorial.columns else [])
    base = hoja_sectorial[id_cols].drop_duplicates("codigo_ejecutor").sort_values("codigo_ejecutor")
    pivote = hoja_sectorial.pivot_table(
        index="codigo_ejecutor", columns="sector", values="indice_experiencia_sector", aggfunc="first",
    ).round(1)
    matriz = base.merge(pivote, on="codigo_ejecutor", how="left").reset_index(drop=True)
    return matriz


# =============================================================================
# ESCRITURA EN EXCEL COMO TABLAS (filtros, bandas de color, ancho automático)
# =============================================================================

def escribir_hoja_tabla(writer: pd.ExcelWriter, df: pd.DataFrame, nombre_hoja: str) -> None:
    df.to_excel(writer, sheet_name=nombre_hoja, index=False)
    ws = writer.sheets[nombre_hoja]

    n_filas, n_cols = df.shape
    if n_filas == 0 or n_cols == 0:
        return

    ultima_col = get_column_letter(n_cols)
    rango = f"A1:{ultima_col}{n_filas + 1}"
    tabla = Table(displayName=nombre_hoja, ref=rango)
    tabla.tableStyleInfo = TableStyleInfo(
        name=ESTILO_TABLA, showRowStripes=True, showColumnStripes=False,
        showFirstColumn=False, showLastColumn=False,
    )
    ws.add_table(tabla)

    for i, col in enumerate(df.columns, start=1):
        # str(v) sobre cada valor crudo (no .astype(str) de la columna): ver
        # nota en metodologia_ICH.py — .astype(str) en columnas object con
        # NaN a veces deja el NaN como float en vez de 'nan', y rompe len().
        muestra = df[col].head(200).tolist()
        ancho = max([len(str(col))] + [len(str(v)) for v in muestra]) + 2
        ws.column_dimensions[get_column_letter(i)].width = min(ancho, 40)

    ws.freeze_panes = "A2"


def colorear_columnas_riesgo(
    ws, df: pd.DataFrame,
    columnas=("indice_experiencia", "puntaje_riesgo", "nivel_riesgo"),
    col_nivel: str = "nivel_riesgo",
) -> None:
    if col_nivel not in df.columns:
        return
    mapa = {
        "Riesgo Bajo": (FILL_RIESGO_BAJO, FONT_RIESGO_BAJO),
        "Riesgo Medio": (FILL_RIESGO_MEDIO, FONT_RIESGO_MEDIO),
        "Riesgo Alto": (FILL_RIESGO_ALTO, FONT_RIESGO_ALTO),
    }
    col_idx = {c: i + 1 for i, c in enumerate(df.columns)}
    objetivo = [col_idx[c] for c in columnas if c in col_idx]
    if not objetivo:
        return
    for fila_i, nivel in enumerate(df[col_nivel], start=2):
        # nivel_riesgo_sector incluye la variante "Riesgo Alto (Sin proyectos
        # en el sector)" -> se colorea igual que "Riesgo Alto" por prefijo.
        estilo = mapa.get(nivel) or next(
            (v for k, v in mapa.items() if isinstance(nivel, str) and nivel.startswith(k)), None
        )
        if estilo is None:
            continue
        fill, font = estilo
        for col_i in objetivo:
            celda = ws.cell(row=fila_i, column=col_i)
            celda.fill = fill
            celda.font = font


def escribir_excel_por_pasos(salida: Path, hojas: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(salida, engine="openpyxl") as writer:
        wb = writer.book
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        for nombre_hoja, df in hojas.items():
            escribir_hoja_tabla(writer, df, nombre_hoja)

        colorear_columnas_riesgo(writer.sheets["3_Indice_Riesgo"], hojas["3_Indice_Riesgo"])

        if "4_IE_Sectorial" in hojas:
            colorear_columnas_riesgo(
                writer.sheets["4_IE_Sectorial"], hojas["4_IE_Sectorial"],
                columnas=("indice_experiencia_sector", "puntaje_riesgo_sector", "nivel_riesgo_sector"),
                col_nivel="nivel_riesgo_sector",
            )


# =============================================================================
# EXPLICACIÓN DE LA METODOLOGÍA (para --verbose)
# =============================================================================

def explicar_metodologia() -> None:
    logger.info("=" * 70)
    logger.info("METODOLOGÍA — Índice de Experiencia (IE)")
    logger.info("=" * 70)
    logger.info(
        "Hoja 1_DatosFuente: antiguedad_anios = (fecha_corte - "
        "fecha_inicio_ejecutor) en años. n_proyectos = n° de proyectos del "
        "ejecutor con estado en ESTADOS_VALIDOS_IE=%s (hiperparámetro "
        "editable). valor_total_ejecutado = SUMA(valor_total_proyecto) de "
        "esos mismos proyectos. cuartil = grupo de comparación por escala "
        "(Q1..Q4, ~mismo tamaño cada uno).", ESTADOS_VALIDOS_IE,
    )
    logger.info(
        "Hoja 2_Rankings_Percentiles: percentil_X = posición/N_grupo x 100, "
        "calculado DENTRO del cuartil para cada variable por separado. "
        "Empates: posición = menores_estrictos + ceil(empatados/2) — "
        "fórmula replicada del Excel de referencia del usuario (no es rank "
        "promedio estándar)."
    )
    logger.info(
        "Hoja 3_Indice_Riesgo: indice_experiencia = (pct_antiguedad x %.0f%%) "
        "+ (pct_n_proyectos x %.0f%%) + (pct_valor x %.0f%%). puntaje_riesgo "
        "= 100 - indice_experiencia. nivel_riesgo: puntaje_riesgo < %s -> "
        "Bajo | %s-%s -> Medio | >=%s -> Alto. Coloreado: verde claro=Bajo, "
        "amarillo claro=Medio, rosado=Alto.",
        PESO_ANTIGUEDAD * 100, PESO_N_PROYECTOS * 100, PESO_VALOR * 100,
        UMBRAL_RIESGO_BAJO, UMBRAL_RIESGO_BAJO, UMBRAL_RIESGO_ALTO, UMBRAL_RIESGO_ALTO,
    )
    logger.info(
        "Ejecutores sin historial (sin proyectos en el maestro o sin "
        "fecha_inicio_ejecutor): la variable 'no aplica' — quedan con "
        "indice/puntaje en blanco, nivel_riesgo='No aplica (sin historial)'."
    )
    logger.info(
        "Hoja 4_IE_Sectorial: MISMA fórmula y pesos, pero recalculada por "
        "sector (filtro puro) — cada ejecutor se compara solo contra otros "
        "ejecutores CON proyectos en ese sector. Antigüedad = años desde su "
        "primer proyecto EN ese sector (no antigüedad institucional). Si el "
        "ejecutor no tiene proyectos en el sector: indice=0, riesgo=100, "
        "nivel='Riesgo Alto (Sin proyectos en el sector)' — es información "
        "real (cero trayectoria ahí), no un dato faltante. Si el sector "
        "tiene menos de %s ejecutores evaluables, no se pueden formar "
        "cuartiles: se usa un único grupo de comparación "
        "('%s'). Columna 'confianza_muestral' (Baja/Media/Alta según n° de "
        "proyectos en el sector) es informativa, NO altera el índice.",
        N_CUARTILES, CUARTIL_UNICO_SECTOR,
    )
    logger.info(
        "Hoja 5_Matriz_IE_x_Sector: vista ancha (ejecutor x sector) del "
        "indice_experiencia_sector, para filtrar rápido al buscar el "
        "ejecutor idóneo de un proyecto nuevo en un sector dado."
    )
    logger.info("=" * 70)


def imprimir_resumen(hoja3: pd.DataFrame, n_sin_historial: int) -> None:
    logger.info("=" * 70)
    logger.info("RESUMEN — RECUENTO DEL IE")
    logger.info("=" * 70)

    n_total = hoja3["codigo_ejecutor"].nunique()
    logger.info("Ejecutores en total: %s", n_total)

    conteo_nivel = hoja3["nivel_riesgo"].value_counts()
    for nivel in ["Riesgo Bajo", "Riesgo Medio", "Riesgo Alto", "No aplica (sin historial)"]:
        logger.info("  %s: %s ejecutores (%.1f%%)", nivel, conteo_nivel.get(nivel, 0),
                    100 * conteo_nivel.get(nivel, 0) / n_total if n_total else 0)

    if "cuartil" in hoja3.columns:
        conteo_cuartil = hoja3["cuartil"].value_counts(dropna=False)
        logger.info("Ejecutores evaluables por cuartil de valor:")
        for cuartil in NOMBRES_CUARTIL:
            logger.info("  %s: %s ejecutores", cuartil, conteo_cuartil.get(cuartil, 0))

    logger.warning(
        "ALERTA — %s ejecutores sin historial en el SGR: la variable no "
        "aplica, quedaron con indice/puntaje en blanco.", n_sin_historial,
    )
    if n_sin_historial:
        codigos_sin_historial = hoja3.loc[
            hoja3["nivel_riesgo"].eq("No aplica (sin historial)"), "codigo_ejecutor"
        ].head(5).tolist()
        logger.info("Ejemplos de codigo_ejecutor sin historial (máx. 5): %s", codigos_sin_historial)

    logger.info("=" * 70)


# =============================================================================
# LOGGING / CLI
# =============================================================================

def configurar_logging(verbose: bool, guardar_log: bool, carpeta_log: Path) -> None:
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()

    formato = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", "%H:%M:%S")

    consola = logging.StreamHandler(sys.stdout)
    consola.setLevel(logging.INFO if verbose else logging.WARNING)
    consola.setFormatter(formato)
    logger.addHandler(consola)

    if guardar_log:
        carpeta_log.mkdir(parents=True, exist_ok=True)
        nombre_log = f"metodologia_IE_{datetime.now():%Y%m%d_%H%M%S}.log"
        archivo = logging.FileHandler(carpeta_log / nombre_log, encoding="utf-8")
        archivo.setLevel(logging.DEBUG)
        archivo.setFormatter(formato)
        logger.addHandler(archivo)
        logger.debug("Log guardado en: %s", carpeta_log / nombre_log)


def resolver_path_maestro(path_maestro: Path) -> Path:
    if path_maestro.is_dir():
        candidato = path_maestro / NOMBRE_MAESTRO_DEFAULT
        if not candidato.exists():
            raise FileNotFoundError(
                f"No se encontró '{NOMBRE_MAESTRO_DEFAULT}' dentro de {path_maestro}. "
                f"Pasa la ruta directa al archivo con --path_maestro."
            )
        return candidato
    if not path_maestro.exists():
        raise FileNotFoundError(f"--path_maestro no existe: {path_maestro}")
    return path_maestro


def parsear_argumentos(argv=None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Calcula el IE a partir del Excel Maestro, en 3 hojas (replicando la "
                    "estructura de PE Experiencia SGR Cálculo.xlsx), formateadas como "
                    "Tablas de Excel para poder auditarlo paso a paso."
    )
    parser.add_argument(
        "--path_maestro", required=True, type=Path,
        help="Ruta al EXCEL_MAESTRO_ICS.xlsx, o carpeta que lo contiene.",
    )
    parser.add_argument(
        "--output", "-o", type=Path, default=None,
        help=f"Ruta/nombre del excel de salida (default: {NOMBRE_SALIDA_DEFAULT} "
             f"en la misma carpeta del maestro).",
    )
    parser.add_argument(
        "--fecha_corte", type=str, default=None,
        help="Fecha de corte para la antigüedad, formato YYYY-MM-DD (default: hoy).",
    )
    parser.add_argument(
        "--verbose", "-v", action="store_true",
        help="Imprime la explicación de cada fórmula/variable y el recuento por nivel de riesgo.",
    )
    parser.add_argument(
        "--log", action="store_true",
        help="Guarda un archivo .log con el detalle de la ejecución en la carpeta del maestro.",
    )
    return parser.parse_args(argv)


def main(argv=None) -> None:
    args = parsear_argumentos(argv)

    path_maestro = resolver_path_maestro(args.path_maestro.expanduser().resolve())
    carpeta = path_maestro.parent
    salida = args.output.expanduser().resolve() if args.output else carpeta / NOMBRE_SALIDA_DEFAULT
    fecha_corte = pd.Timestamp(args.fecha_corte) if args.fecha_corte else pd.Timestamp.now().normalize()

    configurar_logging(verbose=args.verbose, guardar_log=args.log, carpeta_log=carpeta)

    if args.verbose:
        explicar_metodologia()
        logger.info("Fecha de corte usada para antigüedad: %s", fecha_corte.date())

    ejecutores, proyectos = leer_maestro(path_maestro)
    hojas, n_sin_historial, _ = calcular_ie_por_pasos(ejecutores, proyectos, fecha_corte)

    if args.verbose:
        logger.info("Calculando IE sectorizado (filtro puro por sector, ver docstring)...")
    hoja_sector = calcular_ie_sectorizado(ejecutores, proyectos, fecha_corte)
    hojas["4_IE_Sectorial"] = hoja_sector
    hojas["5_Matriz_IE_x_Sector"] = armar_matriz_ie_sector(hoja_sector)

    escribir_excel_por_pasos(salida, hojas)

    if args.verbose:
        imprimir_resumen(hojas["3_Indice_Riesgo"], n_sin_historial)
        logger.info(
            "IE Sectorial: %s sectores x %s ejecutores = %s filas en 4_IE_Sectorial.",
            hoja_sector["sector"].nunique(), hoja_sector["codigo_ejecutor"].nunique(), len(hoja_sector),
        )

    print(f"Metodología IE generada en: {salida} (5 hojas: 3 del IE global + IE sectorial + matriz)")


if __name__ == "__main__":
    main()