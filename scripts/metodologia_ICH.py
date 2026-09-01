"""
metodologia_ICH_v2.py
========================
Segunda version del Indice de Cumplimiento Historico (ICH), calibrada tras
la auditoria de justicia hecha con Daniel sobre el ICH oficial
(metodologia_ICH.py). Los pasos 1-4 (TCP_i, Pen_i, peso_i, Base_e) NO
cambian -- son identicos al oficial. Lo que cambia es como se combinan
Base_e con las incidencias de valor y numero de proyectos.

QUE CAMBIA Y POR QUE (resumen de la calibracion, con evidencia numerica
sobre los 1.990 ejecutores del maestro real)
------------------------------------------------------------------------
1) DE PRODUCTO A SUMA PONDERADA (pedido explicito de Daniel, y coherente
   con IE/ICCI/IMA que ya son sumas ponderadas en este proyecto):
   el ICH oficial es Base_e x Inc_Valor x Inc_N (producto). Un ejecutor
   con cumplimiento mediocre pero MUY grande en valor/n_proyectos puede
   terminar con el ICH mas alto de su grupo -- caso real: ejecutor
   6666948, Base_e=0.57 (percentil 64.6% SOLO en cumplimiento) pero
   ich_percentil=1.0 (el MEJOR de 582 ejecutores de su grupo) en la
   formula oficial.
   Se probaron pesos 70/15/15, 60/20/20 y 50/25/25 con la prueba de
   "pares invertidos" (dos ejecutores del mismo grupo, uno cumple >=15pp
   peor que el otro: ¿el de peor cumplimiento termina con mejor puntaje
   final solo por tamaño?):
       Oficial (producto):        28.6% de pares invertidos
       Lineal 70% Base_e/15%/15%:  3.6% de pares invertidos  <- ADOPTADA
       Lineal 60%/20%/20%:         8.7%
       Lineal 50%/25%/25%:        15.7%
   70/15/15 fue la que mejor reflejo la jerarquia pedida por Daniel:
   "Base_e es lo mas importante (variable de cumplimiento real), las
   incidencias son un reconocimiento adicional, no el motor principal".

2) ACOTAMIENTO POR MIN-MAX EN VEZ DE PERCENTIL:
   Base_e ya vive en [0,1], pero incidencia_valor e incidencia_n
   (logaritmos) no tienen techo natural. Para sumarlos con pesos hace
   falta llevarlos a la misma escala. Se probaron 4 formas (percentil,
   x/(1+x), min-max, sigmoide-MAD), todas con pesos 70/15/15, midiendo
   el mismo % de pares invertidos:
       Percentil:            3.57%
       x/(1+x):               2.54%
       Min-max dentro grupo:  0.44%  <- ADOPTADA (la mas justa, por lejos)
       Sigmoide (MAD):         4.79%
   Min-max preserva la MAGNITUD real de la diferencia (el doble de
   proyectos pesa el doble, no "un peldaño mas" como con percentil).

3) DOS GRUPOS DE COMPARACION DISTINTOS (hallazgo de la verificacion de
   grupos): capacidad_institucional es un buen grupo de comparacion para
   Base_e (el cumplimiento si depende de la capacidad del ejecutor), pero
   NO es un proxy valido de tamaño de portafolio -- la correlacion de
   Spearman entre capacidad_institucional y ve_e es -0.14, y con
   total_proyectos es 0.16 (practicamente nula). Del grupo de capacidad
   institucional MAS BAJA, 209 de 582 ejecutores estan en el cuartil de
   valor MAS ALTO de todo el universo -- son "chicos" en capacidad pero
   manejan algunos de los presupuestos mas grandes del sistema.
   Por eso:
       - base_e_norm  se normaliza (min-max) DENTRO de capacidad_institucional
       - valor_norm y n_norm se normalizan (min-max) DENTRO de un grupo
         de TAMAÑO DE PORTAFOLIO (grupo_tamano_portafolio, cuartiles T1-T4
         combinando rank de ve_e y de total_proyectos), NO capacidad_institucional.

FORMULA FINAL
--------------
    Por proyecto i (identico al oficial):
        TCP_i = periodos_cumplidos_i / periodos_evaluados_i   (tau=10%)
        Pen_i = 1 / (1 + p * ln(1 + R_i))                      (p=0.50)
        peso_i = Valor_i / Ve_e

    Por ejecutor e:
        Base_e = SUMA_i (TCP_i x Pen_i x peso_i)                (identico al oficial)

        razon_valor_e = Ve_e                     (valor total ejecutado)
        razon_n_e     = total_proyectos_e         (En Ejecucion + Terminado)

        grupo_capacidad = capacidad_institucional
        grupo_tamano    = cuartil de (rank(ve_e) + rank(total_proyectos))/2,
                          calculado sobre TODO el universo evaluable
                          (T1_pequeño .. T4_grande)

        base_e_norm  = (Base_e - min) / (max - min)         DENTRO de grupo_capacidad
        valor_norm   = (razon_valor - min) / (max - min)    DENTRO de grupo_tamano
        n_norm       = (razon_n - min) / (max - min)        DENTRO de grupo_tamano

        ICH_e = 0.70 * base_e_norm + 0.15 * valor_norm + 0.15 * n_norm

    Normalizacion final para riesgo (igual criterio que el oficial):
        ich_percentil = rank(pct=True) de ICH_e DENTRO de grupo_capacidad
        puntaje_riesgo = (1 - ich_percentil) * 100
        nivel_riesgo: <33 Bajo | 33-67 Medio | >=67 Alto

PERIODOS SIN ACTIVIDAD, FECHA DE CORTE, ESTADOS VALIDOS, EJECUTORES SIN
DATOS: identico al ICH oficial (ver metodologia_ICH.py) -- no se repite
esa logica aqui, se reutiliza importando el modulo.

SALIDA -- 9 HOJAS (2 mas que el oficial, para las nuevas normalizaciones)
---------------------------------------------------------------------------
    0_Detalle_Periodos     -> igual al oficial (detalle crudo por periodo)
    1_TCP_i                -> igual al oficial
    2_Pen_i                -> igual al oficial
    3_Peso_i               -> igual al oficial
    4_Base_e               -> igual al oficial
    5_Grupo_Tamano         -> NUEVA: ve_e, total_proyectos, rank, indice de
                              tamaño y grupo_tamano_portafolio (T1-T4) por
                              ejecutor -- para poder auditar a que grupo de
                              tamaño quedo asignado cada uno.
    6_Normalizacion        -> NUEVA: base_e_norm, valor_norm, n_norm (min-max
                              dentro de cada grupo respectivo) y el ICH_e
                              resultante de la suma ponderada 70/15/15.
    7_Resultado_ICH        -> tabla final consolidada con percentil, puntaje
                              y nivel de riesgo (mismo formato/colores que
                              el oficial).

USO
----
    python metodologia_ICH_v2.py --path_maestro ruta/EXCEL_MAESTRO_ICS.xlsx
    python metodologia_ICH_v2.py --path_maestro ruta/carpeta --verbose
    python metodologia_ICH_v2.py --path_maestro ruta/carpeta --output Metodologia_ICH_v2.xlsx
"""

from __future__ import annotations

import argparse
import logging
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Reutiliza TODO lo que no cambia del ICH oficial: lectura del maestro,
# filtros de estado/fecha de corte, calculo de TCP_i/Pen_i/peso_i/Base_e,
# detalle por periodo, helpers de escritura de Excel.
import metodologia_ICH as ich_v1

# =============================================================================
# PARAMETROS METODOLOGICOS DE LA V2 (ver docstring)
# =============================================================================

PESO_BASE_E = 0.70   # HIPERPARAMETRO -- validado via prueba de pares invertidos
PESO_VALOR = 0.15    # HIPERPARAMETRO
PESO_N = 0.15        # HIPERPARAMETRO
GRUPO_CAPACIDAD = ich_v1.GRUPO_COMPARACION       # "capacidad_institucional", para base_e_norm
GRUPO_TAMANO = "grupo_tamano_portafolio"          # nuevo grupo, para valor_norm y n_norm
UMBRAL_RIESGO_BAJO = ich_v1.UMBRAL_RIESGO_BAJO
UMBRAL_RIESGO_MEDIO = ich_v1.UMBRAL_RIESGO_MEDIO
ESTADOS_PARA_N = ich_v1.ESTADOS_PARA_N

NOMBRE_SALIDA_DEFAULT = "Metodologia_ICH_v2.xlsx"

logger = logging.getLogger("metodologia_ICH_v2")


# =============================================================================
# HOJA 5: GRUPO DE TAMAÑO DE PORTAFOLIO
# =============================================================================

def calcular_grupo_tamano(df_peso: pd.DataFrame, df_proyectos: pd.DataFrame, universo: pd.DataFrame) -> pd.DataFrame:
    """
    Construye, para CADA ejecutor del universo (no solo los evaluables en
    Base_e), un indice de tamaño de portafolio = promedio del rank
    percentual de ve_e y de total_proyectos, y lo corta en 4 cuartiles
    (T1_pequeño .. T4_grande). Este grupo reemplaza a
    capacidad_institucional SOLO para normalizar incidencia_valor e
    incidencia_n (ver docstring del modulo: capacidad_institucional no
    correlaciona con tamaño real de portafolio).
    """
    ve_e_por_ejecutor = df_peso[["codigo_ejecutor", "ve_e"]].drop_duplicates()
    n_proy = ich_v1.calcular_num_proyectos(df_proyectos, estados=ESTADOS_PARA_N)

    tabla = universo[["codigo_ejecutor"]].drop_duplicates().merge(
        ve_e_por_ejecutor, on="codigo_ejecutor", how="left"
    ).merge(n_proy, on="codigo_ejecutor", how="left")
    tabla["ve_e"] = tabla["ve_e"].fillna(0.0)
    tabla["n_proyectos"] = tabla["n_proyectos"].fillna(0)

    tabla["rank_valor_pct"] = tabla["ve_e"].rank(pct=True)
    tabla["rank_n_pct"] = tabla["n_proyectos"].rank(pct=True)
    tabla["indice_tamano"] = (tabla["rank_valor_pct"] + tabla["rank_n_pct"]) / 2
    tabla[GRUPO_TAMANO] = pd.qcut(
        tabla["indice_tamano"], 4, labels=["T1_pequeno", "T2", "T3", "T4_grande"]
    )
    return tabla[[
        "codigo_ejecutor", "ve_e", "n_proyectos", "rank_valor_pct", "rank_n_pct",
        "indice_tamano", GRUPO_TAMANO,
    ]].sort_values("codigo_ejecutor").reset_index(drop=True)


# =============================================================================
# HOJA 6: NORMALIZACION MIN-MAX + COMBINACION LINEAL
# =============================================================================

def minmax_dentro_de_grupo(serie: pd.Series, grupo: pd.Series) -> pd.Series:
    df_tmp = pd.DataFrame({"valor": serie, "grupo": grupo})
    gmin = df_tmp.groupby("grupo", dropna=False)["valor"].transform("min")
    gmax = df_tmp.groupby("grupo", dropna=False)["valor"].transform("max")
    rango = (gmax - gmin).replace(0, np.nan)
    return ((df_tmp["valor"] - gmin) / rango).fillna(0.0)


def calcular_normalizacion_y_ich(
    hoja5_base_e: pd.DataFrame, hoja5_grupo_tamano: pd.DataFrame, ejecutores: pd.DataFrame,
) -> pd.DataFrame:
    """
    hoja5_base_e: codigo_ejecutor, base_e (de calcular_detalle_base_e de la v1)
    hoja5_grupo_tamano: salida de calcular_grupo_tamano (ve_e, n_proyectos,
                        grupo_tamano_portafolio) para TODO el universo
    ejecutores: hoja Ejecutores del maestro (para capacidad_institucional)
    """
    universo_cap = ejecutores[["codigo_ejecutor", GRUPO_CAPACIDAD]].drop_duplicates()

    tabla = universo_cap.merge(hoja5_grupo_tamano, on="codigo_ejecutor", how="left")
    tabla = tabla.merge(hoja5_base_e, on="codigo_ejecutor", how="left")
    tabla["base_e"] = tabla["base_e"].fillna(0.0)

    tabla["base_e_norm"] = minmax_dentro_de_grupo(tabla["base_e"], tabla[GRUPO_CAPACIDAD])
    tabla["valor_norm"] = minmax_dentro_de_grupo(tabla["ve_e"], tabla[GRUPO_TAMANO])
    tabla["n_norm"] = minmax_dentro_de_grupo(tabla["n_proyectos"], tabla[GRUPO_TAMANO])

    tabla["ich_e"] = (
        PESO_BASE_E * tabla["base_e_norm"]
        + PESO_VALOR * tabla["valor_norm"]
        + PESO_N * tabla["n_norm"]
    )

    return tabla[[
        "codigo_ejecutor", GRUPO_CAPACIDAD, GRUPO_TAMANO,
        "base_e", "base_e_norm", "ve_e", "valor_norm", "n_proyectos", "n_norm",
        "ich_e",
    ]].sort_values("codigo_ejecutor").reset_index(drop=True)


# =============================================================================
# HOJA 7: RESULTADO FINAL (percentil, puntaje, nivel de riesgo)
# =============================================================================

def calcular_resultado_final(hoja6: pd.DataFrame, base_e_df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    sin_datos_mask = ~hoja6["codigo_ejecutor"].isin(base_e_df["codigo_ejecutor"])
    n_sin_datos = int(sin_datos_mask.sum())

    resultado = hoja6.copy()
    tam_grupo = resultado.groupby(GRUPO_CAPACIDAD, dropna=False)["codigo_ejecutor"].transform("count")
    grupos_solitarios = sorted(
        resultado.loc[tam_grupo == 1, GRUPO_CAPACIDAD].dropna().unique().tolist()
    )
    if grupos_solitarios:
        logger.warning(
            "ALERTA — %s ejecutor(es) están solos en su grupo de %s (sin nadie "
            "con quien compararse). Su percentil sera automáticamente 1.0 "
            "(Riesgo Bajo) sin importar su ich_e real. Grupos afectados: %s",
            int((tam_grupo == 1).sum()), GRUPO_CAPACIDAD, grupos_solitarios,
        )

    resultado["ich_percentil"] = resultado.groupby(GRUPO_CAPACIDAD, dropna=False)["ich_e"].rank(
        pct=True, method="average"
    )
    resultado.loc[sin_datos_mask.values, "ich_percentil"] = 0.0
    resultado["puntaje_riesgo"] = (1 - resultado["ich_percentil"]) * 100

    condiciones = [
        resultado["ich_percentil"] >= UMBRAL_RIESGO_BAJO,
        resultado["ich_percentil"] >= UMBRAL_RIESGO_MEDIO,
    ]
    niveles = ["Riesgo Bajo", "Riesgo Medio"]
    resultado["nivel_riesgo"] = np.select(condiciones, niveles, default="Riesgo Alto")
    resultado = resultado.sort_values("puntaje_riesgo").reset_index(drop=True)

    columnas = [
        "codigo_ejecutor", GRUPO_CAPACIDAD, GRUPO_TAMANO, "base_e", "base_e_norm",
        "ve_e", "valor_norm", "n_proyectos", "n_norm", "ich_e",
        "ich_percentil", "puntaje_riesgo", "nivel_riesgo",
    ]
    return resultado[columnas], n_sin_datos


# =============================================================================
# PIPELINE COMPLETO -> 8 HOJAS (0 a 7)
# =============================================================================

def calcular_ich_v2_por_pasos(ejecutores, proyectos, periodos, reprogramaciones):
    """
    proyectos/periodos deben llegar YA FILTRADOS (ver ich_v1.filtrar_proyectos_validos
    y ich_v1.filtrar_periodos_por_fecha_corte), igual que en la v1.
    """
    # Pasos 1-4: identicos al ICH oficial, se reutilizan tal cual.
    hoja1_tcp = ich_v1.calcular_tcp_por_proyecto(periodos)

    df_pen = ich_v1.calcular_pen_por_proyecto(reprogramaciones)
    hoja2_pen = df_pen.merge(proyectos[["bpin", "codigo_ejecutor"]].drop_duplicates(), on="bpin", how="left")
    hoja2_pen["p_suavidad"] = ich_v1.P_SUAVIDAD
    hoja2_pen = hoja2_pen[["codigo_ejecutor", "bpin", "reprogramaciones_no_permitidas", "p_suavidad", "pen_i"]]
    hoja2_pen = hoja2_pen.sort_values(["codigo_ejecutor", "bpin"]).reset_index(drop=True)

    df_peso, n_sin_valor = ich_v1.calcular_peso_por_proyecto(proyectos)
    if n_sin_valor:
        logger.warning(
            "ALERTA — %s proyectos no tienen 'valor_total_proyecto' válido y "
            "quedaron fuera de ve_e/peso_i/Base_e.", n_sin_valor,
        )
    hoja3_peso = df_peso.sort_values(["codigo_ejecutor", "bpin"]).reset_index(drop=True)

    hoja4_base, base_e_df = ich_v1.calcular_detalle_base_e(hoja1_tcp, df_pen, df_peso)

    # Paso 5 (NUEVO): grupo de tamaño de portafolio, sobre TODO el universo
    # de ejecutores (no solo los evaluables), para que ningun ejecutor
    # quede sin grupo asignado.
    hoja5_grupo_tamano = calcular_grupo_tamano(df_peso, proyectos, ejecutores)

    # Paso 6 (NUEVO): normalizacion min-max dentro de cada grupo + ICH_e
    hoja6_normalizacion = calcular_normalizacion_y_ich(base_e_df, hoja5_grupo_tamano, ejecutores)

    # Paso 7: resultado final (percentil/puntaje/nivel) — mismo criterio que la v1
    hoja7_resultado, n_sin_datos = calcular_resultado_final(hoja6_normalizacion, base_e_df)

    hojas = {
        "1_TCP_i": hoja1_tcp,
        "2_Pen_i": hoja2_pen,
        "3_Peso_i": hoja3_peso,
        "4_Base_e": hoja4_base,
        "5_Grupo_Tamano": hoja5_grupo_tamano,
        "6_Normalizacion": hoja6_normalizacion,
        "7_Resultado_ICH": hoja7_resultado,
    }
    return hojas, n_sin_datos, proyectos


# =============================================================================
# ESCRITURA DEL EXCEL (reutiliza los helpers de la v1 tal cual)
# =============================================================================

def _armar_instrucciones_detalle(periodos, proyectos, resultado_para_banner, reprogramaciones):
    """Construye, en un solo barrido, la lista de instrucciones (banner /
    row / subtotal) para la hoja 0_Detalle_Periodos, EN VEZ de escribir
    celda por celda directamente -- ver nota de rendimiento en
    escribir_excel_v2 sobre por que esto importa.

    NOTA DE RENDIMIENTO (2): la version original recorria cada proyecto
    de cada ejecutor con dos groupby anidados + .iterrows() interno
    (~526,000 filas de periodo) para calcular pct_ejecucion, acumulado y
    el texto de TCP fila por fila -- eso tardaba ~15s solo en construir
    la lista de instrucciones. Aqui esas mismas columnas se calculan
    VECTORIZADAS con numpy/pandas (cumsum por bpin, mascara de "ultima
    fila del proyecto", etc.) y solo se hace UN recorrido final con
    itertuples (mucho mas rapido que iterrows) para armar las tuplas de
    instruccion -- baja el tiempo a ~1s sobre los mismos ~554,000 datos.
    """
    detalle = ich_v1.armar_detalle_periodos(periodos, proyectos)

    detalle["pct_ejec"] = np.where(
        detalle["valor_programado"] != 0,
        detalle["valor_ejecutado"] / detalle["valor_programado"],
        np.nan,
    )
    detalle["acumulado"] = detalle.groupby("bpin")["valor_programado"].cumsum()
    is_last_bpin = (detalle.groupby("bpin").cumcount(ascending=False) == 0).to_numpy()
    detalle["valor_proyecto_final"] = np.where(
        is_last_bpin & detalle["valor_total_proyecto"].notna().to_numpy(),
        detalle["valor_total_proyecto"], np.nan,
    )
    cumple_obj = detalle["cumple"].astype(object).where(detalle["cumple"].notna(), None)
    detalle["tcp_texto"] = [
        "Sin actividad (excluido)" if v is None else ("Cumple" if v else "No Cumple")
        for v in cumple_obj
    ]
    is_first_ejec = (detalle.groupby("codigo_ejecutor").cumcount() == 0).to_numpy()

    reprog_por_bpin = reprogramaciones.set_index("bpin")["reprogramaciones_no_permitidas"]
    bpin_ejec = detalle[["codigo_ejecutor", "bpin"]].drop_duplicates()
    bpin_ejec["reprog"] = bpin_ejec["bpin"].map(reprog_por_bpin).fillna(0)
    r_total_por_ejec = bpin_ejec.groupby("codigo_ejecutor")["reprog"].sum()

    cumplidos_por_ejec = detalle.groupby("codigo_ejecutor")["cumple"].sum()
    evaluados_por_ejec = detalle["cumple"].notna().groupby(detalle["codigo_ejecutor"]).sum()
    bpin_valor = detalle.groupby("bpin")["valor_total_proyecto"].first().fillna(0.0)
    resumen = resultado_para_banner.set_index("codigo_ejecutor")

    cols = [
        "codigo_ejecutor", "bpin", "periodo", "valor_programado", "valor_ejecutado",
        "pct_ejec", "tcp_texto", "desviacion", "acumulado", "valor_proyecto_final",
    ]
    is_first_list = is_first_ejec.tolist()
    is_last_list = is_last_bpin.tolist()

    instrucciones = []
    for i, row in enumerate(detalle[cols].itertuples(index=False, name=None)):
        (codigo, bpin, periodo, prog, ejec, pct_ejec, tcp_texto,
         desviacion, acumulado, valor_final) = row
        if is_first_list[i]:
            if codigo in resumen.index:
                r = resumen.loc[codigo]
                n_val = int(r["total_proyectos"]) if pd.notna(r["total_proyectos"]) else 0
                ve_val = float(r["ve_e"]) if pd.notna(r["ve_e"]) else 0.0
                ich_val = float(r["ich"]) if pd.notna(r["ich"]) else 0.0
                nivel_val = r["nivel_riesgo"]
            else:
                n_val, ve_val, ich_val, nivel_val = 0, 0.0, 0.0, "Sin resultado"
            cum = cumplidos_por_ejec.get(codigo, 0)
            evalu = evaluados_por_ejec.get(codigo, 0)
            r_total = int(r_total_por_ejec.get(codigo, 0))
            pct_tcp = (cum / evalu * 100) if evalu else 0.0
            texto_banner = (
                f"Ejecutor {codigo}   |   N={n_val}   V_e=${ve_val/1e6:,.1f}M   "
                f"R={r_total}   TCP={int(cum)}/{evalu}={pct_tcp:.1f}%   "
                f"ICH={ich_val:.4f}   Riesgo: {nivel_val}"
            )
            instrucciones.append(("banner", texto_banner))

        prog_f = float(prog) if pd.notna(prog) else 0.0
        ejec_f = float(ejec) if pd.notna(ejec) else 0.0
        pe = None if pd.isna(pct_ejec) else float(pct_ejec)
        dev = None if pd.isna(desviacion) else float(desviacion)
        vf = None if pd.isna(valor_final) else float(valor_final)
        instrucciones.append((
            "row", codigo, bpin, periodo, prog_f, ejec_f, pe,
            tcp_texto, dev, float(acumulado), vf,
        ))

        if is_last_list[i]:
            valor_proy = float(bpin_valor.get(bpin, 0.0))
            instrucciones.append(("subtotal", bpin, valor_proy))

    return instrucciones


def escribir_excel_v2(
    salida: Path,
    hojas: dict[str, pd.DataFrame],
    periodos: pd.DataFrame,
    proyectos: pd.DataFrame,
    reprogramaciones: pd.DataFrame,
) -> None:
    """
    NOTA DE RENDIMIENTO: la hoja 0_Detalle_Periodos tiene ~550,000 filas
    (una por bpin+periodo). Escribirla con la API "normal" de openpyxl
    celda por celda (ws.cell(row=..., column=..., value=...)) es viable,
    pero usar ws.merge_cells() para las ~28,000 filas de banner/subtotal
    NO lo es: esa funcion revisa solapamiento contra TODOS los rangos ya
    combinados en cada llamada (costo O(n) por llamada => O(n^2) total),
    y con >20,000 merges se vuelve impracticable (probado: pasa de
    segundos a *varios minutos*). La solucion es insertar los rangos
    directamente en el set interno `ws.merged_cells.ranges` (con .add),
    evitando el chequeo de solapamiento -- es seguro aqui porque los
    rangos que construimos (una fila completa por banner/subtotal) nunca
    se solapan entre si por construccion.
    """
    from openpyxl import Workbook
    from openpyxl.worksheet.cell_range import CellRange

    resultado_para_banner = hojas["7_Resultado_ICH"].rename(
        columns={"n_proyectos": "total_proyectos", "ich_e": "ich"}
    )
    instrucciones = _armar_instrucciones_detalle(
        periodos, proyectos, resultado_para_banner, reprogramaciones
    )

    FILL_ENCABEZADO = ich_v1.FILL_ENCABEZADO
    FILL_BANNER_EJECUTOR = ich_v1.FILL_BANNER_EJECUTOR
    FILL_SUBTOTAL = ich_v1.FILL_SUBTOTAL
    FONT_HEADER = Font(bold=True, color="FFFFFF")
    FONT_BANNER = Font(bold=True, color="FFFFFF")
    FONT_SUBTOTAL = Font(italic=True, bold=True)

    encabezados = [
        "codigo_ejecutor", "bpin", "periodo", "valor_programado", "valor_ejecutado",
        "pct_ejecucion", "pct_cumplimiento", "tcp", "desviacion_pct",
        "prog_acumulado", "valor_proyecto",
    ]
    n_cols = len(encabezados)

    wb = Workbook()
    ws = wb.active
    ws.title = "0_Detalle_Periodos"
    ws.append(encabezados)

    fila = 1
    filas_banner = []
    filas_subtotal = []
    for instr in instrucciones:
        tipo = instr[0]
        fila += 1
        if tipo == "banner":
            _, texto_banner = instr
            ws.append([texto_banner] + [None] * (n_cols - 1))
            filas_banner.append(fila)
        elif tipo == "row":
            (_, codigo, bpin, periodo, prog, ejec, pct_ejec,
             tcp_texto, desviacion, acumulado, valor_proyecto_final) = instr
            ws.append([codigo, bpin, periodo, prog, ejec, pct_ejec, pct_ejec,
                       tcp_texto, desviacion, acumulado, valor_proyecto_final])
        elif tipo == "subtotal":
            _, bpin, valor_proy = instr
            texto_subtotal = (
                f"SUBTOTAL {bpin}   Valor proyecto (SUMA programados) = "
                f"${valor_proy/1e6:,.1f}M"
            )
            ws.append([texto_subtotal] + [None] * (n_cols - 2) + [valor_proy])
            filas_subtotal.append(fila)

    for i in range(1, n_cols + 1):
        c = ws.cell(row=1, column=i)
        c.font = FONT_HEADER
        c.fill = FILL_ENCABEZADO

    for fb in filas_banner:
        c = ws.cell(row=fb, column=1)
        c.font = FONT_BANNER
        c.fill = FILL_BANNER_EJECUTOR
        ws.merged_cells.ranges.add(CellRange(min_col=1, min_row=fb, max_col=n_cols, max_row=fb))

    for fs in filas_subtotal:
        c_sub = ws.cell(row=fs, column=1)
        c_sub.font = FONT_SUBTOTAL
        c_sub.fill = FILL_SUBTOTAL
        c_val = ws.cell(row=fs, column=n_cols)
        c_val.fill = FILL_SUBTOTAL
        c_val.number_format = "$#,##0"
        ws.merged_cells.ranges.add(
            CellRange(min_col=1, min_row=fs, max_col=n_cols - 1, max_row=fs)
        )

    anchos = [16, 18, 14, 15, 15, 11, 18, 22, 14, 15, 17]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.freeze_panes = "A2"

    def _limpiar_na(v):
        try:
            if pd.isna(v):
                return None
        except (TypeError, ValueError):
            pass
        return v

    def _escribir_hoja_tabla_directo(wb, df, nombre_hoja):
        ws2 = wb.create_sheet(nombre_hoja)
        ws2.append(list(df.columns))
        for row in df.itertuples(index=False, name=None):
            ws2.append([_limpiar_na(v) for v in row])

        n_filas, n_colsx = df.shape
        if n_filas == 0 or n_colsx == 0:
            return ws2

        ultima_col = get_column_letter(n_colsx)
        rango = f"A1:{ultima_col}{n_filas + 1}"
        tabla = Table(displayName=ich_v1._nombre_tabla_valido(nombre_hoja), ref=rango)
        tabla.tableStyleInfo = TableStyleInfo(
            name=ich_v1.ESTILO_TABLA, showRowStripes=True, showColumnStripes=False,
            showFirstColumn=False, showLastColumn=False,
        )
        ws2.add_table(tabla)

        for i, col in enumerate(df.columns, start=1):
            muestra = df[col].head(200).tolist()
            ancho = max([len(str(col))] + [len(str(v)) for v in muestra]) + 2
            ws2.column_dimensions[get_column_letter(i)].width = min(ancho, 40)
        return ws2

    for nombre_hoja, df in hojas.items():
        ws_creada = _escribir_hoja_tabla_directo(wb, df, nombre_hoja)
        if nombre_hoja == "7_Resultado_ICH":
            ich_v1.colorear_columnas_riesgo(ws_creada, df)

    wb.save(salida)


# =============================================================================
# RESUMEN --verbose
# =============================================================================

def explicar_metodologia_v2() -> None:
    logger.info("=" * 70)
    logger.info("METODOLOGÍA — Índice de Cumplimiento Histórico v2 (ICH v2)")
    logger.info("=" * 70)
    logger.info(
        "Pasos 1-4 (TCP_i, Pen_i, peso_i, Base_e): IDÉNTICOS al ICH oficial "
        "(tau=%.0f%%, p=%.2f).", ich_v1.TAU_TOLERANCIA * 100, ich_v1.P_SUAVIDAD,
    )
    logger.info(
        "Paso 5 (NUEVO): grupo_tamano_portafolio = cuartil del promedio de "
        "rank(ve_e) y rank(total_proyectos), calculado sobre TODO el "
        "universo -- reemplaza a capacidad_institucional SOLO para "
        "normalizar valor/n (correlación con tamaño real: ~0)."
    )
    logger.info(
        "Paso 6 (NUEVO): base_e_norm = min-max DENTRO de capacidad_institucional; "
        "valor_norm y n_norm = min-max DENTRO de grupo_tamano_portafolio. "
        "ICH_e = %.2f*base_e_norm + %.2f*valor_norm + %.2f*n_norm (suma "
        "ponderada, ya NO es un producto).", PESO_BASE_E, PESO_VALOR, PESO_N,
    )
    logger.info(
        "Paso 7: ich_percentil = rank(pct=True) de ICH_e DENTRO de "
        "capacidad_institucional. puntaje_riesgo = (1-percentil)*100. "
        "Umbrales: <33 Bajo (percentil>%.2f) | 33-67 Medio | >=67 Alto.",
        UMBRAL_RIESGO_BAJO,
    )


def imprimir_resumen_v2(hoja7: pd.DataFrame, n_sin_datos: int) -> None:
    conteo = hoja7["nivel_riesgo"].value_counts()
    logger.info("=" * 70)
    logger.info("RESUMEN — ICH v2")
    logger.info("Ejecutores evaluados: %s (sin datos evaluables: %s)", len(hoja7), n_sin_datos)
    for nivel in ["Riesgo Bajo", "Riesgo Medio", "Riesgo Alto"]:
        logger.info("  %s: %s", nivel, int(conteo.get(nivel, 0)))
    logger.info("Distribución por grupo_tamano_portafolio:")
    logger.info("%s", hoja7["grupo_tamano_portafolio"].value_counts().to_string())


# =============================================================================
# CLI
# =============================================================================

def parsear_argumentos(argv=None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Calcula el ICH v2 (suma ponderada 70/15/15, min-max por grupo, dos "
                    "grupos de comparación) a partir del Excel Maestro, en 8 hojas."
    )
    parser.add_argument(
        "--path_maestro", required=True, type=Path,
        help="Ruta al EXCEL_MAESTRO_ICS.xlsx, o carpeta que lo contiene.",
    )
    parser.add_argument(
        "--output", "-o", type=Path, default=None,
        help=f"Ruta/nombre del excel de salida (default: {NOMBRE_SALIDA_DEFAULT} "
             f"en la misma carpeta del maestro).",
    )
    parser.add_argument(
        "--fecha_corte", type=str, default=None,
        help="Fecha de corte YYYY-MM-DD para Periodos (default: igual a la v1, "
             f"{ich_v1.FECHA_CORTE_ICH.date()}).",
    )
    parser.add_argument("--verbose", "-v", action="store_true")
    parser.add_argument("--log", action="store_true")
    return parser.parse_args(argv)


def main(argv=None) -> None:
    args = parsear_argumentos(argv)

    path_maestro = ich_v1.resolver_path_maestro(args.path_maestro.expanduser().resolve())
    carpeta = path_maestro.parent
    salida = args.output.expanduser().resolve() if args.output else carpeta / NOMBRE_SALIDA_DEFAULT
    fecha_corte = pd.Timestamp(args.fecha_corte) if args.fecha_corte else ich_v1.FECHA_CORTE_ICH

    ich_v1.configurar_logging(verbose=args.verbose, guardar_log=args.log, carpeta_log=carpeta)
    logger.setLevel(logging.DEBUG)
    logger.handlers = ich_v1.logger.handlers  # comparte los mismos handlers configurados arriba

    if args.verbose:
        explicar_metodologia_v2()

    ejecutores, proyectos, periodos, reprogramaciones = ich_v1.leer_maestro(path_maestro)
    proyectos, periodos = ich_v1.filtrar_proyectos_validos(proyectos, periodos)
    periodos = ich_v1.filtrar_periodos_por_fecha_corte(periodos, fecha_corte)
    hojas, n_sin_datos, proyectos = calcular_ich_v2_por_pasos(ejecutores, proyectos, periodos, reprogramaciones)

    escribir_excel_v2(salida, hojas, periodos, proyectos, reprogramaciones)

    if args.verbose:
        imprimir_resumen_v2(hojas["7_Resultado_ICH"], n_sin_datos)

    print(f"Metodología ICH v2 generada en: {salida} (8 hojas: detalle por periodo + 7 pasos del cálculo)")


if __name__ == "__main__":
    main()
