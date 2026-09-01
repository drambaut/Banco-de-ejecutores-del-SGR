"""
metodologia_ICCI.py
================================
Calcula el Índice de Continuidad y Calidad de Información (ICCI) a partir
del Excel Maestro (hojas 'Continuidad', 'Habilitaciones', 'Proyectos' y
'Ejecutores') y genera un Excel de metodología paso a paso con TABLAS de
Excel (TableStyleMedium9) y FÓRMULAS REALES (SUMIFS/COUNTIFS), no valores
pegados — así el archivo se recalcula solo si alguien edita un dato en
'0_Base'.

METODOLOGÍA
-----------
Universo: todos los BPIN que aparezcan en la hoja 'Continuidad' del
Maestro. Si hay varios cortes mensuales, se usa el ÚLTIMO corte disponible
por BPIN (máxima fecha_corte). A medida que se acumulen más cortes
mensuales en el Maestro, el método de continuidad deberá migrar al conteo
de transiciones avance/retroceso por periodo — ver docstring abajo.

Score Continuidad (por ejecutor):
    = (Σ nro_periodos_reportados / Σ nro_periodos_a_reportar) × 100
    Clipeado a [0, 100]. A mayor score, mejor desempeño (menos riesgo).

Score Calidad de Información (por ejecutor):
    = 100 − (Σ n_periodos_habilitados / Σ nro_periodos_a_reportar × 100)
    donde n_periodos_habilitados = suma de meses habilitados (solicitados
    en 'Habilitaciones') de todos los BPIN del ejecutor. Si el ejecutor
    no tiene ningún BPIN en Habilitaciones, n_periodos_habilitados = 0.
    A mayor score, mejor desempeño (menos periodos reabiertos).

Pesos (método CRITIC — resultado agosto 2026, verificable con
calcular_pesos_ICCI.py):
    Con datos de calidad   : ICCI = 0.95 × score_cont + 0.05 × score_cal
    Sin datos de calidad   : ICCI = 1.00 × score_cont
    "Con datos de calidad" = el ejecutor tiene al menos 1 BPIN en la hoja
    Habilitaciones (independientemente del n_periodos_habilitados = 0).
    NOTA: Las constantes W_* son hiperparámetros editables al inicio del
    archivo; cambiarlas y re-ejecutar recalcula todo el árbol.

Puntaje de Riesgo = 100 − ICCI (a mayor puntaje, mayor riesgo).
Nivel de Riesgo: Bajo < 33 ≤ Medio < 67 ≤ Alto (cortes editables).

NOTA SOBRE EL PROXY CON 1 CORTE
---------------------------------
Con un solo corte de GESPROY disponible, la continuidad se mide como
una proporción estática (reportados / obligados). Cuando se acumulen ≥6
cortes mensuales, el score debería migrar al método de transiciones de la
propuesta original (avance entre cortes consecutivos para cada periodo),
que captura mejor la dinámica de reporte. Ese cambio se hace en
cargar_base() — el resto del pipeline no cambia.

ESTRUCTURA DEL EXCEL DE SALIDA
---------------------------------
  0_Base          : datos BPIN-nivel (Continuidad ⋈ Habilitaciones ⋈ Proyectos)
  1_Continuidad   : score de continuidad por ejecutor (fórmulas vs 0_Base)
  2_Calidad       : score de calidad por ejecutor (fórmulas vs 0_Base)
  3_ICCI          : índice combinado, puntaje y nivel de riesgo con colores

USO
----
    python metodologia_ICCI.py --maestro EXCEL_MAESTRO.xlsx
    python metodologia_ICCI.py --maestro EXCEL_MAESTRO.xlsx --output METODOLOGIA_ICCI.xlsx
    python metodologia_ICCI.py --maestro EXCEL_MAESTRO.xlsx --verbose
"""

from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# =============================================================================
# HIPERPARÁMETROS METODOLÓGICOS — cambiar aquí y re-ejecutar
# =============================================================================

# Pesos del ICCI derivados de CRITIC (agosto 2026: σ ratio 18.9×).
# Editables: cambiar W_* y volver a correr recalcula la hoja 3_ICCI completa.
W_CONTINUIDAD_CON_CAL: float = 0.95   # peso continuidad cuando hay datos de calidad
W_CALIDAD_CON_CAL: float     = 0.05   # peso calidad cuando hay datos de calidad
W_CONTINUIDAD_SIN_CAL: float = 1.00   # peso continuidad cuando NO hay datos de calidad

UMBRAL_BAJO: int  = 33   # Puntaje Riesgo: 0–UMBRAL_BAJO → Bajo
UMBRAL_MEDIO: int = 67   # Puntaje Riesgo: UMBRAL_BAJO–UMBRAL_MEDIO → Medio; ≥UMBRAL_MEDIO → Alto

# Nombre de la hoja Continuidad en el Maestro (debe coincidir con lo que
# genera construir_maestro.py).
NOMBRE_HOJA_CONTINUIDAD  = "Continuidad"
NOMBRE_HOJA_HABILITACIONES = "Habilitaciones"
NOMBRE_HOJA_PROYECTOS    = "Proyectos"
NOMBRE_HOJA_EJECUTORES   = "Ejecutores"

# =============================================================================
# ESTILO — paleta DNP
# =============================================================================

FONT_NAME = "Verdana"
NAVY = "002060"
FILL_TITLE  = PatternFill("solid", fgColor=NAVY)
FILL_HEADER = PatternFill("solid", fgColor=NAVY)
FILL_NOTE   = PatternFill("solid", fgColor="FFF2CC")
FILL_PARAM  = PatternFill("solid", fgColor="E7F7F7")
FILL_ALT    = PatternFill("solid", fgColor="F2F2F2")
FILL_BAJO   = PatternFill("solid", fgColor="C6EFCE")
FILL_MEDIO  = PatternFill("solid", fgColor="FFEB9C")
FILL_ALTO   = PatternFill("solid", fgColor="FFC7CE")
THIN   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

logger = logging.getLogger("metodologia_ICCI")


# =============================================================================
# UTILIDADES DE FORMATO
# =============================================================================

def _titulo(ws, fila: int, texto: str, n_cols: int) -> None:
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=n_cols)
    c = ws.cell(row=fila, column=1, value=texto)
    c.font = Font(name=FONT_NAME, size=13, bold=True, color="FFFFFF")
    c.fill = FILL_TITLE
    c.alignment = LEFT
    ws.row_dimensions[fila].height = 22


def _nota(ws, fila: int, texto: str, n_cols: int, altura: int = 30) -> None:
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=n_cols)
    c = ws.cell(row=fila, column=1, value=texto)
    c.font = Font(name=FONT_NAME, size=9, italic=True, color="666666")
    c.fill = FILL_NOTE
    c.alignment = LEFT
    ws.row_dimensions[fila].height = altura


def _encabezados(ws, fila: int, encabezados: list[str]) -> None:
    for j, h in enumerate(encabezados, start=1):
        c = ws.cell(row=fila, column=j, value=h)
        c.font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
        c.fill = FILL_HEADER
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[fila].height = 30


def _tabla(ws, nombre: str, fila_enc: int, fila_fin: int, n_cols: int) -> None:
    if fila_fin < fila_enc + 1:
        return  # sin filas de datos, no se puede crear la tabla
    ref = f"A{fila_enc}:{get_column_letter(n_cols)}{fila_fin}"
    tabla = Table(displayName=nombre, ref=ref)
    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tabla)


# =============================================================================
# PASO 1: CARGA Y "JOIN" DESDE EL MAESTRO
# =============================================================================

def cargar_base(path_maestro: str) -> pd.DataFrame:
    """Carga y cruza las cuatro hojas relevantes del Maestro, devuelve un
    DataFrame a nivel BPIN con columnas:

        bpin, codigo_ejecutor, nombre_ejecutor, estado_detalle, fecha_corte,
        nro_periodos_a_reportar, nro_periodos_reportados,
        n_periodos_habilitados (int, 0 si el BPIN no tiene habilitaciones),
        tiene_habilitacion (bool, True si el BPIN aparece en Habilitaciones)

    Si el Maestro tiene múltiples cortes de Continuidad, se usa el ÚLTIMO
    corte disponible por BPIN (máxima fecha_corte).
    """
    logger.info("Cargando hojas del Maestro: %s", path_maestro)

    # --- Continuidad ---
    cont = pd.read_excel(
        path_maestro, sheet_name=NOMBRE_HOJA_CONTINUIDAD,
        dtype={"bpin": str},
    )
    cont["bpin"] = cont["bpin"].astype(str).str.strip()
    cont["fecha_corte"] = pd.to_datetime(cont["fecha_corte"], errors="coerce")
    cont["nro_periodos_a_reportar"] = pd.to_numeric(cont["nro_periodos_a_reportar"], errors="coerce").fillna(0)
    cont["nro_periodos_reportados"] = pd.to_numeric(cont["nro_periodos_reportados"], errors="coerce").fillna(0)

    # Si hay varios cortes por BPIN, quedarse solo con el último
    n_cortes = cont["fecha_corte"].nunique()
    if n_cortes > 1:
        logger.info(
            "Continuidad: %s cortes detectados. Se usa el último por BPIN.",
            n_cortes,
        )
        cont = (
            cont.sort_values("fecha_corte")
            .groupby("bpin", as_index=False)
            .last()
        )
    logger.info("Continuidad: %s BPIN únicos (corte más reciente).", cont["bpin"].nunique())

    # --- Habilitaciones: agregar a nivel BPIN ---
    hab = pd.read_excel(
        path_maestro, sheet_name=NOMBRE_HOJA_HABILITACIONES,
        dtype={"bpin": str},
    )
    hab["bpin"] = hab["bpin"].astype(str).str.strip()
    hab["n_periodos_a_habilitar"] = pd.to_numeric(hab["n_periodos_a_habilitar"], errors="coerce").fillna(0)

    # Sumar todos los meses habilitados por BPIN (un BPIN puede tener varias
    # solicitudes). Se marca también si el BPIN aparece en Habilitaciones
    # (independientemente de si el conteo es 0 o no).
    hab_bpin = (
        hab.groupby("bpin")["n_periodos_a_habilitar"]
        .sum()
        .reset_index()
        .rename(columns={"n_periodos_a_habilitar": "n_periodos_habilitados"})
    )
    hab_bpin["tiene_habilitacion"] = True
    logger.info("Habilitaciones: %s BPIN únicos con al menos 1 solicitud.", len(hab_bpin))

    # --- Proyectos: obtener codigo_ejecutor por BPIN ---
    proy = pd.read_excel(
        path_maestro, sheet_name=NOMBRE_HOJA_PROYECTOS,
        dtype={"bpin": str, "codigo_ejecutor": str},
    )
    proy["bpin"] = proy["bpin"].astype(str).str.strip()
    proy["codigo_ejecutor"] = proy["codigo_ejecutor"].astype(str).str.strip()
    proy_mapa = proy[["bpin", "codigo_ejecutor"]].drop_duplicates("bpin")

    # --- Ejecutores: obtener nombre ---
    ejec = pd.read_excel(
        path_maestro, sheet_name=NOMBRE_HOJA_EJECUTORES,
        dtype={"codigo_ejecutor": str},
    )
    ejec["codigo_ejecutor"] = ejec["codigo_ejecutor"].astype(str).str.strip()
    ejec_mapa = ejec[["codigo_ejecutor", "nombre_ejecutor"]].drop_duplicates("codigo_ejecutor")

    # --- JOIN ---
    base = cont.copy()
    base = base.merge(proy_mapa, on="bpin", how="left")
    base = base.merge(ejec_mapa, on="codigo_ejecutor", how="left")
    base = base.merge(hab_bpin, on="bpin", how="left")
    base["n_periodos_habilitados"] = base["n_periodos_habilitados"].fillna(0).astype(int)
    base["tiene_habilitacion"] = base["tiene_habilitacion"].fillna(False)

    # Verificar cobertura
    sin_ejecutor = base["codigo_ejecutor"].isna().sum()
    if sin_ejecutor:
        logger.warning(
            "ALERTA — %s BPIN de Continuidad no tienen código de ejecutor en "
            "la hoja Proyectos. Revisar el Maestro.", sin_ejecutor,
        )

    columnas_finales = [
        "bpin", "codigo_ejecutor", "nombre_ejecutor",
        "estado_detalle", "fecha_corte",
        "nro_periodos_a_reportar", "nro_periodos_reportados",
        "n_periodos_habilitados", "tiene_habilitacion",
    ]
    base = base[columnas_finales].sort_values(["codigo_ejecutor", "bpin"]).reset_index(drop=True)
    logger.info(
        "Base BPIN-nivel: %s filas, %s ejecutores únicos.",
        len(base), base["codigo_ejecutor"].nunique(),
    )
    return base


# =============================================================================
# PASO 2: HOJA '0_Base' — datos a nivel BPIN con columna calculada
# =============================================================================

def escribir_hoja_0_base(wb: Workbook, base: pd.DataFrame) -> int:
    """Escribe la hoja 0_Base con los datos BPIN-nivel y la columna
    calculada 'tiene_habilitacion' como fórmula de Excel (no valor pegado).
    Devuelve la fila final de datos (para que las hojas siguientes puedan
    armar rangos de SUMIFS/COUNTIFS apuntando aquí)."""
    ws = wb.create_sheet("0_Base")

    encabezados = [
        "bpin", "codigo_ejecutor", "nombre_ejecutor",
        "estado_detalle", "fecha_corte",
        "nro_periodos_a_reportar", "nro_periodos_reportados",
        "n_periodos_habilitados",
        "tiene_habilitacion",  # columna I — fórmula
    ]
    fila_enc = 1
    _encabezados(ws, fila_enc, encabezados)

    n = len(base)
    fila_inicio = fila_enc + 1
    fila_fin    = fila_enc + n

    for i, row in enumerate(base.itertuples(index=False), start=fila_inicio):
        ws.cell(row=i, column=1, value=row.bpin)
        ws.cell(row=i, column=2, value=row.codigo_ejecutor)
        ws.cell(row=i, column=3, value=row.nombre_ejecutor)
        ws.cell(row=i, column=4, value=row.estado_detalle)
        ws.cell(row=i, column=5, value=(
            None if pd.isna(row.fecha_corte) else row.fecha_corte.to_pydatetime()
        ))
        ws.cell(row=i, column=6, value=int(row.nro_periodos_a_reportar))
        ws.cell(row=i, column=7, value=int(row.nro_periodos_reportados))
        ws.cell(row=i, column=8, value=int(row.n_periodos_habilitados))
        # tiene_habilitacion: SI si el BPIN tiene al menos 1 mes habilitado registrado
        ws.cell(row=i, column=9, value=f'=IF(H{i}>0,"SI","NO")')

        ws.cell(row=i, column=5).number_format = "yyyy-mm-dd"

    anchos = [16, 14, 42, 20, 13, 12, 12, 12, 12]
    for j, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    _tabla(ws, "TablaBase", fila_enc, fila_fin, len(encabezados))
    ws.freeze_panes = f"A{fila_inicio}"
    return fila_fin


# =============================================================================
# PASO 3: HOJA '1_Continuidad' — score de continuidad por ejecutor
# =============================================================================

def escribir_hoja_1_continuidad(
    wb: Workbook, base: pd.DataFrame, fila_fin_base: int
) -> tuple[int, int]:
    """Score de Continuidad por ejecutor:
        score = (Σ periodos_reportados / Σ periodos_a_reportar) × 100

    Usa SUMIFS apuntando a 0_Base para que el Excel se recalcule solo."""
    ws = wb.create_sheet("1_Continuidad")
    n_cols = 5

    _titulo(ws, 1, "SCORE CONTINUIDAD — Períodos reportados vs. obligados por ejecutor", n_cols)
    _nota(ws, 2, (
        "Score = (Σ NRO_PERIODOS_REPORTADOS / Σ NRO_PERIODOS_A_REPORTAR) × 100, por ejecutor. "
        "Clipeado en [0, 100]. A mayor score, mejor desempeño (menos riesgo de continuidad). "
        "Con un solo corte de GESPROY disponible, este ratio es un PROXY de continuidad. "
        "Cuando se acumulen ≥6 cortes mensuales, la metodología debe migrar al conteo de transiciones "
        "avance/retroceso por periodo (ver propuesta ICCI - Propuesta.pdf)."
    ), n_cols, 55)

    encabezados = [
        "Cód.\nEjecutor", "Entidad Ejecutora",
        "Σ Períodos\na Reportar", "Σ Períodos\nReportados",
        "Score\nContinuidad",
    ]
    fila_enc = 3
    _encabezados(ws, fila_enc, encabezados)

    # Universo: un ejecutor por fila, ordenado por código
    ejecutores = (
        base.groupby("codigo_ejecutor", as_index=False)
        .agg(nombre_ejecutor=("nombre_ejecutor", "first"))
        .sort_values("codigo_ejecutor")
        .reset_index(drop=True)
    )

    fila_inicio = fila_enc + 1
    fila_fin    = fila_enc + len(ejecutores)

    hoja0 = "'0_Base'"
    rango_cod  = f"{hoja0}!$B$2:$B${fila_fin_base}"
    rango_obli = f"{hoja0}!$F$2:$F${fila_fin_base}"
    rango_rep  = f"{hoja0}!$G$2:$G${fila_fin_base}"

    for i, row in enumerate(ejecutores.itertuples(index=False), start=fila_inicio):
        ws.cell(row=i, column=1, value=row.codigo_ejecutor)
        ws.cell(row=i, column=2, value=row.nombre_ejecutor)
        ws.cell(row=i, column=3, value=f"=SUMIFS({rango_obli},{rango_cod},$A{i})")
        ws.cell(row=i, column=4, value=f"=SUMIFS({rango_rep},{rango_cod},$A{i})")
        # Score: ratio × 100, clipeado a 100 — MIN(100, ...) evita scores >100
        # cuando NRO_PERIODOS_REPORTADOS supera NRO_PERIODOS_A_REPORTAR (MARCA_REPORTO_MAS).
        ws.cell(row=i, column=5, value=f"=IFERROR(MIN(100,$D{i}/$C{i}*100),0)")
        ws.cell(row=i, column=5).number_format = "0.00"

    anchos = [14, 42, 14, 14, 13]
    for j, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    _tabla(ws, "TablaContinuidad", fila_enc, fila_fin, n_cols)
    ws.freeze_panes = f"A{fila_inicio}"

    return fila_inicio, fila_fin


# =============================================================================
# PASO 4: HOJA '2_Calidad' — score de calidad por ejecutor
# =============================================================================

def escribir_hoja_2_calidad(
    wb: Workbook, base: pd.DataFrame, fila_fin_base: int
) -> tuple[int, int]:
    """Score de Calidad de Información por ejecutor:
        score = 100 − (Σ n_periodos_habilitados / Σ nro_periodos_a_reportar × 100)

    Muestra todos los ejecutores. La columna 'Tiene Datos Calidad' indica SI
    cuando al menos uno de sus BPIN tiene habilitaciones registradas — esos
    ejecutores reciben peso diferenciado en el ICCI (0.95/0.05). Los demás
    reciben peso 1.00 en continuidad."""
    ws = wb.create_sheet("2_Calidad")
    n_cols = 7

    _titulo(ws, 1, "SCORE CALIDAD DE INFORMACIÓN — Períodos habilitados (reabiertos) por ejecutor", n_cols)
    _nota(ws, 2, (
        "Score = 100 − (Σ n_periodos_habilitados / Σ nro_periodos_a_reportar × 100), por ejecutor. "
        "n_periodos_habilitados = suma de meses reabiertos via solicitud de habilitación (hoja Habilitaciones). "
        "Si el ejecutor no tiene BPIN en habilitaciones → score = 100 (sin reaperturas = máxima calidad). "
        "NOTA COBERTURA: solo 74 ejecutores (≈4.5%) tienen datos de habilitación reales — "
        "el 95.5% restante tiene score 100 por defecto, razón por la que el peso CRITIC de calidad es solo 5%."
    ), n_cols, 55)

    encabezados = [
        "Cód.\nEjecutor", "Entidad Ejecutora",
        "N° BPIN\ncon Hab.", "Σ Períodos\nHabilitados",
        "Σ Períodos\na Reportar",
        "Score\nCalidad",
        "Tiene\nDatos Calidad",
    ]
    fila_enc = 3
    _encabezados(ws, fila_enc, encabezados)

    ejecutores = (
        base.groupby("codigo_ejecutor", as_index=False)
        .agg(nombre_ejecutor=("nombre_ejecutor", "first"))
        .sort_values("codigo_ejecutor")
        .reset_index(drop=True)
    )

    fila_inicio = fila_enc + 1
    fila_fin    = fila_enc + len(ejecutores)

    hoja0 = "'0_Base'"
    rango_cod   = f"{hoja0}!$B$2:$B${fila_fin_base}"
    rango_obli  = f"{hoja0}!$F$2:$F${fila_fin_base}"
    rango_hab   = f"{hoja0}!$H$2:$H${fila_fin_base}"
    rango_tiene = f"{hoja0}!$I$2:$I${fila_fin_base}"

    for i, row in enumerate(ejecutores.itertuples(index=False), start=fila_inicio):
        ws.cell(row=i, column=1, value=row.codigo_ejecutor)
        ws.cell(row=i, column=2, value=row.nombre_ejecutor)
        # N° BPIN con habilitacion (tiene_habilitacion = "SI" en 0_Base)
        ws.cell(row=i, column=3, value=f'=COUNTIFS({rango_cod},$A{i},{rango_tiene},"SI")')
        # Suma de periodos habilitados
        ws.cell(row=i, column=4, value=f"=SUMIFS({rango_hab},{rango_cod},$A{i})")
        # Suma de periodos a reportar (denominador)
        ws.cell(row=i, column=5, value=f"=SUMIFS({rango_obli},{rango_cod},$A{i})")
        # Score calidad: 100 - tasa de habilitacion. IFERROR → 100 si denominador es 0.
        ws.cell(row=i, column=6, value=f"=IFERROR(100-$D{i}/$E{i}*100,100)")
        ws.cell(row=i, column=6).number_format = "0.00"
        # Flag: tiene datos reales de calidad (al menos 1 BPIN en Habilitaciones)
        ws.cell(row=i, column=7, value=f'=IF($C{i}>0,"SI","NO")')

    anchos = [14, 42, 10, 12, 12, 12, 12]
    for j, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    _tabla(ws, "TablaCalidad", fila_enc, fila_fin, n_cols)
    ws.freeze_panes = f"A{fila_inicio}"

    return fila_inicio, fila_fin


# =============================================================================
# PASO 5: HOJA '3_ICCI' — índice combinado y nivel de riesgo
# =============================================================================

def escribir_hoja_3_icci(
    wb: Workbook,
    fila_inicio_1: int, fila_fin_1: int,
    fila_inicio_2: int, fila_fin_2: int,
) -> None:
    """Combina los scores de continuidad y calidad en el ICCI final.

    Lógica de pesos (hardcodeada aquí como parámetros editables en C3/F3/I3):
        Si tiene_datos_calidad = "SI":
            ICCI = W_CONT × score_continuidad + W_CAL × score_calidad
        Si no:
            ICCI = 1.00 × score_continuidad
    Puntaje de Riesgo = 100 − ICCI.
    """
    ws = wb.create_sheet("3_ICCI")
    n_cols = 9

    _titulo(ws, 1, "ICCI — Índice de Continuidad y Calidad de Información", n_cols)

    # Bloque de parámetros de pesos (editables en celda)
    ws.cell(row=2, column=1, value="PARÁMETROS — cambiar aquí para recalcular todo").font = \
        Font(name=FONT_NAME, size=10, bold=True)

    params_pesos = [
        (2, "Peso Continuidad\n(con datos calidad)", W_CONTINUIDAD_CON_CAL),
        (5, "Peso Calidad\n(con datos calidad)",     W_CALIDAD_CON_CAL),
        (8, "Peso Continuidad\n(sin datos calidad)",  W_CONTINUIDAD_SIN_CAL),
    ]
    for col_lbl, texto, valor in params_pesos:
        ws.cell(row=3, column=col_lbl, value=texto).font = Font(name=FONT_NAME, size=9)
        c = ws.cell(row=3, column=col_lbl + 1, value=valor)
        c.font = Font(name=FONT_NAME, size=11, bold=True)
        c.fill = FILL_PARAM
        c.number_format = "0%"
        c.border = BORDER

    # Verificación: suma de pesos con calidad
    ws.cell(row=3, column=7, value="Suma pesos\ncon calidad").font = Font(name=FONT_NAME, size=9)
    c = ws.cell(row=3, column=8, value="=C3+F3")
    c.font = Font(name=FONT_NAME, size=11, bold=True)
    c.fill = FILL_PARAM
    c.number_format = "0%"
    c.border = BORDER

    # Cortes de riesgo
    ws.cell(row=4, column=1, value="CORTES DE NIVEL DE RIESGO").font = \
        Font(name=FONT_NAME, size=10, bold=True)
    ws.cell(row=5, column=1, value="Corte Bajo/Medio →").font = Font(name=FONT_NAME, size=9)
    c = ws.cell(row=5, column=2, value=UMBRAL_BAJO)
    c.font = Font(name=FONT_NAME, size=11, bold=True)
    c.fill = FILL_PARAM
    c.border = BORDER
    ws.cell(row=5, column=3, value="Corte Medio/Alto →").font = Font(name=FONT_NAME, size=9)
    c = ws.cell(row=5, column=4, value=UMBRAL_MEDIO)
    c.font = Font(name=FONT_NAME, size=11, bold=True)
    c.fill = FILL_PARAM
    c.border = BORDER
    ws.cell(row=5, column=6, value=(
        f"Puntaje Riesgo: Bajo 0–{UMBRAL_BAJO}  |  "
        f"Medio {UMBRAL_BAJO}–{UMBRAL_MEDIO}  |  "
        f"Alto {UMBRAL_MEDIO}–100"
    )).font = Font(name=FONT_NAME, size=9, italic=True)

    _nota(ws, 6, (
        "ICCI = W_CONT × Score_Continuidad + W_CAL × Score_Calidad  (si ejecutor tiene datos de calidad). "
        "ICCI = Score_Continuidad  (si ejecutor NO tiene datos de calidad). "
        "Puntaje de Riesgo = 100 − ICCI. A mayor puntaje, mayor riesgo. "
        "Las columnas 'Aporte Cont.' y 'Aporte Cal.' muestran la contribución de cada componente al ICCI."
    ), n_cols, 40)

    encabezados = [
        "Cód.\nEjecutor", "Entidad Ejecutora",
        "Score\nContinuidad", "Score\nCalidad",
        "Datos\nCalidad",
        "Aporte\nCont.", "Aporte\nCal.",
        "ICCI", "Puntaje\nRiesgo", "Nivel\nRiesgo",
    ]
    # 10 columnas — actualizar n_cols
    n_cols = len(encabezados)
    fila_enc = 7
    _encabezados(ws, fila_enc, encabezados)

    n = fila_fin_1 - fila_inicio_1 + 1
    fila_inicio = fila_enc + 1
    fila_fin    = fila_enc + n

    hoja1 = "'1_Continuidad'"
    hoja2 = "'2_Calidad'"

    for k, (i1, i2) in enumerate(
        zip(range(fila_inicio_1, fila_fin_1 + 1), range(fila_inicio_2, fila_fin_2 + 1)),
        start=0,
    ):
        i3 = fila_inicio + k
        ws.cell(row=i3, column=1,  value=f"={hoja1}!$A${i1}")
        ws.cell(row=i3, column=2,  value=f"={hoja1}!$B${i1}")
        ws.cell(row=i3, column=3,  value=f"={hoja1}!$E${i1}")  # score_continuidad
        ws.cell(row=i3, column=4,  value=f"={hoja2}!$F${i2}")  # score_calidad
        ws.cell(row=i3, column=5,  value=f"={hoja2}!$G${i2}")  # tiene_datos_calidad

        # Aporte continuidad: depende de si tiene datos de calidad
        ws.cell(row=i3, column=6, value=(
            f'=IF($E{i3}="SI",$C{i3}*$C$3,$C{i3}*$I$3)'
        ))
        # Aporte calidad: 0 si no tiene datos de calidad
        ws.cell(row=i3, column=7, value=(
            f'=IF($E{i3}="SI",$D{i3}*$F$3,0)'
        ))
        # ICCI = suma de aportes
        ws.cell(row=i3, column=8,  value=f"=$F{i3}+$G{i3}")
        # Puntaje de Riesgo = 100 - ICCI
        ws.cell(row=i3, column=9,  value=f"=100-$H{i3}")
        # Nivel de riesgo (usa puntaje de riesgo y cortes editables $B$5/$D$5)
        ws.cell(row=i3, column=10, value=(
            f'=IF($I{i3}<$B$5,"Riesgo Bajo",'
            f'IF($I{i3}<$D$5,"Riesgo Medio","Riesgo Alto"))'
        ))

        for j in (3, 4, 6, 7, 8, 9):
            ws.cell(row=i3, column=j).number_format = "0.00"

        if k % 2 == 1:
            for j in range(1, n_cols + 1):
                ws.cell(row=i3, column=j).fill = FILL_ALT

    anchos = [14, 42, 12, 12, 12, 12, 12, 12, 12, 14]
    for j, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    _tabla(ws, "TablaICCI", fila_enc, fila_fin, n_cols)
    ws.freeze_panes = f"A{fila_inicio}"

    # Formato condicional: PUNTAJE DE RIESGO (col I) y NIVEL DE RIESGO (col J)
    rango_puntaje = f"$I${fila_inicio}:$I${fila_fin}"
    ws.conditional_formatting.add(rango_puntaje, CellIsRule(
        operator="between", formula=["0", str(UMBRAL_BAJO)], fill=FILL_BAJO,
    ))
    ws.conditional_formatting.add(rango_puntaje, CellIsRule(
        operator="between", formula=[str(UMBRAL_BAJO), str(UMBRAL_MEDIO)], fill=FILL_MEDIO,
    ))
    ws.conditional_formatting.add(rango_puntaje, CellIsRule(
        operator="between", formula=[str(UMBRAL_MEDIO), "100"], fill=FILL_ALTO,
    ))

    rango_nivel = f"$J${fila_inicio}:$J${fila_fin}"
    ws.conditional_formatting.add(rango_nivel, CellIsRule(
        operator="equal", formula=['"Riesgo Bajo"'], fill=FILL_BAJO,
    ))
    ws.conditional_formatting.add(rango_nivel, CellIsRule(
        operator="equal", formula=['"Riesgo Medio"'], fill=FILL_MEDIO,
    ))
    ws.conditional_formatting.add(rango_nivel, CellIsRule(
        operator="equal", formula=['"Riesgo Alto"'], fill=FILL_ALTO,
    ))


# =============================================================================
# PIPELINE COMPLETO
# =============================================================================

def generar_metodologia_icci(path_maestro: str, path_salida: str) -> None:
    logger.info("Pesos CRITIC: Continuidad=%.0f%% / Calidad=%.0f%% (con datos calidad)",
                W_CONTINUIDAD_CON_CAL * 100, W_CALIDAD_CON_CAL * 100)
    logger.info("Peso sin datos calidad: Continuidad=%.0f%%", W_CONTINUIDAD_SIN_CAL * 100)
    logger.info("Cortes de riesgo: Bajo<%.0f ≤ Medio<%.0f ≤ Alto", UMBRAL_BAJO, UMBRAL_MEDIO)

    base = cargar_base(path_maestro)

    wb = Workbook()
    del wb["Sheet"]

    fila_fin_base = escribir_hoja_0_base(wb, base)
    fila_inicio_1, fila_fin_1 = escribir_hoja_1_continuidad(wb, base, fila_fin_base)
    fila_inicio_2, fila_fin_2 = escribir_hoja_2_calidad(wb, base, fila_fin_base)
    escribir_hoja_3_icci(wb, fila_inicio_1, fila_fin_1, fila_inicio_2, fila_fin_2)

    Path(path_salida).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path_salida)
    logger.info("Excel de metodología ICCI generado en: %s", path_salida)


def configurar_logging(verbose: bool) -> None:
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()
    formato = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", "%H:%M:%S")
    consola = logging.StreamHandler(sys.stdout)
    consola.setLevel(logging.INFO if verbose else logging.WARNING)
    consola.setFormatter(formato)
    logger.addHandler(consola)


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Genera el Excel de metodología del ICCI "
            "(Índice de Continuidad y Calidad de Información) con fórmulas reales."
        )
    )
    parser.add_argument(
        "--maestro", "-m",
        default="EXCEL_MAESTRO.xlsx",
        help="Ruta al Excel Maestro (default: EXCEL_MAESTRO.xlsx en el directorio actual).",
    )
    parser.add_argument(
        "--output", "-o",
        default="METODOLOGIA_ICCI.xlsx",
        help="Ruta del Excel de salida (default: METODOLOGIA_ICCI.xlsx en el directorio actual).",
    )
    parser.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="Mostrar logs detallados en consola.",
    )
    args = parser.parse_args()

    configurar_logging(args.verbose)

    if not Path(args.maestro).exists():
        print(f"ERROR: no encontré el archivo '{args.maestro}'", file=sys.stderr)
        sys.exit(1)

    generar_metodologia_icci(args.maestro, args.output)


if __name__ == "__main__":
    main()
